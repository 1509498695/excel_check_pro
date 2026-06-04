"""执行接口测试。"""

from __future__ import annotations

import subprocess
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import httpx
import pandas as pd
import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from backend.app.api import source_api
from backend.app.api.schemas import DataSource, VariableTag
from backend.app.auth.service import create_access_token, hash_password
from backend.app.database import async_session_factory
from backend.app.integrations import feishu_bot, feishu_client
from backend.app.integrations.feishu_client import (
    FEISHU_APP_PERMISSION_MISSING,
    FEISHU_API_ERROR,
    FEISHU_DOCUMENT_NOT_FOUND,
    FEISHU_DOCUMENT_PERMISSION_DENIED,
    FEISHU_INVALID_URL,
    FeishuClientError,
    FeishuSheetMetadata,
    FeishuSheetTable,
)
from backend.app.loaders import local_reader as local_reader_module
from backend.app.loaders.local_reader import load_local_variables
from backend.app.models import FeishuBotConfigRecord, User, UserProjectRole
from backend.app.security.crypto import encrypt_secret
from backend.run import app


TEST_DATA_PATH = Path(__file__).resolve().parent / "data" / "minimal_rules.xlsx"


@pytest.fixture(autouse=True)
def _allow_test_local_files(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """执行接口大量复用本地测试 Excel，测试环境显式放入 allowlist。"""
    monkeypatch.setattr(
        local_reader_module,
        "settings",
        SimpleNamespace(
            local_file_root_allowlist=(TEST_DATA_PATH.parent, tmp_path),
            runtime_upload_dir=tmp_path / "uploads",
            svn_cache_dir=tmp_path / "svn-cache",
        ),
    )


async def _seed_feishu_bot_config(project_id: int) -> None:
    """写入项目级飞书机器人配置，供 metadata 接口测试复用。"""
    feishu_bot._TOKEN_CACHE.clear()
    feishu_bot._TOKEN_LOCKS.clear()
    async with async_session_factory() as session:
        session.add(
            FeishuBotConfigRecord(
                project_id=project_id,
                app_id="metadata_app",
                app_secret_cipher=encrypt_secret("metadata_secret"),
                default_chat_id="",
            )
        )
        await session.commit()


def _install_feishu_metadata_mock(
    monkeypatch: pytest.MonkeyPatch,
    handler,
) -> None:
    transport = httpx.MockTransport(handler)

    def _factory(timeout: float = 10.0) -> httpx.AsyncClient:
        return httpx.AsyncClient(
            transport=transport,
            base_url=feishu_bot.FEISHU_OPEN_BASE_URL,
            timeout=timeout,
        )

    monkeypatch.setattr(feishu_bot, "_create_async_client", _factory)
    monkeypatch.setattr(feishu_client, "_create_async_client", _factory)


def _feishu_token_response() -> httpx.Response:
    return httpx.Response(
        200,
        json={
            "code": 0,
            "msg": "ok",
            "tenant_access_token": "t_metadata",
            "expire": 7200,
        },
    )


def _feishu_sheets_response() -> httpx.Response:
    return httpx.Response(
        200,
        json={
            "code": 0,
            "msg": "success",
            "data": {
                "sheets": [
                    {
                        "sheet_id": "gid001",
                        "title": "Sheet1",
                        "index": 0,
                        "hidden": False,
                        "resource_type": "sheet",
                        "grid_properties": {"row_count": 2, "column_count": 2},
                    },
                    {
                        "sheet_id": "gid002",
                        "title": "Sheet2",
                        "index": 1,
                        "hidden": False,
                        "resource_type": "sheet",
                        "grid_properties": {"row_count": 2, "column_count": 1},
                    },
                ]
            },
        },
    )


def _feishu_preview_sheets_response() -> httpx.Response:
    return httpx.Response(
        200,
        json={
            "code": 0,
            "msg": "success",
            "data": {
                "sheets": [
                    {
                        "sheet_id": "gid_preview",
                        "title": "Preview",
                        "index": 0,
                        "hidden": False,
                        "resource_type": "sheet",
                        "grid_properties": {"row_count": 6, "column_count": 3},
                    }
                ]
            },
        },
    )


def _feishu_preview_values_response(values: list[list[Any]]) -> httpx.Response:
    return httpx.Response(
        200,
        json={
            "code": 0,
            "msg": "success",
            "data": {
                "spreadsheetToken": "shtcnabc123",
                "valueRange": {
                    "range": "gid_preview!A1:C6",
                    "values": values,
                },
            },
        },
    )


def _create_composite_test_workbook(target_path: Path) -> Path:
    """创建组合变量测试所需的最小 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": [1, 2, 3],
                "ItemName": ["Gold", "Wood", "Stone"],
                "Desc": ["黄金+100", "木头+200", "石头+300"],
            }
        ).to_excel(writer, sheet_name="items", index=False)
        pd.DataFrame({"RefID": [1, 2, 9]}).to_excel(
            writer,
            sheet_name="drops",
            index=False,
        )

    return target_path


def _create_dual_composite_test_workbook(target_path: Path) -> Path:
    """创建双组合变量关联比对测试所需的 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [10001, 10002],
                "INT_ConditionType": [4, 3],
                "INT_RequireRule": [1, 1],
            }
        ).to_excel(writer, sheet_name="left_items", index=False)
        pd.DataFrame(
            {
                "INT_ID": [10001, 10002],
                "INT_ConditionType": [5, 3],
                "INT_RequireRule": [1, 1],
            }
        ).to_excel(writer, sheet_name="right_items", index=False)

    return target_path


def _create_mapping_exclusion_test_workbook(target_path: Path) -> Path:
    """创建多组映射排除范围判定值测试所需的 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1, 2, 3],
                "INT_Group": [1, 0, 1],
                "INT_Faction": [0, 0, 0],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    return target_path


def _create_paginated_test_workbook(target_path: Path) -> Path:
    """创建用于执行结果分页测试的 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame({"ID": list(range(1, 11))}).to_excel(
            writer,
            sheet_name="items",
            index=False,
        )
        pd.DataFrame({"RefID": list(range(1, 46))}).to_excel(
            writer,
            sheet_name="drops",
            index=False,
        )

    return target_path


def _create_whitespace_header_workbook(target_path: Path) -> Path:
    """创建带尾部空格 Sheet/列名的 Excel，用于回归原始标识读取链路。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1001, 1002],
                "STR_ABSwitch  ": ["on", "off"],
                "DESC3": ["开", "关"],
            }
        ).to_excel(writer, sheet_name="Quest  ", index=False)

    return target_path


@pytest.mark.anyio
async def test_execute_engine_returns_three_rule_results() -> None:
    """验证一次请求能同时覆盖空值、重复值和跨表映射缺失。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "Execution Completed"
    assert payload["meta"]["total_rows_scanned"] > 0
    assert payload["meta"]["failed_sources"] == []

    abnormal_results = payload["data"]["abnormal_results"]
    assert isinstance(abnormal_results, list)
    assert any(
        item["rule_name"] == "not_null" and item["level"] == "error"
        for item in abnormal_results
    )
    assert any(
        item["rule_name"] == "unique" and item["level"] == "warning"
        for item in abnormal_results
    )
    assert any(
        item["rule_name"] == "cross_table_mapping" and item["level"] == "error"
        for item in abnormal_results
    )

    for item in abnormal_results:
        assert set(item) == {
            "level",
            "rule_name",
            "location",
            "row_index",
            "raw_value",
            "message",
        }


@pytest.mark.anyio
async def test_execute_engine_loads_feishu_variables(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """执行流水线应能加载飞书变量并参与规则执行。"""
    _install_feishu_runtime_stubs(monkeypatch)

    response = await auth_client.post(
        "/api/v1/engine/execute",
        json={
            "sources": [
                {
                    "id": "src_feishu",
                    "type": "feishu",
                    "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
                }
            ],
            "variables": [
                {
                    "tag": "[items-name]",
                    "source_id": "src_feishu",
                    "sheet": "Items",
                    "column": "Name",
                }
            ],
            "rules": [
                {
                    "rule_type": "not_null",
                    "params": {"target_tags": ["[items-name]"]},
                }
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["meta"]["failed_sources"] == []
    assert payload["meta"]["total_rows_scanned"] == 3
    abnormal_results = payload["data"]["abnormal_results"]
    assert any(
        item["rule_name"] == "not_null"
        and item["row_index"] == 3
        and item["location"] == "[items-name] -> Name"
        for item in abnormal_results
    )


@pytest.mark.anyio
async def test_execute_engine_marks_failed_feishu_source(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """飞书读取失败时只标记 failed_sources，不让整个任务崩溃。"""

    async def _explode(*_args, **_kwargs):
        raise FeishuClientError(FEISHU_API_ERROR, "飞书 API 临时不可用")

    monkeypatch.setattr(feishu_client, "list_spreadsheet_sheets", _explode)

    response = await auth_client.post(
        "/api/v1/engine/execute",
        json={
            "sources": [
                {
                    "id": "src_feishu",
                    "type": "feishu",
                    "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
                }
            ],
            "variables": [
                {
                    "tag": "[items-name]",
                    "source_id": "src_feishu",
                    "sheet": "Items",
                    "column": "Name",
                }
            ],
            "rules": [
                {
                    "rule_type": "not_null",
                    "params": {"target_tags": ["[items-name]"]},
                }
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["meta"]["failed_sources"] == ["src_feishu"]
    assert payload["data"]["abnormal_results"] == []


def _install_feishu_runtime_stubs(monkeypatch: pytest.MonkeyPatch) -> None:
    async def _list_sheets(*_args, **_kwargs):
        return [
            FeishuSheetMetadata(
                sheet_id="gid_items",
                title="Items",
                index=0,
                row_count=4,
                column_count=3,
                hidden=False,
                resource_type="sheet",
            )
        ]

    async def _read_values(*_args, **_kwargs):
        return FeishuSheetTable(
            spreadsheet_token="shtcnabc123",
            sheet_id="gid_items",
            sheet_title="Items",
            range="gid_items!A1:C4",
            columns=["ID", "Name", "Group"],
            rows=[],
            raw_values=[
                ["ID", "Name", "Group"],
                [1, "Alpha", "A"],
                [2, "", "B"],
                [3, "Gamma", "C"],
            ],
        )

    monkeypatch.setattr(feishu_client, "list_spreadsheet_sheets", _list_sheets)
    monkeypatch.setattr(feishu_client, "read_sheet_values", _read_values)


@pytest.mark.anyio
async def test_execute_engine_supports_expected_value_set_for_eq_ne(
    tmp_path: Path,
) -> None:
    """验证个人校验执行器的等于/不等于可使用英文逗号规则集。"""
    workbook_path = tmp_path / "expected_value_set.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame({"ID": [0, 1, 2, 3]}).to_excel(
            writer,
            sheet_name="items",
            index=False,
        )

    payload = {
        "sources": [
            {"id": "src_set", "type": "local_excel", "path": str(workbook_path)}
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_set",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [
            {
                "rule_type": "fixed_value_compare",
                "params": {
                    "target_tag": "[items-id]",
                    "operator": "eq",
                    "expected_value": "0, 1,2",
                    "expected_value_mode": "set",
                    "rule_name": "ID 必须属于 0/1/2",
                    "location": "items -> ID",
                },
            },
            {
                "rule_type": "fixed_value_compare",
                "params": {
                    "target_tag": "[items-id]",
                    "operator": "ne",
                    "expected_value": "2,4",
                    "expected_value_mode": "set",
                    "rule_name": "ID 不应属于 2/4",
                    "location": "items -> ID",
                },
            },
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert [item["rule_name"] for item in abnormal_results] == [
        "ID 必须属于 0/1/2",
        "ID 不应属于 2/4",
    ]
    assert abnormal_results[0]["raw_value"] == 3
    assert abnormal_results[1]["raw_value"] == 2
    assert "规则集中的任一值" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_supports_expected_value_set_in_composite_conditions(
    tmp_path: Path,
) -> None:
    """验证组合变量筛选和断言中的固定值规则集也会生效。"""
    workbook_path = tmp_path / "composite_expected_value_set.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1001, 1002, 1003, 1004],
                "INT_Faction": [0, 1, 2, 3],
                "INT_Group": ["A", "B", "C", "D"],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [
            {
                "id": "src_composite_set",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-composite]",
                "source_id": "src_composite_set",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_Faction", "INT_Group"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_type": "composite_condition_check",
                "params": {
                    "target_tag": "[items-composite]",
                    "rule_name": "组合规则集校验",
                    "composite_config": {
                        "global_filters": [
                            {
                                "condition_id": "filter-faction-set",
                                "field": "INT_Faction",
                                "operator": "eq",
                                "value_source": "literal",
                                "expected_value": "0, 1",
                                "expected_value_mode": "set",
                            }
                        ],
                        "branches": [
                            {
                                "branch_id": "branch-group-set",
                                "filters": [],
                                "assertions": [
                                    {
                                        "condition_id": "assert-group-set",
                                        "field": "INT_Group",
                                        "operator": "eq",
                                        "value_source": "literal",
                                        "expected_value": "A,C",
                                        "expected_value_mode": "set",
                                    }
                                ],
                            }
                        ],
                    },
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["raw_value"] == "B"
    assert "规则集中的任一值" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_composite_branch_not_null_reports_excel_empty_cells(
    tmp_path: Path,
) -> None:
    """验证组合分支非空断言能报告 Excel 空单元格。"""
    workbook_path = tmp_path / "composite_not_null_empty_cells.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1, 2, 3, 4, 5],
                "STR_ABSwitch": [None, None, None, "GreenServer:0", "   "],
                "DESC": ["设置名字", "给自己设置头像", "城墙升级到3级", "正常值", "空白字符串"],
            }
        ).to_excel(writer, sheet_name="Quest", index=False)

    payload = {
        "sources": [
            {
                "id": "quests",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[quests-Quest-composite]",
                "source_id": "quests",
                "sheet": "Quest",
                "variable_kind": "composite",
                "columns": ["INT_ID", "STR_ABSwitch", "DESC"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_type": "composite_condition_check",
                "params": {
                    "target_tag": "[quests-Quest-composite]",
                    "rule_name": "ABSwitch 非空校验",
                    "composite_config": {
                        "global_filters": [],
                        "branches": [
                            {
                                "branch_id": "branch-not-null",
                                "filters": [],
                                "assertions": [
                                    {
                                        "condition_id": "assert-abswitch-not-null",
                                        "field": "STR_ABSwitch",
                                        "operator": "not_null",
                                    }
                                ],
                            }
                        ],
                    },
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert [item["row_index"] for item in abnormal_results] == [2, 3, 4, 6]
    assert {item["location"] for item in abnormal_results} == {"Quest -> STR_ABSwitch"}
    assert all("不能为空" in item["message"] for item in abnormal_results)


@pytest.mark.anyio
async def test_execute_engine_composite_branch_not_null_reports_empty_key_column(
    tmp_path: Path,
) -> None:
    """验证组合变量 Key 列为空时也可被组合分支非空断言报告。"""
    workbook_path = tmp_path / "composite_not_null_empty_key_column.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "STR_ABSwitch": [None, None, "GreenServer:0", "   "],
                "DESC": ["设置名字", "给自己设置头像", "正常值", "空白字符串"],
            }
        ).to_excel(writer, sheet_name="Quest", index=False)

    payload = {
        "sources": [
            {
                "id": "quests",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[quests-Quest-composite]",
                "source_id": "quests",
                "sheet": "Quest",
                "variable_kind": "composite",
                "columns": ["STR_ABSwitch", "DESC"],
                "key_column": "STR_ABSwitch",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_type": "composite_condition_check",
                "params": {
                    "target_tag": "[quests-Quest-composite]",
                    "rule_name": "ABSwitch Key 非空校验",
                    "composite_config": {
                        "global_filters": [],
                        "branches": [
                            {
                                "branch_id": "branch-not-null",
                                "filters": [],
                                "assertions": [
                                    {
                                        "condition_id": "assert-abswitch-key-not-null",
                                        "field": "STR_ABSwitch",
                                        "operator": "not_null",
                                    }
                                ],
                            }
                        ],
                    },
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert [item["row_index"] for item in abnormal_results] == [2, 3, 5]
    assert {item["location"] for item in abnormal_results} == {"Quest -> STR_ABSwitch"}
    assert all("不能为空" in item["message"] for item in abnormal_results)


@pytest.mark.anyio
async def test_execute_engine_multi_composite_mapping_exclusion_requires_expected_value_hit(
    tmp_path: Path,
) -> None:
    """验证个人校验多组映射排除范围需同时命中行号和判定值。"""
    workbook_path = _create_mapping_exclusion_test_workbook(
        tmp_path / "mapping_exclusion_expected_value.xlsx"
    )
    payload = {
        "sources": [
            {
                "id": "src_mapping",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-mapping]",
                "source_id": "src_mapping",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_Group", "INT_Faction"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_type": "multi_composite_mapping_check",
                "params": {
                    "target_tag": "[items-mapping]",
                    "rule_name": "个人映射排除判定值",
                    "mapping_config": {
                        "nodes": [
                            {
                                "node_id": "mapping-node-1",
                                "variable_tag": "[items-mapping]",
                                "filters": [
                                    {
                                        "condition_id": "filter-group",
                                        "field": "INT_Group",
                                        "operator": "eq",
                                        "value_source": "literal",
                                        "expected_value": "1",
                                        "exclusion_ranges": [
                                            {
                                                "range_id": "row-3-miss",
                                                "start_row": 3,
                                                "end_row": 3,
                                                "expected_value": "9",
                                            }
                                        ],
                                    }
                                ],
                            }
                        ]
                    },
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["row_index"] == 3
    assert "筛选失败排除行号范围或判定值" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_filters_rules_by_selected_ids() -> None:
    """验证 selected_rule_ids 只会执行被勾选的规则。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-not-null",
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_id": "rule-unique",
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_id": "rule-cross",
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
        "selected_rule_ids": ["rule-cross"],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert abnormal_results
    assert all(item["rule_name"] == "cross_table_mapping" for item in abnormal_results)


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_unsupported_rule() -> None:
    """验证未注册规则类型会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [{"rule_type": "missing_rule", "params": {}}],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "Unsupported rule_type" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_invalid_rule_params() -> None:
    """验证非法规则参数会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": "not-a-list"},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "params.target_tags" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_unknown_source_id() -> None:
    """验证变量引用不存在的 source_id 会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[bad-source]",
                "source_id": "src_missing",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "unknown source_id" in response.json()["detail"]


@pytest.mark.anyio
async def test_local_pick_returns_real_selected_path(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证本地文件选择接口会返回真实路径，不会复制文件。"""

    monkeypatch.setattr(
        source_api,
        "settings",
        SimpleNamespace(enable_local_picker=True),
    )
    monkeypatch.setattr(
        source_api,
        "_show_local_file_dialog",
        lambda source_type: str(TEST_DATA_PATH) if source_type == "local_excel" else "",
    )

    response = await auth_client.post(
        "/api/v1/sources/local-pick",
        json={"source_type": "local_excel"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "ok"
    assert payload["data"]["source_type"] == "local_excel"
    assert payload["data"]["selected_path"] == str(TEST_DATA_PATH.resolve())


@pytest.mark.anyio
async def test_local_pick_returns_cancelled_when_user_closes_dialog(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证用户取消选择时接口会返回 cancelled。"""

    monkeypatch.setattr(
        source_api,
        "settings",
        SimpleNamespace(enable_local_picker=True),
    )
    monkeypatch.setattr(source_api, "_show_local_file_dialog", lambda _source_type: "")

    response = await auth_client.post(
        "/api/v1/sources/local-pick",
        json={"source_type": "local_excel"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 204
    assert payload["msg"] == "cancelled"
    assert payload["data"]["selected_path"] == ""


@pytest.mark.anyio
async def test_local_directory_validate_returns_normalized_absolute_directory(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证本地目录校验接口会返回规范化后的绝对目录路径。"""

    response = await auth_client.post(
        "/api/v1/sources/local-directory-validate",
        json={"directory_path": f" {tmp_path} "},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "ok"
    assert payload["data"]["directory_path"] == str(tmp_path.resolve())


def test_show_local_file_dialog_returns_subprocess_stdout(monkeypatch: pytest.MonkeyPatch) -> None:
    """验证子进程方案下，正常返回时取真实路径并去除两端空白。"""

    captured_args: dict[str, Any] = {}

    def fake_run(args: list[str], **kwargs: Any) -> SimpleNamespace:
        captured_args["args"] = args
        captured_args["kwargs"] = kwargs
        return SimpleNamespace(returncode=0, stdout=" C:/tmp/example.xlsx \n", stderr="")

    monkeypatch.setattr(source_api.subprocess, "run", fake_run)

    result = source_api._show_local_file_dialog("local_excel")

    assert result == "C:/tmp/example.xlsx"
    assert captured_args["args"][0] == source_api.sys.executable
    assert captured_args["kwargs"]["timeout"] == source_api._PICKER_SUBPROCESS_TIMEOUT_SECONDS
    assert captured_args["kwargs"]["capture_output"] is True


def test_show_local_file_dialog_returns_empty_when_user_cancelled(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证用户在子进程弹窗中取消时，函数返回空串。"""

    monkeypatch.setattr(
        source_api.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(returncode=0, stdout="", stderr=""),
    )

    assert source_api._show_local_file_dialog("local_excel") == ""


def test_show_local_file_dialog_raises_runtime_error_on_subprocess_timeout(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证子进程超时会被转换为 RuntimeError，由路由层映射成 500。"""

    def raise_timeout(*_args: Any, **_kwargs: Any) -> SimpleNamespace:
        raise subprocess.TimeoutExpired(cmd="python", timeout=1)

    monkeypatch.setattr(source_api.subprocess, "run", raise_timeout)

    with pytest.raises(RuntimeError, match="超时"):
        source_api._show_local_file_dialog("local_excel")


def test_show_local_file_dialog_raises_runtime_error_on_nonzero_exit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证子进程异常退出时，会带上 stderr 抛出 RuntimeError。"""

    monkeypatch.setattr(
        source_api.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(
            returncode=1,
            stdout="",
            stderr="tk init failed",
        ),
    )

    with pytest.raises(RuntimeError, match="tk init failed"):
        source_api._show_local_file_dialog("local_excel")


@pytest.mark.anyio
async def test_source_metadata_returns_excel_sheet_and_column_structure(
    auth_client: AsyncClient,
) -> None:
    """验证变量池元数据接口会返回 Excel 的 Sheet 与列结构。"""

    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_test",
            "type": "local_excel",
            "path": str(TEST_DATA_PATH),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_id"] == "src_test"
    assert payload["data"]["source_type"] == "local_excel"
    assert payload["data"]["sheets"] == [
        {"name": "items", "columns": ["ID", "Name"]},
        {"name": "drops", "columns": ["RefID"]},
    ]


@pytest.mark.anyio
async def test_source_metadata_rejects_csv_for_variable_pool_dropdown(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证 CSV 数据源会被明确拦截。"""

    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("id\n1\n", encoding="utf-8")

    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_csv",
            "type": "local_csv",
            "path": str(csv_path),
        },
    )

    assert response.status_code == 400
    assert "变量池下拉提取目前仅支持 Excel 与 SVN" in response.json()["detail"]


@pytest.mark.anyio
async def test_source_metadata_returns_feishu_sheet_and_columns(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """飞书 metadata 接口应返回兼容变量池的 Sheet 与列结构。"""
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid001!A1:B1":
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "msg": "success",
                    "data": {
                        "spreadsheetToken": "shtcnabc123",
                        "valueRange": {
                            "range": "gid001!A1:B1",
                            "values": [["id", "name"]],
                        },
                    },
                },
            )
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid002!A1:A1":
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "msg": "success",
                    "data": {
                        "spreadsheetToken": "shtcnabc123",
                        "valueRange": {
                            "range": "gid002!A1:A1",
                            "values": [["status"]],
                        },
                    },
                },
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"] == {
        "source_id": "src_feishu",
        "source_type": "feishu",
        "sheets": [
            {"name": "Sheet1", "sheet_id": "gid001", "columns": ["id", "name"]},
            {"name": "Sheet2", "sheet_id": "gid002", "columns": ["status"]},
        ],
        "authorization_status": "authorized",
    }


@pytest.mark.anyio
async def test_source_metadata_feishu_can_skip_columns_for_sheet_list(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """礼包弹窗轻量读取 Sheet 列表时，不应逐个读取表头列。"""
    await _seed_feishu_bot_config(test_project_id)
    requested_paths: list[str] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requested_paths.append(request.url.path)
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_sheets_response()
        if "/values/" in request.url.path:
            return httpx.Response(500, json={"code": 500, "msg": "headers should be skipped"})
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/metadata?include_columns=false",
        json={
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"] == {
        "source_id": "src_feishu",
        "source_type": "feishu",
        "sheets": [
            {"name": "Sheet1", "sheet_id": "gid001", "columns": []},
            {"name": "Sheet2", "sheet_id": "gid002", "columns": []},
        ],
        "authorization_status": "authorized",
    }
    assert "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query" in requested_paths
    assert not any("/values/" in path for path in requested_paths)


@pytest.mark.anyio
async def test_source_metadata_feishu_requires_login() -> None:
    """飞书 metadata 需要当前项目上下文；未登录请求应返回 401。"""
    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/metadata",
            json={
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
        )

    assert response.status_code == 401
    assert response.json()["detail"] == "未提供认证令牌"


@pytest.mark.anyio
async def test_source_metadata_feishu_invalid_url_maps_error(
    auth_client: AsyncClient,
) -> None:
    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "not-a-url",
        },
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["code"] == FEISHU_INVALID_URL
    assert detail["msg"] == "请输入合法的飞书电子表格链接"


@pytest.mark.anyio
@pytest.mark.parametrize(
    ("api_response", "expected_status", "expected_code", "expected_message"),
    [
        (
            httpx.Response(403, json={"code": 1254030, "msg": "permission denied"}),
            403,
            FEISHU_DOCUMENT_PERMISSION_DENIED,
            "机器人暂无该表格权限，请发送授权请求到群。",
        ),
        (
            httpx.Response(404, json={"code": 1254040, "msg": "not found"}),
            404,
            FEISHU_DOCUMENT_NOT_FOUND,
            None,
        ),
    ],
)
async def test_source_metadata_feishu_api_errors_map_to_http_detail(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
    api_response: httpx.Response,
    expected_status: int,
    expected_code: str,
    expected_message: str | None,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return api_response
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
        },
    )

    assert response.status_code == expected_status
    detail = response.json()["detail"]
    assert detail["code"] == expected_code
    if expected_message is not None:
        assert detail["msg"] == expected_message


@pytest.mark.anyio
async def test_source_metadata_feishu_app_permission_missing(
    auth_client: AsyncClient,
) -> None:
    feishu_bot._TOKEN_CACHE.clear()
    feishu_bot._TOKEN_LOCKS.clear()

    response = await auth_client.post(
        "/api/v1/sources/metadata",
        json={
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
        },
    )

    assert response.status_code == 403
    detail = response.json()["detail"]
    assert detail["code"] == FEISHU_APP_PERMISSION_MISSING
    assert "飞书应用凭证不可用" in detail["msg"]


@pytest.mark.anyio
async def test_column_preview_returns_top_rows_for_variable_detail(
    auth_client: AsyncClient,
) -> None:
    """验证列预览接口会返回变量详情页签所需的前几行数据。"""

    response = await auth_client.post(
        "/api/v1/sources/column-preview",
        json={
            "source": {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            },
            "sheet": "items",
            "column": "ID",
            "limit": 3,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_id"] == "src_test"
    assert payload["data"]["sheet"] == "items"
    assert payload["data"]["column"] == "ID"
    assert payload["data"]["preview_limit"] == 3
    assert payload["data"]["total_rows"] == 5
    assert payload["data"]["preview_rows"] == [
        {"row_index": 2, "value": 1},
        {"row_index": 3, "value": 2},
        {"row_index": 4, "value": 2},
    ]


@pytest.mark.anyio
async def test_column_preview_without_limit_returns_full_column_for_detail_dialog(
    auth_client: AsyncClient,
) -> None:
    """验证详情弹窗在不传 limit 时会返回当前列的完整预览。"""

    response = await auth_client.post(
        "/api/v1/sources/column-preview",
        json={
            "source": {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            },
            "sheet": "items",
            "column": "ID",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_path"] == str(TEST_DATA_PATH)
    assert payload["data"]["total_rows"] == 5
    assert payload["data"]["loaded_rows"] == 5
    assert payload["data"]["loaded_all_rows"] is True
    assert payload["data"]["preview_limit"] == 5
    assert payload["data"]["preview_rows"] == [
        {"row_index": 2, "value": 1},
        {"row_index": 3, "value": 2},
        {"row_index": 4, "value": 2},
        {"row_index": 5, "value": None},
        {"row_index": 6, "value": "   "},
    ]


@pytest.mark.anyio
async def test_feishu_column_preview_returns_rows_with_real_row_index(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """飞书单列预览应过滤目标列空值，并保留真实表格行号。"""
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    [1, "Alpha", "A"],
                    [2, "", "B"],
                    [3, None, "C"],
                    [4, "   ", "D"],
                    [5, "Omega", "E"],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/column-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "Preview",
            "column": "name",
            "limit": 3,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["variable_kind"] == "single"
    assert payload["source_id"] == "src_feishu"
    assert payload["source_type"] == "feishu"
    assert payload["sheet"] == "Preview"
    assert payload["column"] == "name"
    assert payload["preview_limit"] == 3
    assert payload["total_rows"] == 5
    assert payload["loaded_rows"] == 2
    assert payload["loaded_all_rows"] is False
    assert payload["preview_rows"] == [
        {"row_index": 2, "value": "Alpha"},
        {"row_index": 6, "value": "Omega"},
    ]


@pytest.mark.anyio
async def test_feishu_column_preview_without_limit_returns_full_column(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    [1, "Alpha", "A"],
                    [2, "Beta", "B"],
                    [3, "Gamma", "C"],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/column-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "gid_preview",
            "column": "id",
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["preview_limit"] == 3
    assert payload["loaded_rows"] == 3
    assert payload["loaded_all_rows"] is True
    assert payload["preview_rows"] == [
        {"row_index": 2, "value": 1},
        {"row_index": 3, "value": 2},
        {"row_index": 4, "value": 3},
    ]


@pytest.mark.anyio
async def test_composite_preview_returns_json_mapping_for_same_sheet(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证组合变量预览接口会返回 key->object 的完整 JSON 映射。"""

    workbook_path = _create_composite_test_workbook(tmp_path / "composite_preview.xlsx")

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "items",
            "columns": ["ID", "ItemName", "Desc"],
            "key_column": "ID",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["variable_kind"] == "composite"
    assert payload["data"]["sheet"] == "items"
    assert payload["data"]["columns"] == ["ID", "ItemName", "Desc"]
    assert payload["data"]["key_column"] == "ID"
    assert payload["data"]["has_duplicate_keys"] is False
    assert payload["data"]["duplicate_keys_preview"] == []
    assert payload["data"]["total_rows"] == 3
    assert payload["data"]["loaded_rows"] == 3
    assert payload["data"]["mapping"] == {
        "1": {"ItemName": "Gold", "Desc": "黄金+100"},
        "2": {"ItemName": "Wood", "Desc": "木头+200"},
        "3": {"ItemName": "Stone", "Desc": "石头+300"},
    }
    assert payload["data"]["total_keys"] == 3
    assert payload["data"]["page"] == 1
    assert payload["data"]["page_size"] == 3
    assert payload["data"]["total_pages"] == 1


@pytest.mark.anyio
async def test_composite_preview_paginates_local_mapping(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """本地组合变量分页只返回当前页 mapping，同时保留全量统计。"""
    workbook_path = tmp_path / "composite_paged_preview.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": [1, 2, 3, 4, 5],
                "ItemName": ["Gold", "Wood", "Stone", "Iron", "Gem"],
                "Desc": ["A", "B", "C", "D", "E"],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "items",
            "columns": ["ID", "ItemName", "Desc"],
            "key_column": "ID",
            "page": 2,
            "size": 2,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["total_rows"] == 5
    assert payload["total_keys"] == 5
    assert payload["page"] == 2
    assert payload["page_size"] == 2
    assert payload["total_pages"] == 3
    assert payload["loaded_rows"] == 2
    assert payload["loaded_all_rows"] is False
    assert payload["mapping"] == {
        "3": {"ItemName": "Stone", "Desc": "C"},
        "4": {"ItemName": "Iron", "Desc": "D"},
    }


@pytest.mark.anyio
async def test_composite_preview_detects_duplicate_keys_outside_current_page(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """重复 key 即使不在当前页内，也必须阻断未追加序号的预览。"""
    workbook_path = tmp_path / "composite_paged_duplicate_keys.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": ["A", "B", "C", "D", "A"],
                "Name": ["Alpha", "Beta", "Gamma", "Delta", "AlphaAgain"],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "items",
            "columns": ["ID", "Name"],
            "key_column": "ID",
            "page": 1,
            "size": 2,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["A"]
    assert payload["total_keys"] == 5
    assert payload["mapping"] == {}
    assert payload["loaded_rows"] == 0


@pytest.mark.anyio
async def test_composite_preview_pagination_keeps_original_index_suffix(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """分页追加序号时，后缀应沿用原始数据行位置而不是页内位置。"""
    workbook_path = tmp_path / "composite_paged_append_index.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": ["A", "A", "B", "C"],
                "Name": ["Alpha", "Alpha2", "Beta", "Gamma"],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "items",
            "columns": ["ID", "Name"],
            "key_column": "ID",
            "append_index_to_key": True,
            "page": 2,
            "size": 2,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["mapping"] == {
        "B_2": {"Name": "Beta"},
        "C_3": {"Name": "Gamma"},
    }


@pytest.mark.anyio
async def test_composite_preview_supports_appending_index_to_duplicate_keys(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证组合变量预览可将重复 key 生成为原值_序号。"""
    workbook_path = tmp_path / "composite_duplicate_keys.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": ["A", "A", "B"],
                "INT_Group": [1, 2, 3],
                "INT_Faction": [0, 1, 1],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "items",
            "columns": ["INT_ID", "INT_Group", "INT_Faction"],
            "key_column": "INT_ID",
            "append_index_to_key": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["A"]
    assert payload["mapping"] == {
        "A_0": {"INT_Group": 1, "INT_Faction": 0},
        "A_1": {"INT_Group": 2, "INT_Faction": 1},
        "B_2": {"INT_Group": 3, "INT_Faction": 1},
    }


@pytest.mark.anyio
async def test_composite_preview_reports_duplicate_keys_without_append_mode(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证重复 key 且未开启追加序号时，预览仍返回重复标记供前端提示。"""
    workbook_path = tmp_path / "composite_duplicate_keys_without_append.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "STR_ParamType": ["list", "list", "other"],
                "INT_Group": [1, 2, 3],
                "INT_Faction": [0, 1, 1],
            }
        ).to_excel(writer, sheet_name="switch", index=False)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "switch",
            "columns": ["STR_ParamType", "INT_Group", "INT_Faction"],
            "key_column": "STR_ParamType",
            "append_index_to_key": False,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["list"]
    assert payload["mapping"] == {}
    assert payload["loaded_rows"] == 0


@pytest.mark.anyio
async def test_feishu_composite_preview_returns_mapping_and_skips_empty_keys(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    [1, "Alpha", "A"],
                    ["", "NoKey", "B"],
                    [3, None, "C"],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "Preview",
            "columns": ["id", "name", "group"],
            "key_column": "id",
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["variable_kind"] == "composite"
    assert payload["source_type"] == "feishu"
    assert payload["sheet"] == "Preview"
    assert payload["columns"] == ["id", "name", "group"]
    assert payload["key_column"] == "id"
    assert payload["has_duplicate_keys"] is False
    assert payload["duplicate_keys_preview"] == []
    assert payload["total_rows"] == 3
    assert payload["loaded_rows"] == 2
    assert payload["loaded_all_rows"] is True
    assert payload["mapping"] == {
        "1": {"name": "Alpha", "group": "A"},
        "3": {"name": None, "group": "C"},
    }
    assert payload["total_keys"] == 2
    assert payload["page"] == 1
    assert payload["page_size"] == 2
    assert payload["total_pages"] == 1


@pytest.mark.anyio
async def test_feishu_composite_preview_paginates_mapping(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    [1, "Alpha", "A"],
                    ["", "NoKey", "B"],
                    [3, "Gamma", "C"],
                    [4, "Delta", "D"],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "Preview",
            "columns": ["id", "name", "group"],
            "key_column": "id",
            "page": 2,
            "size": 1,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["total_rows"] == 4
    assert payload["total_keys"] == 3
    assert payload["page"] == 2
    assert payload["page_size"] == 1
    assert payload["total_pages"] == 3
    assert payload["loaded_rows"] == 1
    assert payload["mapping"] == {
        "3": {"name": "Gamma", "group": "C"},
    }


@pytest.mark.anyio
async def test_feishu_composite_preview_duplicate_keys_without_append_mode(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    ["A", "Alpha", 1],
                    ["A", "Beta", 2],
                    ["B", "Gamma", 3],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "Preview",
            "columns": ["id", "name", "group"],
            "key_column": "id",
            "append_index_to_key": False,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["A"]
    assert payload["mapping"] == {}
    assert payload["loaded_rows"] == 0


@pytest.mark.anyio
async def test_feishu_composite_preview_appends_index_to_duplicate_keys(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [
                    ["id", "name", "group"],
                    ["A", "Alpha", 1],
                    ["A", "Beta", 2],
                    ["B", "Gamma", 3],
                ]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)

    response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_feishu",
                "type": "feishu",
                "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
            },
            "sheet": "Preview",
            "columns": ["id", "name", "group"],
            "key_column": "id",
            "append_index_to_key": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["A"]
    assert payload["mapping"] == {
        "A_0": {"name": "Alpha", "group": 1},
        "A_1": {"name": "Beta", "group": 2},
        "B_2": {"name": "Gamma", "group": 3},
    }


@pytest.mark.anyio
@pytest.mark.parametrize(
    ("request_payload", "expected_message"),
    [
        (
            {
                "endpoint": "/api/v1/sources/column-preview",
                "body": {"sheet": "Missing", "column": "id", "limit": 3},
            },
            "未找到指定 Sheet",
        ),
        (
            {
                "endpoint": "/api/v1/sources/column-preview",
                "body": {"sheet": "Preview", "column": "missing", "limit": 3},
            },
            "未找到指定列",
        ),
        (
            {
                "endpoint": "/api/v1/sources/composite-preview",
                "body": {
                    "sheet": "Preview",
                    "columns": ["id", "name"],
                    "key_column": "group",
                    "append_index_to_key": False,
                },
            },
            "主键列必须包含在组合列中",
        ),
    ],
)
async def test_feishu_preview_validation_errors_return_chinese_messages(
    auth_client: AsyncClient,
    test_project_id: int,
    monkeypatch: pytest.MonkeyPatch,
    request_payload: dict[str, Any],
    expected_message: str,
) -> None:
    await _seed_feishu_bot_config(test_project_id)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == feishu_bot.TENANT_ACCESS_TOKEN_PATH:
            return _feishu_token_response()
        if request.url.path == "/open-apis/sheets/v3/spreadsheets/shtcnabc123/sheets/query":
            return _feishu_preview_sheets_response()
        if request.url.path == "/open-apis/sheets/v2/spreadsheets/shtcnabc123/values/gid_preview!A1:C6":
            return _feishu_preview_values_response(
                [["id", "name", "group"], [1, "Alpha", "A"]]
            )
        return httpx.Response(404, json={"code": 404, "msg": "not found"})

    _install_feishu_metadata_mock(monkeypatch, handler)
    body = {
        "source": {
            "id": "src_feishu",
            "type": "feishu",
            "pathOrUrl": "https://demo.feishu.cn/sheets/shtcnabc123",
        },
        **request_payload["body"],
    }

    response = await auth_client.post(request_payload["endpoint"], json=body)

    assert response.status_code == 400
    assert expected_message in response.json()["detail"]


@pytest.mark.anyio
async def test_preview_endpoints_preserve_raw_sheet_and_column_names(
    auth_client: AsyncClient,
    tmp_path: Path,
) -> None:
    """验证预览接口读取 Excel 时保留原始 Sheet 名和列名，不裁掉尾部空格。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "whitespace_headers.xlsx")

    column_response = await auth_client.post(
        "/api/v1/sources/column-preview",
        json={
            "source": {
                "id": "src_space",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "Quest  ",
            "column": "STR_ABSwitch  ",
        },
    )
    composite_response = await auth_client.post(
        "/api/v1/sources/composite-preview",
        json={
            "source": {
                "id": "src_space",
                "type": "local_excel",
                "path": str(workbook_path),
            },
            "sheet": "Quest  ",
            "columns": ["INT_ID", "STR_ABSwitch  ", "DESC3"],
            "key_column": "INT_ID",
        },
    )

    assert column_response.status_code == 200
    column_payload = column_response.json()["data"]
    assert column_payload["sheet"] == "Quest  "
    assert column_payload["column"] == "STR_ABSwitch  "
    assert column_payload["preview_rows"] == [
        {"row_index": 2, "value": "on"},
        {"row_index": 3, "value": "off"},
    ]

    assert composite_response.status_code == 200
    composite_payload = composite_response.json()["data"]
    assert composite_payload["sheet"] == "Quest  "
    assert composite_payload["columns"] == ["INT_ID", "STR_ABSwitch  ", "DESC3"]
    assert composite_payload["mapping"] == {
        "1001": {"STR_ABSwitch  ": "on", "DESC3": "开"},
        "1002": {"STR_ABSwitch  ": "off", "DESC3": "关"},
    }


@pytest.mark.anyio
async def test_execute_engine_supports_trimmed_sheet_and_column_identifiers(
    tmp_path: Path,
) -> None:
    """验证执行链路可兼容已被 trim 的 Sheet/列配置，并正确解析真实表头。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "trimmed_execute.xlsx")
    payload = {
        "sources": [
            {
                "id": "src_space",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[quest-switch]",
                "source_id": "src_space",
                "sheet": "Quest",
                "column": "STR_ABSwitch",
            }
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[quest-switch]"]},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    response_payload = response.json()
    assert response_payload["meta"]["failed_sources"] == []
    assert response_payload["meta"]["total_rows_scanned"] == 2
    assert response_payload["data"]["abnormal_results"] == []


def test_load_local_variables_preserves_raw_excel_identifiers(
    tmp_path: Path,
) -> None:
    """验证执行链路加载变量时也保留原始 Sheet 名和列名。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "whitespace_execute.xlsx")

    loaded_variables = load_local_variables(
        [
            DataSource(
                id="src_space",
                type="local_excel",
                path=str(workbook_path),
            )
        ],
        [
            VariableTag(
                tag="[quest-switch-single]",
                source_id="src_space",
                sheet="Quest  ",
                variable_kind="single",
                column="STR_ABSwitch  ",
                expected_type="str",
            ),
            VariableTag(
                tag="[quest-switch-composite]",
                source_id="src_space",
                sheet="Quest  ",
                variable_kind="composite",
                columns=["INT_ID", "STR_ABSwitch  ", "DESC3"],
                key_column="INT_ID",
                expected_type="json",
            ),
        ],
    )

    single_frame = loaded_variables["[quest-switch-single]"]
    assert single_frame.columns.tolist() == ["STR_ABSwitch  ", "_row_index"]
    assert single_frame["STR_ABSwitch  "].tolist() == ["on", "off"]

    composite_frame = loaded_variables["[quest-switch-composite]"]
    assert composite_frame.columns.tolist() == [
        "__key__",
        "INT_ID",
        "STR_ABSwitch  ",
        "DESC3",
        "_row_index",
    ]
    assert composite_frame["__key__"].tolist() == ["1001", "1002"]
    assert composite_frame["INT_ID"].tolist() == [1001, 1002]


@pytest.mark.anyio
async def test_execute_engine_accepts_composite_variable_without_breaking_rules(
    tmp_path: Path,
) -> None:
    """验证 TaskTree 中包含组合变量时，现有三类规则仍可对单变量正常执行。"""

    workbook_path = _create_composite_test_workbook(tmp_path / "composite_execute.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_combo",
                "sheet": "items",
                "variable_kind": "single",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_combo",
                "sheet": "drops",
                "variable_kind": "single",
                "column": "RefID",
            },
            {
                "tag": "[items-json]",
                "source_id": "src_combo",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["ID", "ItemName", "Desc"],
                "key_column": "ID",
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    response_payload = response.json()
    assert response_payload["msg"] == "Execution Completed"
    assert response_payload["meta"]["failed_sources"] == []
    assert response_payload["meta"]["total_rows_scanned"] > 0
    assert any(
        item["rule_name"] == "cross_table_mapping" and item["raw_value"] == 9
        for item in response_payload["data"]["abnormal_results"]
    )


@pytest.mark.anyio
async def test_execute_engine_supports_dual_composite_compare(tmp_path: Path) -> None:
    """验证个人校验执行接口支持双组合变量按 Key 关联后比较字段值。"""
    workbook_path = _create_dual_composite_test_workbook(tmp_path / "dual_composite_rules.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_dual",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[left-json]",
                "source_id": "src_dual",
                "sheet": "left_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "expected_type": "json",
            },
            {
                "tag": "[right-json]",
                "source_id": "src_dual",
                "sheet": "right_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[left-json]",
                    "reference_tag": "[right-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "双组合变量比对",
                    "comparisons": [
                        {
                            "comparison_id": "compare-condition-type",
                            "left_field": "INT_ConditionType",
                            "operator": "eq",
                            "right_field": "INT_ConditionType",
                        },
                        {
                            "comparison_id": "compare-require-rule",
                            "left_field": "INT_RequireRule",
                            "operator": "eq",
                            "right_field": "INT_RequireRule",
                        },
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["rule_name"] == "双组合变量比对"
    assert "Key 10001" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_supports_same_dual_composite_filtered_compare(
    tmp_path: Path,
) -> None:
    """验证同一组合变量可拆成左右筛选子集后按 Key 比对。"""
    workbook_path = tmp_path / "same_dual_composite_filters.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [10001, 10001, 10002, 10002],
                "SIDE": ["left", "right", "left", "right"],
                "VALUE": [1, 2, 3, 3],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_same_dual", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-same-json]",
                "source_id": "src_same_dual",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "SIDE", "VALUE"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-same-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-same-json]",
                    "reference_tag": "[items-same-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "同变量筛选比对",
                    "left_filters": [
                        {
                            "condition_id": "left-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "left",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "right",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-value",
                            "left_field": "VALUE",
                            "operator": "eq",
                            "right_field": "VALUE",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert "Key 10001" in abnormal_results[0]["message"]
    assert "左侧筛选" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_same_dual_compare_uses_explicit_business_key(
    tmp_path: Path,
) -> None:
    """验证 append_index_to_key 开启时同变量筛选可按业务字段对齐。"""
    workbook_path = tmp_path / "same_dual_business_key.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1, 1, 2, 2],
                "INT_Index": [1012, 1010, 1012, 1010],
                "INT_Level": [1, 1, 2, 2],
                "INT_FreeRewardSubType": [7, 7, 8, 8],
                "INT_FreeRewardValue": [100, 100, 200, 200],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_business_key", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-reward-json]",
                "source_id": "src_business_key",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": [
                    "INT_ID",
                    "INT_Index",
                    "INT_Level",
                    "INT_FreeRewardSubType",
                    "INT_FreeRewardValue",
                ],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-business-key",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-reward-json]",
                    "reference_tag": "[items-reward-json]",
                    "key_check_mode": "baseline_only",
                    "left_key_field": "INT_Level",
                    "right_key_field": "INT_Level",
                    "rule_name": "奖励配置按等级对齐",
                    "left_filters": [
                        {
                            "condition_id": "left-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1012",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1010",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-subtype",
                            "left_field": "INT_FreeRewardSubType",
                            "operator": "eq",
                            "right_field": "INT_FreeRewardSubType",
                        },
                        {
                            "comparison_id": "compare-value",
                            "left_field": "INT_FreeRewardValue",
                            "operator": "eq",
                            "right_field": "INT_FreeRewardValue",
                        },
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    assert response.json()["data"]["abnormal_results"] == []


@pytest.mark.anyio
async def test_execute_engine_dual_compare_rejects_duplicate_explicit_key_after_filter(
    tmp_path: Path,
) -> None:
    """验证显式关联 Key 字段在筛选后仍要求唯一。"""
    workbook_path = tmp_path / "same_dual_duplicate_business_key.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1, 2, 3],
                "INT_Index": [1012, 1012, 1010],
                "INT_Level": [1, 1, 1],
                "VALUE": [100, 100, 100],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_dup_business_key", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-dup-business-key]",
                "source_id": "src_dup_business_key",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_Index", "INT_Level", "VALUE"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-duplicate-business-key",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-dup-business-key]",
                    "reference_tag": "[items-dup-business-key]",
                    "key_check_mode": "baseline_only",
                    "left_key_field": "INT_Level",
                    "right_key_field": "INT_Level",
                    "rule_name": "显式 Key 重复",
                    "left_filters": [
                        {
                            "condition_id": "left-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1012",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1010",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-value",
                            "left_field": "VALUE",
                            "operator": "eq",
                            "right_field": "VALUE",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "关联 Key 字段 INT_Level 存在重复值" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_dual_bidirectional_checks_filtered_right_keys(
    tmp_path: Path,
) -> None:
    """验证双向缺失 Key 只基于筛选后的右侧 Key 集合。"""
    workbook_path = tmp_path / "same_dual_bidirectional_business_key.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1, 2, 3],
                "INT_Index": [1012, 1010, 9999],
                "INT_Level": [1, 2, 99],
                "VALUE": [100, 200, 999],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_bidir_key", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-bidir-business-key]",
                "source_id": "src_bidir_key",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_Index", "INT_Level", "VALUE"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-bidir-business-key",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-bidir-business-key]",
                    "reference_tag": "[items-bidir-business-key]",
                    "key_check_mode": "bidirectional",
                    "left_key_field": "INT_Level",
                    "right_key_field": "INT_Level",
                    "rule_name": "双向筛选 Key",
                    "left_filters": [
                        {
                            "condition_id": "left-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1012",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-index",
                            "field": "INT_Index",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "1010",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-value",
                            "left_field": "VALUE",
                            "operator": "eq",
                            "right_field": "VALUE",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 2
    assert {item["raw_value"] for item in abnormal_results} == {1, 2}
    assert all(item["raw_value"] != 99 for item in abnormal_results)


@pytest.mark.anyio
async def test_execute_engine_dual_composite_empty_filter_returns_warning(
    tmp_path: Path,
) -> None:
    """验证双组合变量筛选为空时返回 warning 而不是静默通过。"""
    workbook_path = tmp_path / "dual_empty_filter_warning.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [10001],
                "SIDE": ["left"],
                "VALUE": [1],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_empty_dual", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-empty-json]",
                "source_id": "src_empty_dual",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "SIDE", "VALUE"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-empty-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-empty-json]",
                    "reference_tag": "[items-empty-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "同变量空筛选",
                    "left_filters": [
                        {
                            "condition_id": "left-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "missing",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "left",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-value",
                            "left_field": "VALUE",
                            "operator": "eq",
                            "right_field": "VALUE",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["level"] == "warning"
    assert "左侧筛选后无数据" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_dual_composite_duplicate_key_after_filter_returns_400(
    tmp_path: Path,
) -> None:
    """验证筛选后的重复 Key 会阻断双组合变量比对。"""
    workbook_path = tmp_path / "dual_duplicate_key_after_filter.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [10001, 10001, 10002],
                "SIDE": ["left", "left", "right"],
                "VALUE": [1, 1, 1],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    payload = {
        "sources": [{"id": "src_duplicate_dual", "type": "local_excel", "path": str(workbook_path)}],
        "variables": [
            {
                "tag": "[items-duplicate-json]",
                "source_id": "src_duplicate_dual",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "SIDE", "VALUE"],
                "key_column": "INT_ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-duplicate-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[items-duplicate-json]",
                    "reference_tag": "[items-duplicate-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "同变量重复 Key",
                    "left_filters": [
                        {
                            "condition_id": "left-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "left",
                        }
                    ],
                    "right_filters": [
                        {
                            "condition_id": "right-side",
                            "field": "SIDE",
                            "operator": "eq",
                            "value_source": "literal",
                            "expected_value": "right",
                        }
                    ],
                    "comparisons": [
                        {
                            "comparison_id": "compare-value",
                            "left_field": "VALUE",
                            "operator": "eq",
                            "right_field": "VALUE",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "Key" in response.json()["detail"]
    assert "10001" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_returns_configured_display_field_value(tmp_path: Path) -> None:
    """异常结果按规则配置输出当前关联变量内的显示字段值。"""
    workbook_path = tmp_path / "display_field_rules.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": [1, 2, 3],
                "Faction": [1, 1, 1],
                "Name": ["Ok", "Ok", "Bad"],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    composite_config = {
        "global_filters": [],
        "branches": [
            {
                "branch_id": "branch-display",
                "filters": [],
                "assertions": [
                    {
                        "condition_id": "name-must-ok",
                        "field": "Name",
                        "operator": "eq",
                        "value_source": "literal",
                        "expected_value": "Ok",
                    }
                ],
            }
        ],
    }
    payload = {
        "sources": [
            {
                "id": "src_display",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-composite]",
                "source_id": "src_display",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["ID", "Faction", "Name"],
                "key_column": "ID",
                "expected_type": "json",
            }
        ],
        "rules": [
            {
                "rule_type": "composite_condition_check",
                "params": {
                    "target_tag": "[items-composite]",
                    "rule_name": "名称显示字段校验",
                    "display_field": "Name",
                    "composite_config": composite_config,
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["display_value"] == "Bad"


@pytest.mark.anyio
async def test_execute_engine_uses_appended_composite_keys_during_runtime(tmp_path: Path) -> None:
    """验证执行链路中的组合变量 __key__ 与预览使用相同的原值_序号口径。"""
    workbook_path = _create_dual_composite_test_workbook(tmp_path / "dual_composite_suffix_keys.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_dual",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[left-json]",
                "source_id": "src_dual",
                "sheet": "left_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            },
            {
                "tag": "[right-json]",
                "source_id": "src_dual",
                "sheet": "right_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[left-json]",
                    "reference_tag": "[right-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "双组合变量比对",
                    "comparisons": [
                        {
                            "comparison_id": "compare-suffixed-key",
                            "left_field": "__key__",
                            "operator": "eq",
                            "right_field": "__key__",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    assert response.json()["data"]["abnormal_results"] == []


@pytest.mark.anyio
async def test_execute_engine_persists_latest_result_and_supports_server_side_pagination(
    tmp_path: Path,
    auth_headers: dict[str, str],
    test_project_id: int,
) -> None:
    """验证个人校验执行后会返回第一页结果，并支持按 result_id 翻页。"""
    workbook_path = _create_paginated_test_workbook(tmp_path / "paged_execute.xlsx")
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-cross",
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            }
        ],
        "selected_rule_ids": ["rule-cross"],
        "page": 1,
        "size": 20,
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
        headers=auth_headers,
    ) as client:
        execute_response = await client.post("/api/v1/engine/execute", json=payload)

        assert execute_response.status_code == 200
        execute_payload = execute_response.json()
        assert execute_payload["msg"] == "Execution Completed"
        assert execute_payload["meta"]["result_id"] > 0
        assert execute_payload["data"]["page"] == 1
        assert execute_payload["data"]["size"] == 20
        assert execute_payload["data"]["total"] == 35
        assert len(execute_payload["data"]["list"]) == 20
        assert execute_payload["data"]["abnormal_results"] == execute_payload["data"]["list"]

        result_id = execute_payload["meta"]["result_id"]
        page_two_response = await client.get(
            f"/api/v1/engine/results/{result_id}",
            params={"page": 2, "size": 20},
        )
        export_response = await client.get(f"/api/v1/engine/results/{result_id}/export")

    assert page_two_response.status_code == 200
    page_two_payload = page_two_response.json()
    assert page_two_payload["meta"]["result_id"] == result_id
    assert page_two_payload["data"]["page"] == 2
    assert page_two_payload["data"]["size"] == 20
    assert page_two_payload["data"]["total"] == 35
    assert len(page_two_payload["data"]["list"]) == 15
    assert page_two_payload["data"]["list"][0]["raw_value"] == 31

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_response.content))
    assert workbook.sheetnames == ["统计摘要", "异常明细"]
    assert workbook["统计摘要"]["B2"].value == "个人校验"
    assert workbook["异常明细"].max_row == 36
    assert workbook["异常明细"]["A1"].value == "级别"

    async with async_session_factory() as session:
        other_user = User(
            username="other-workbench-user",
            hashed_password=hash_password("testpass"),
            primary_project_id=test_project_id,
        )
        session.add(other_user)
        await session.flush()
        session.add(
            UserProjectRole(
                user_id=other_user.id,
                project_id=test_project_id,
                role="user",
            )
        )
        await session.commit()
        other_token = create_access_token(other_user.id, project_id=test_project_id)

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
        headers={"Authorization": f"Bearer {other_token}"},
    ) as client:
        forbidden_export_response = await client.get(
            f"/api/v1/engine/results/{result_id}/export"
        )

    assert forbidden_export_response.status_code == 404


@pytest.mark.anyio
async def test_execute_engine_runs_svn_remote_source(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """SVN 远端 URL 数据源走 prepare_remote_svn_source 落到本地 fixture，
    应当与等价的 local_excel 配置产出相同的异常结果。"""
    from backend.app.loaders import svn_cache

    fixture_xlsx = tmp_path / "remote_quests.xlsx"
    pd.DataFrame({"ID": [1, 2, 3], "Name": ["A", "", "C"]}).to_excel(
        fixture_xlsx,
        sheet_name="items",
        index=False,
    )

    def _fake_prepare(source, *, user_scope=None, force_refresh=False):
        return fixture_xlsx

    monkeypatch.setattr(svn_cache, "prepare_remote_svn_source", _fake_prepare)

    payload = {
        "sources": [
            {
                "id": "src_remote",
                "type": "svn",
                "pathOrUrl": "https://samosvn/data/project/samo/GameDatas/datas_qa88/remote_quests.xlsx",
            }
        ],
        "variables": [
            {"tag": "[items-name]", "source_id": "src_remote", "sheet": "items", "column": "Name"}
        ],
        "rules": [
            {
                "rule_id": "rule_not_null_name",
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-name]"]},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200, response.text
    body = response.json()
    abnormal_results = body["data"]["abnormal_results"]
    assert any(item["rule_name"] for item in abnormal_results)
    # 至少能命中一行空 Name 异常
    assert len(abnormal_results) >= 1
    failed_sources = body["meta"]["failed_sources"]
    assert "src_remote" not in failed_sources
