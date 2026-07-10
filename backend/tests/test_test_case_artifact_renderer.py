"""Canonical Generation Run artifact renderer contract tests."""

from __future__ import annotations

import json
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from backend.app.test_cases.artifact_renderer import build_canonical_test_case_workbook
from backend.app.test_cases.case_quality_audit import (
    _audit_cases,
    _recommended_run_status,
    _safe_repair_fields,
)


def _cases() -> list[dict[str, object]]:
    return [
        {
            "case_id": "TC-0001",
            "fields": {
                "primary_module": "开关/入口",
                "secondary_module": "活动入口",
                "checkpoint": "活动开启时入口出现",
                "preconditions": "活动开关开启，当前时间为 2026-07-10 10:00",
                "steps": "1. 登录账号\n2. 打开主界面",
                "expected_results": "主界面展示活动入口，入口可点击且不报错",
                "priority": "P0",
                "remarks": "来源 Requirement Atom：RA-0001",
            },
            "atom_refs": ["RA-0001"],
        },
        {
            "case_id": "TC-0002",
            "fields": {
                "primary_module": "开关/入口",
                "secondary_module": "活动入口",
                "checkpoint": "活动关闭时入口消失",
                "preconditions": "活动开关关闭",
                "steps": "1. 重新登录\n2. 查看主界面",
                "expected_results": "主界面不展示活动入口",
                "priority": "P1",
                "remarks": "来源 Requirement Atom：RA-0002",
            },
            "atom_refs": ["RA-0002"],
        },
    ]


def test_renderer_uses_fixed_four_sheet_template_contract() -> None:
    output = build_canonical_test_case_workbook(
        cases=_cases(),
        title="重塑活动改系统功能",
        source_summary="上传文件：需求.xlsx",
        blueprint={"modules": ["活动入口"]},
        atoms=[
            {
                "atom_id": "RA-0001",
                "atom_type": "rule",
                "source_sheet_name": "详案",
                "source_row_start": 12,
                "requirement_text": "活动开启时显示入口",
            },
            {
                "atom_id": "RA-Q001",
                "atom_type": "open_question",
                "source_sheet_name": "详案",
                "source_row_start": 18,
                "requirement_text": "战力是否计入排行待确认",
            },
        ],
        coverage_audit={"uncovered_atom_ids": []},
        quality_audit={
            "status": "completed",
            "blocking_count": 0,
            "warning_count": 0,
            "issues": [],
        },
        metadata={
            "title": "重塑活动改系统功能",
            "run_id": 7001,
            "status": "completed",
            "strict_mode": False,
        },
    )

    workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
    assert workbook.sheetnames == ["测试用例", "用例蓝图", "生成说明", "覆盖审计"]

    sheet = workbook["测试用例"]
    assert sheet.max_column == 12
    assert [sheet.cell(10, column).value for column in range(2, 13)] == [
        "一级模块",
        "二级模块",
        "检查点",
        None,
        None,
        None,
        None,
        None,
        "T/F/D/N/A",
        "T/F/D/N/A",
        "T/F/D/N/A",
    ]
    assert sheet["A11"].value == 1
    assert sheet["A12"].value == 2
    assert sheet["B11"].value == "开关/入口"
    assert "B11:B12" in {str(item) for item in sheet.merged_cells.ranges}
    assert "C11:C12" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet["B2"].value == "=COUNTA(A11:A12)"
    assert sheet["J3"].value == '=COUNTIF(J11:J12,"T")'
    assert sheet["L8"].value == "=IFERROR(L3/(L3+L4),0)"
    assert "C:\\" not in str(sheet["B7"].value)

    validations = {validation.formula1: str(validation.sqref) for validation in sheet.data_validations.dataValidation}
    assert validations['"P0,P1,P2,P3"'] == "H11:H12"
    assert validations['"T,F,D,N/A"'] == "J11:L12"

    audit = workbook["覆盖审计"]
    assert audit["F2"].value == "covered"
    assert audit["F3"].value == "not_case_bearing"


def test_quality_repair_normalizes_legacy_fields_without_inventing_expectation() -> None:
    repaired = _safe_repair_fields(
        {
            "module": "活动入口",
            "feature": "活动入口",
            "scenario": "入口开放",
            "title": "活动入口开放",
            "steps": "登录并查看入口",
            "expected_results": "显示正确",
            "priority": "high",
        },
        case_id="TC-0001",
        atom_refs=["RA-0001"],
    )

    assert repaired["primary_module"] == "开关/入口"
    assert repaired["preconditions"] == "无特殊前置条件"
    assert repaired["priority"] == "P2"
    assert repaired["remarks"] == "来源 Requirement Atom：RA-0001"
    assert repaired["expected_results"] == "显示正确"

    record = SimpleNamespace(
        case_id="TC-0001",
        fields_json=json.dumps(repaired, ensure_ascii=False),
        atom_refs_json=json.dumps(["RA-0001"], ensure_ascii=False),
    )
    issues = _audit_cases(
        [record],
        atom_text={"RA-0001": "入口在活动开启 10 分钟后展示"},
    )
    assert {item["code"] for item in issues} == {
        "GENERIC_EXPECTATION",
        "NUMERIC_EXPECTATION_MISSING",
    }
    assert _recommended_run_status(
        case_count=1,
        blocking_count=1,
        coverage_status="completed",
    ) == "partial_completed"
    assert _recommended_run_status(
        case_count=0,
        blocking_count=0,
        coverage_status="completed",
    ) == "failed"
