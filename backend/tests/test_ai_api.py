"""AI 规则助手接口回归。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select

from backend.app.ai.hint_extractor import extract_workflow_hints_from_text
from backend.app.ai.prompts import build_prompt_optimize_system_prompt
from backend.app.ai.providers import ProviderConnectionError
from backend.app.database import async_session_factory
from backend.app.models import (
    AiProviderCredentialRecord,
    AiRuleDraftRecord,
    Project,
    User,
    WorkbenchConfigRecord,
)
from backend.app.security.crypto import decrypt_secret


DATA_DIR = Path(__file__).resolve().parent / "data"
MINIMAL_XLSX = DATA_DIR / "minimal_rules.xlsx"
SERVER_CONFIG_NATURAL_DESCRIPTION = (
    "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls"
    "的switch分页进行工作流模拟，规则是验证STR_ServersParam字段的配置数据格式。"
    "验证示例：313:1;315:1;317:1;666:0;10010:1;888:1。"
    "配置的冒号后面只能配置1或0，需要过滤掉DES字段里包含废弃的字段。"
    "可以先进行校验配置结果，然后再使用当前的功能添加校验规则。"
)
BATTLEPASS_NATURAL_DESCRIPTION = (
    "检查https://samosvn/data/project/samo/GameDatas/datas_qa88/battlepass.xls的level_reward分页，"
    "筛选INT_Index=1011和INT_Index=1010的配置，以INT_Level字段为key值，"
    "判断INT_FreeRewardSubType、INT_FreeRewardValue、INT_PayRewardSubType1、INT_PayRewardValue1的四个字段，"
    "筛选出来的两组配置是不是相等"
)
BPCHERKS_COMMA_DUAL_DESCRIPTION = (
    "筛选INT_Index=1012,1010,以INT_Level为key值，判断："
    "INT_FreeRewardSubType, INT_FreeRewardValue, INT_FreeRewardSubType1, INT_FreeRewardValue1, "
    "INT_FreeRewardSubType2, INT_FreeRewardValue2等字段的值相等"
)
BPCHERKS_SHORT_TEMPLATE_DUAL_DESCRIPTION = """筛选：
- INT_Index = 1012,1010

Key值选择：INT_Level

判定：INT_FreeRewardSubType,INT_FreeRewardValue,INT_FreeRewardSubType1,INT_FreeRewardValue1,INT_FreeRewardSubType2,INT_FreeRewardValue2 在 INT_Index=1012 和 INT_Index=1010 两组中必须相等"""
BPCHERKS_COMPARE_FIELDS = [
    "INT_FreeRewardSubType",
    "INT_FreeRewardValue",
    "INT_PayRewardSubType1",
    "INT_PayRewardValue1",
    "INT_PayRewardSubType2",
    "INT_PayRewardValue2",
]
SERVER_CONFIG_PATTERN = r"^(?:(?:all|\d+(?:-\d+)?):[01](;(?:all|\d+(?:-\d+)?):[01])*)?$"
QUESTS_SWITCH_DESCRIPTION = (
    "校验https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls配置表的Quest分页，"
    "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
    "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
)
QUESTS_SWITCH_OPTIMIZED_DESCRIPTION = """规则目标：
基于用户输入和项目元数据线索，将原始规则描述整理为更适合 AI 校验解析的结构化表达。

目标变量：
需要用户确认：未选择目标变量。

筛选条件：
global：DESC3 eq 升级p1建筑到p2级p4次,完成p2等级的p1科研

关联 Key：
需要用户确认：未识别到 Key 字段。

校验逻辑：
校验https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls配置表的Quest分页，筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。STR_ABSwitch字段=GreenServer:0 or SLG2:0。

比较字段：
- STR_ABSwitch

期望规则类型：
composite_condition_check

需要用户确认：
请确认数据源、Sheet、字段和变量标签；AI 校验时可尝试自动补齐。"""


def test_rule_prompt_optimize_system_prompt_documents_unsupported_aggregation() -> None:
    """优化提示词要明确聚合类规则不能硬套到现有规则类型。"""
    prompt = build_prompt_optimize_system_prompt()
    assert "聚合" in prompt
    assert "平均值" in prompt
    assert "不要强行映射成已支持类型" in prompt


@pytest.fixture(autouse=True)
def _reset_ai_draft_rate_limit() -> None:
    """AI 草稿接口有用户级限流，单文件回归里每例清空避免测试相互影响。"""
    from backend.app.api import ai_api

    ai_api._draft_call_times.clear()  # noqa: SLF001 - 测试隔离限流桶。


async def _get_test_user_id() -> int:
    async with async_session_factory() as session:
        user_id = await session.scalar(select(User.id).where(User.username == "testuser"))
        assert user_id is not None
        return user_id


async def _seed_workbench_config(project_id: int, user_id: int) -> None:
    config = {
        "version": 1,
        "configured": True,
        "sources": [
            {
                "id": "src_demo",
                "type": "local_excel",
                "pathOrUrl": str(MINIMAL_XLSX),
            }
        ],
        "variables": [
            {
                "tag": "[src_demo-items-ID]",
                "source_id": "src_demo",
                "sheet": "items",
                "variable_kind": "single",
                "column": "ID",
                "expected_type": "str",
            }
        ],
        "groups": [
            {
                "group_id": "ungrouped",
                "group_name": "默认规则组",
                "builtin": True,
            }
        ],
        "rules": [],
        "local_path_replacement_presets": [],
        "svn_path_replacement_presets": [],
    }
    async with async_session_factory() as session:
        session.add(
            WorkbenchConfigRecord(
                project_id=project_id,
                user_id=user_id,
                config_json=json.dumps(config, ensure_ascii=False),
            )
        )
        await session.commit()


async def _seed_test_composite_variable_config(project_id: int, user_id: int) -> None:
    config = {
        "version": 1,
        "configured": True,
        "sources": [
            {
                "id": "quests",
                "type": "local_excel",
                "pathOrUrl": str(MINIMAL_XLSX),
            }
        ],
        "variables": [
            {
                "tag": "test",
                "source_id": "quests",
                "sheet": "Quest",
                "variable_kind": "composite",
                "columns": ["DESC3", "STR_ABSwitch  ", "INT_ID"],
                "key_column": "INT_ID",
                "append_index_to_key": False,
                "expected_type": "json",
            }
        ],
        "groups": [
            {
                "group_id": "ungrouped",
                "group_name": "默认规则组",
                "builtin": True,
            }
        ],
        "rules": [],
        "local_path_replacement_presets": [],
        "svn_path_replacement_presets": [],
    }
    async with async_session_factory() as session:
        session.add(
            WorkbenchConfigRecord(
                project_id=project_id,
                user_id=user_id,
                config_json=json.dumps(config, ensure_ascii=False),
            )
        )
        await session.commit()


def _create_quests_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quest"
    sheet.append(["INT_ID", "DESC3", "STR_ABSwitch  "])
    sheet.append([1, "升级p1建筑到p2级p4次", "GreenServer:0"])
    strategic = workbook.create_sheet("Strategic_slg2")
    strategic.append(["INT_ID", "INT_Faction", "INT_Group"])
    strategic.append([1, 0, 1])
    workbook.save(path)
    return path


def _create_battlepass_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "level_reward"
    sheet.append(
        [
            "INT_Index",
            "INT_Level",
            "INT_FreeRewardSubType",
            "INT_FreeRewardValue",
            "INT_PayRewardSubType1",
            "INT_PayRewardValue1",
            "INT_PayRewardSubType2",
            "INT_PayRewardValue2",
        ]
    )
    sheet.append([1012, 1, 10, 100, 11, 101, 12, 102])
    sheet.append([1010, 1, 10, 100, 11, 101, 12, 102])
    workbook.save(path)
    return path


def _create_battlepass_user_regression_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "level_reward"
    sheet.append(
        [
            "INT_Index",
            "INT_Level",
            "INT_FreeRewardSubType",
            "INT_FreeRewardValue",
            "INT_FreeRewardSubType1",
            "INT_FreeRewardValue1",
            "INT_FreeRewardSubType2",
            "INT_FreeRewardValue2",
        ]
    )
    sheet.append([1012, 1, 10, 100, 11, 101, 12, 102])
    sheet.append([1010, 1, 10, 100, 11, 101, 12, 102])
    workbook.save(path)
    return path


async def _seed_quests_source_only_config(project_id: int, user_id: int, workbook_path: Path) -> None:
    config = {
        "version": 1,
        "configured": True,
        "sources": [
            {
                "id": "quests",
                "type": "local_excel",
                "pathOrUrl": str(workbook_path),
            }
        ],
        "variables": [],
        "groups": [
            {
                "group_id": "ungrouped",
                "group_name": "默认规则组",
                "builtin": True,
            }
        ],
        "rules": [],
        "local_path_replacement_presets": [],
        "svn_path_replacement_presets": [],
    }
    async with async_session_factory() as session:
        session.add(
            WorkbenchConfigRecord(
                project_id=project_id,
                user_id=user_id,
                config_json=json.dumps(config, ensure_ascii=False),
            )
        )
        await session.commit()


async def _seed_battlepass_source_only_config(project_id: int, user_id: int, workbook_path: Path) -> None:
    config = {
        "version": 1,
        "configured": True,
        "sources": [
            {
                "id": "battlepass",
                "type": "local_excel",
                "pathOrUrl": str(workbook_path),
            }
        ],
        "variables": [],
        "groups": [
            {
                "group_id": "ungrouped",
                "group_name": "默认规则组",
                "builtin": True,
            }
        ],
        "rules": [],
        "local_path_replacement_presets": [],
        "svn_path_replacement_presets": [],
    }
    async with async_session_factory() as session:
        session.add(
            WorkbenchConfigRecord(
                project_id=project_id,
                user_id=user_id,
                config_json=json.dumps(config, ensure_ascii=False),
            )
        )
        await session.commit()


async def _seed_bpcherks_composite_variable_config(project_id: int, user_id: int) -> None:
    config = {
        "version": 1,
        "configured": True,
        "sources": [
            {
                "id": "battlepass",
                "type": "svn",
                "pathOrUrl": "https://samosvn/data/project/samo/GameDatas/datas_qa88/battlepass.xls",
            }
        ],
        "variables": [
            {
                "tag": "bpcherks",
                "source_id": "battlepass",
                "sheet": "level_reward",
                "variable_kind": "composite",
                "columns": [
                    "INT_Index",
                    "INT_Level",
                    "INT_FreeRewardSubType",
                    "INT_FreeRewardValue",
                    "INT_PayRewardSubType1",
                    "INT_PayRewardValue1",
                    "INT_PayRewardSubType2",
                    "INT_PayRewardValue2",
                    "STR_SeasonCond",
                ],
                "key_column": "STR_SeasonCond",
                "append_index_to_key": True,
                "expected_type": "json",
            }
        ],
        "groups": [
            {
                "group_id": "ungrouped",
                "group_name": "默认规则组",
                "builtin": True,
            }
        ],
        "rules": [],
        "local_path_replacement_presets": [],
        "svn_path_replacement_presets": [],
    }
    async with async_session_factory() as session:
        session.add(
            WorkbenchConfigRecord(
                project_id=project_id,
                user_id=user_id,
                config_json=json.dumps(config, ensure_ascii=False),
            )
        )
        await session.commit()


async def _save_provider(auth_client: AsyncClient, api_key: str = "sk-test-secret") -> None:
    response = await auth_client.put(
        "/api/v1/ai/providers/me",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
            "api_key": api_key,
            "extra_headers": {"X-Test": "1"},
        },
    )
    assert response.status_code == 200, response.text


@pytest.mark.anyio
async def test_provider_config_is_encrypted_and_never_returns_plain_key(
    auth_client: AsyncClient,
    test_db,
) -> None:
    """保存模型配置后只返回脱敏 Key，数据库中保存密文。"""
    await _save_provider(auth_client)

    get_response = await auth_client.get("/api/v1/ai/providers/me")
    assert get_response.status_code == 200
    data = get_response.json()["data"]
    assert data["api_key_masked"].startswith("sk-")
    assert data["api_key_masked"].endswith("cret")
    assert "sk-test-secret" not in json.dumps(get_response.json(), ensure_ascii=False)

    user_id = await _get_test_user_id()
    async with async_session_factory() as session:
        record = await session.scalar(
            select(AiProviderCredentialRecord).where(
                AiProviderCredentialRecord.user_id == user_id
            )
        )
        assert record is not None
        assert record.encrypted_api_key != "sk-test-secret"
        assert decrypt_secret(record.encrypted_api_key) == "sk-test-secret"


@pytest.mark.anyio
async def test_provider_test_connection_returns_category_on_failure(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
) -> None:
    """连通性测试失败时返回供应商错误分类，便于前端给出低配置难度提示。"""

    async def fake_test_provider_connection(**_: Any) -> int:
        raise ProviderConnectionError(status_code=401, category="auth", message="API Key 无效")

    monkeypatch.setattr(
        "backend.app.api.ai_api.test_provider_connection",
        fake_test_provider_connection,
    )

    response = await auth_client.post(
        "/api/v1/ai/providers/test",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
            "api_key": "bad-key",
        },
    )
    assert response.status_code == 401
    assert response.json()["detail"]["category"] == "auth"


@pytest.mark.anyio
async def test_provider_test_requires_key_when_no_saved_config(
    auth_client: AsyncClient,
    test_db,
) -> None:
    """首次未保存配置时，测试连接仍需要用户输入 API Key。"""
    response = await auth_client.post(
        "/api/v1/ai/providers/test",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
        },
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "请填写 API Key 或重新保存 AI 配置。"


@pytest.mark.anyio
async def test_provider_test_reuses_saved_key_when_payload_key_is_empty(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
) -> None:
    """保存后测试连接允许留空 API Key，并复用后端密文。"""
    await _save_provider(auth_client)
    captured: dict[str, Any] = {}

    async def fake_test_provider_connection(**kwargs: Any) -> int:
        captured.update(kwargs)
        return 12

    monkeypatch.setattr(
        "backend.app.api.ai_api.test_provider_connection",
        fake_test_provider_connection,
    )

    response = await auth_client.post(
        "/api/v1/ai/providers/test",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
            "api_key": None,
        },
    )
    assert response.status_code == 200, response.text
    assert response.json()["data"]["latency_ms"] == 12
    assert captured["api_key"] == "sk-test-secret"


@pytest.mark.anyio
async def test_provider_save_with_empty_key_keeps_existing_cipher(
    auth_client: AsyncClient,
    test_db,
) -> None:
    """已保存配置再次保存且 API Key 留空时，不清空旧密文。"""
    await _save_provider(auth_client)
    response = await auth_client.put(
        "/api/v1/ai/providers/me",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
            "api_key": None,
            "extra_headers": {},
        },
    )
    assert response.status_code == 200, response.text

    user_id = await _get_test_user_id()
    async with async_session_factory() as session:
        record = await session.scalar(
            select(AiProviderCredentialRecord).where(
                AiProviderCredentialRecord.user_id == user_id
            )
        )
        assert record is not None
        assert decrypt_secret(record.encrypted_api_key) == "sk-test-secret"


@pytest.mark.anyio
async def test_provider_test_with_broken_saved_cipher_returns_clear_error(
    auth_client: AsyncClient,
    test_db,
) -> None:
    """已保存密文损坏时，测试连接提示重新填写而不是误报空 Key。"""
    await _save_provider(auth_client)
    user_id = await _get_test_user_id()
    async with async_session_factory() as session:
        record = await session.scalar(
            select(AiProviderCredentialRecord).where(
                AiProviderCredentialRecord.user_id == user_id
            )
        )
        assert record is not None
        record.encrypted_api_key = "not-a-valid-fernet-token"
        await session.commit()

    response = await auth_client.post(
        "/api/v1/ai/providers/test",
        json={
            "provider_preset": "openai",
            "base_url": "https://api.openai.com/v1",
            "model": "gpt-5.4-mini",
        },
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "请填写 API Key 或重新保存 AI 配置。"


@pytest.mark.anyio
async def test_rule_prompt_optimize_needs_raw_description(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """优化输入为空时返回 needs_input，且不调用模型。"""

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("空输入不应调用模型")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={"raw_description": "   ", "selected_variable_tags": ["test"]},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "needs_input"
    assert data["missing"] == ["请先输入规则描述。"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_needs_selected_variables(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """未选择目标变量时返回 needs_input，且不调用模型。"""

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("未选变量不应调用模型")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={"raw_description": "ID 不能为空", "selected_variable_tags": []},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "needs_input"
    assert data["missing"] == ["请先选择一个或多个目标变量。"]
    assert data["fallback"] is True


@pytest.mark.anyio
async def test_rule_prompt_optimize_auto_complete_allows_empty_variables(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """允许自动补齐时，未选择目标变量也不应在前置校验阶段拦截。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)
    captured: dict[str, Any] = {}

    async def fake_call_provider_json(**kwargs: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        captured.update(kwargs)
        return {
            "status": "optimized",
            "raw_description": "server_config switch STR_ServersParam 格式校验",
            "optimized_description": "规则目标：\n根据描述自动补齐数据源、Sheet、字段和变量后再进行 AI 校验。\n\n需要用户确认：\n请确认数据源、Sheet 和目标字段。",
            "detected_clues": {
                "rule_type_hint": "regex_check",
                "involved_variables": [],
                "target_field": "STR_ServersParam",
                "key_field": None,
                "filters": [],
                "compare_fields": [],
                "compare_operator": None,
            },
            "missing": ["请确认数据源、Sheet 和目标字段。"],
            "warnings": ["未选择目标变量，后续 AI 校验会尝试自动补齐。"],
            "confidence": 0.61,
            "fallback": False,
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "server_config switch STR_ServersParam 格式校验",
            "selected_variable_tags": [],
            "allow_auto_complete": True,
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "optimized"
    assert data["detected_clues"]["target_field"] == "STR_ServersParam"
    assert "allow_auto_complete" in captured["user_prompt"]
    assert "optimized_description 必须使用解析友好的 DSL" in captured["system_prompt"]
    assert "聚合、平均值、求和" in captured["system_prompt"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_fills_fixed_template_fields_from_input(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型返回固定模板但配置线索为空时，应优先用用户固定模板输入补齐。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "status": "optimized",
            "raw_description": "Quest STR_ABSwitch 不能为空",
            "optimized_description": "\n".join(
                [
                    "数据源：需要用户确认：配置表链接或已有数据源标识",
                    "sheet分页：需要用户确认：sheet名",
                    "变量选择：需要用户确认：变量1,变量2",
                    "",
                    "筛选规则1：无",
                    "",
                    "筛选规则2：无",
                    "",
                    "校验规则：规则类型：not_null；断言：STR_ABSwitch 不能为空",
                ]
            ),
            "detected_clues": {
                "rule_type_hint": "not_null",
                "involved_variables": [],
                "target_field": "STR_ABSwitch",
                "key_field": None,
                "filters": [],
                "compare_fields": [],
                "compare_operator": None,
            },
            "missing": [],
            "warnings": [],
            "confidence": 0.91,
            "fallback": False,
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "\n".join(
                [
                    "数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls",
                    "sheet分页：Quest",
                    "变量选择：STR_ABSwitch",
                    "",
                    "筛选规则1：",
                    "",
                    "筛选规则2：",
                    "",
                    "校验规则：STR_ABSwitch 不能为空",
                ]
            ),
            "selected_variable_tags": [],
            "allow_auto_complete": True,
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "optimized"
    assert not data["optimized_description"].startswith("数据源：")
    assert "sheet分页：Quest" not in data["optimized_description"]
    assert "变量选择：STR_ABSwitch" not in data["optimized_description"]
    assert data["optimized_description"].startswith("not_null")
    assert "目标：STR_ABSwitch" in data["optimized_description"]
    assert "断言：STR_ABSwitch 不能为空" in data["optimized_description"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_scattered_model_text_becomes_parser_friendly_template(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型仍返回散段落时，后端要归一成带规则类型短语的固定模板。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "status": "optimized",
            "raw_description": "ID 不能为空",
            "optimized_description": "规则目标：校验 ID 字段不能为空，请后续 AI 校验。",
            "detected_clues": {
                "rule_type_hint": "not_null",
                "involved_variables": ["[src_demo-items-ID]"],
                "target_field": "ID",
                "key_field": None,
                "filters": [],
                "compare_fields": [],
                "compare_operator": None,
            },
            "missing": [],
            "warnings": [],
            "confidence": 0.84,
            "fallback": False,
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "ID 不能为空",
            "selected_variable_tags": ["[src_demo-items-ID]"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "optimized"
    optimized_description = data["optimized_description"]
    assert optimized_description.startswith("not_null")
    assert "变量选择：ID" not in optimized_description
    assert "目标：ID" in optimized_description
    assert "断言：ID 不能为空" in optimized_description


@pytest.mark.anyio
async def test_rule_prompt_optimize_without_provider_returns_failed_fallback(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """AI 未配置时优化接口不抛 500，返回兜底优化结果。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "ID 不能为空",
            "selected_variable_tags": ["[src_demo-items-ID]"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "failed"
    assert data["fallback"] is True
    assert "AI 模型未配置" in " ".join(data["warnings"])
    assert "ID 不能为空" in data["optimized_description"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_dry_run_extracts_local_clues_without_provider(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """dry_run 只做本地线索抽取，不要求模型配置。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize?dry_run=true",
        json={
            "raw_description": "筛选：INT_ID唯一，INT_Faction!=0，Key值选择：INT_ID，判定：INT_Group必须重复",
            "selected_variable_tags": [],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "optimized"
    assert data["fallback"] is True
    assert data["optimized_description"] == ""
    assert data["detected_clues"]["rule_type_hint"] == "composite_condition_check"
    assert data["detected_clues"]["target_field"] == "INT_Group"
    assert data["detected_clues"]["key_field"] == "INT_ID"


@pytest.mark.anyio
async def test_rule_prompt_optimize_short_template_fallback_outputs_dsl(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """模型不可用时，短模板也应兜底优化成组合分支 DSL。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "筛选：INT_ID唯一，INT_Faction!=0，Key值选择：INT_ID，判定：INT_Group必须重复",
            "selected_variable_tags": [],
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "failed"
    assert data["fallback"] is True
    assert data["detected_clues"]["rule_type_hint"] == "composite_condition_check"
    assert data["detected_clues"]["key_field"] == "INT_ID"
    assert data["detected_clues"]["filters"] == [{"field": "INT_Faction", "operator": "ne", "value": "0", "side": "global"}]
    assert data["optimized_description"].startswith("composite_condition_check")
    assert "- INT_Faction != 0" in data["optimized_description"]
    assert "- INT_ID 唯一" in data["optimized_description"]
    assert "Key：INT_ID" in data["optimized_description"]
    assert "断言：INT_Group 必须重复" in data["optimized_description"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_bpcherks_fallback_recognizes_dual_compare(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """模型不可用时，bpcherks 的逗号双筛选描述也应优化成跨组变量校验。"""
    user_id = await _get_test_user_id()
    await _seed_bpcherks_composite_variable_config(test_project_id, user_id)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": BPCHERKS_COMMA_DUAL_DESCRIPTION,
            "selected_variable_tags": ["bpcherks"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "failed"
    assert data["fallback"] is True
    assert data["detected_clues"]["rule_type_hint"] == "dual_composite_compare"
    assert data["detected_clues"]["key_field"] == "INT_Level"
    assert data["detected_clues"]["compare_fields"] == BPCHERKS_COMPARE_FIELDS
    assert data["detected_clues"]["filters"] == [
        {"field": "INT_Index", "operator": "eq", "value": "1012", "side": "left"},
        {"field": "INT_Index", "operator": "eq", "value": "1010", "side": "right"},
    ]
    assert "左侧筛选：INT_Index = 1012" in data["optimized_description"]
    assert "右侧筛选：INT_Index = 1010" in data["optimized_description"]
    assert "断言：左右两组按 Key 对齐后比较字段必须相等" in data["optimized_description"]
    assert "INT_PayRewardSubType1" in data["optimized_description"]
    assert "已根据变量池字段将 INT_FreeRewardSubType1 修正为 INT_PayRewardSubType1" in " ".join(
        data["warnings"]
    )


@pytest.mark.anyio
async def test_rule_prompt_optimize_short_template_bpcherks_recognizes_dual_compare(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """短模板的集合筛选 + Key + 多字段相等应优化为双组 Key 对比 DSL。"""
    user_id = await _get_test_user_id()
    await _seed_bpcherks_composite_variable_config(test_project_id, user_id)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": BPCHERKS_SHORT_TEMPLATE_DUAL_DESCRIPTION,
            "selected_variable_tags": ["bpcherks"],
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "failed"
    assert data["fallback"] is True
    assert data["detected_clues"]["rule_type_hint"] == "dual_composite_compare"
    assert data["detected_clues"]["key_field"] == "INT_Level"
    assert data["detected_clues"]["compare_fields"] == BPCHERKS_COMPARE_FIELDS
    assert data["detected_clues"]["filters"] == [
        {"field": "INT_Index", "operator": "eq", "value": "1012", "side": "left"},
        {"field": "INT_Index", "operator": "eq", "value": "1010", "side": "right"},
    ]
    assert data["optimized_description"].startswith("dual_composite_compare")
    assert "左侧筛选：INT_Index = 1012" in data["optimized_description"]
    assert "右侧筛选：INT_Index = 1010" in data["optimized_description"]
    assert "断言：左右两组按 Key 对齐后比较字段必须相等" in data["optimized_description"]


@pytest.mark.anyio
async def test_rule_prompt_optimize_model_response_and_no_side_effects(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """正常优化只返回文本和线索，不修改个人配置，也不创建草稿历史。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)
    captured: dict[str, Any] = {}

    async def fake_call_provider_json(**kwargs: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        captured.update(kwargs)
        return {
            "status": "optimized",
            "raw_description": "ID 不能为空",
            "optimized_description": "规则目标：\n校验 ID 字段不能为空。\n\n目标变量：\n[src_demo-items-ID]\n\n期望规则类型：\nnot_null",
            "detected_clues": {
                "rule_type_hint": "not_null",
                "involved_variables": ["[src_demo-items-ID]"],
                "target_field": "ID",
                "key_field": None,
                "filters": [],
                "compare_fields": [],
                "compare_operator": None,
            },
            "missing": [],
            "warnings": [],
            "confidence": 0.88,
            "fallback": False,
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    async with async_session_factory() as session:
        before_config = await session.scalar(
            select(WorkbenchConfigRecord.config_json).where(
                WorkbenchConfigRecord.project_id == test_project_id,
                WorkbenchConfigRecord.user_id == user_id,
            )
        )
        before_drafts = (
            await session.execute(
                select(AiRuleDraftRecord).where(
                    AiRuleDraftRecord.project_id == test_project_id,
                    AiRuleDraftRecord.user_id == user_id,
                )
            )
        ).scalars().all()

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "ID 不能为空",
            "selected_variable_tags": ["[src_demo-items-ID]"],
            "context": {"page": "personal_workbench", "mode": "smart_rule"},
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "optimized"
    assert data["optimized_description"].startswith("not_null")
    assert "目标：ID" in data["optimized_description"]
    assert "断言：ID 不能为空" in data["optimized_description"]
    assert data["detected_clues"]["rule_type_hint"] == "not_null"
    assert data["confidence"] == 0.88
    assert "minimal_rules" not in captured["user_prompt"]
    assert "rows" not in captured["user_prompt"].lower()

    async with async_session_factory() as session:
        after_config = await session.scalar(
            select(WorkbenchConfigRecord.config_json).where(
                WorkbenchConfigRecord.project_id == test_project_id,
                WorkbenchConfigRecord.user_id == user_id,
            )
        )
        after_drafts = (
            await session.execute(
                select(AiRuleDraftRecord).where(
                    AiRuleDraftRecord.project_id == test_project_id,
                    AiRuleDraftRecord.user_id == user_id,
                )
            )
        ).scalars().all()
    assert after_config == before_config
    assert len(after_drafts) == len(before_drafts)


@pytest.mark.anyio
async def test_rule_prompt_optimize_invalid_model_response_returns_fallback(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型输出非法时返回 failed + fallback，而不是抛出 500。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "status": "optimized",
            "raw_description": "ID 不能为空",
            "optimized_description": "规则目标：校验 ID 非空",
            "detected_clues": {"unknown": "field"},
            "missing": [],
            "warnings": [],
            "fallback": False,
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-prompt-optimize",
        json={
            "raw_description": "ID 不能为空",
            "selected_variable_tags": ["[src_demo-items-ID]"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["status"] == "failed"
    assert data["fallback"] is True
    assert "模型返回格式不符合协议" in " ".join(data["warnings"])


@pytest.mark.anyio
async def test_rule_draft_ready_compiles_to_existing_not_null_rule(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """ready 草稿由模型意图确定性编译为个人校验可保存的规则配置。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "not_null",
            "confidence": 0.93,
            "reasoning_summary": "items 表 ID 字段可用现有变量表达非空规则。",
            "rule_name": "items-ID-非空",
            "target": {
                "tag": "[src_demo-items-ID]",
                "source_id": "src_demo",
                "sheet": "items",
                "variable_kind": "single",
                "column": "ID",
                "expected_type": "str",
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "items 表 ID 不能为空"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "not_null"
    assert data["draft"]["reuse_variable_tags"] == ["[src_demo-items-ID]"]
    assert data["draft"]["rules_to_add"][0]["rule_type"] == "not_null"
    assert "sk-test-secret" not in json.dumps(data, ensure_ascii=False)


@pytest.mark.anyio
async def test_rule_draft_complete_template_prefers_deterministic_compile(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """固定模板线索完整时应先确定性生成草稿，避免模型输出波动。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整固定模板不应先调用模型")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": """数据源：src_demo
sheet分页：items
变量选择：ID

筛选规则1：

筛选规则2：

校验规则：ID 不能为空""",
            "input_mode": "template",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "src_demo",
                "sheet": "items",
                "target_field": "ID",
                "rule_type_hint": "not_null",
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "not_null"
    assert data["draft"]["rules_to_add"][0]["target_variable_tag"] == "[src_demo-items-ID]"
    assert "结构化线索" in data["reasoning_summary"]


@pytest.mark.anyio
async def test_rule_draft_candidate_critique_blocks_conflicting_assertions_before_model(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """候选批判发现同一规则混合多个最终断言时，不应继续调用模型误添加。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("candidate critique should stop before provider call")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "\n".join(
                [
                    "数据源：demo.xls",
                    "sheet分页：items",
                    "变量选择：Status",
                    "",
                    "校验规则：Status 只能是 A,B 且不能为空",
                ]
            ),
            "allow_auto_complete": True,
            "selected_variable_tags": [],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["rules_to_add"] == []
    assert "多个最终断言" in data["missing"][0]["message"]


@pytest.mark.anyio
async def test_rule_draft_candidate_hint_overrides_wrong_model_rule_type(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型返回相近但错误的类型时，候选收窄后的 workflow_hints 应覆盖为更可信规则。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "unique",
            "confidence": 0.91,
            "reasoning_summary": "错误地判断成唯一校验。",
            "target": {"tag": "[src_demo-items-ID]"},
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "ID 不能为空",
            "selected_variable_tags": ["[src_demo-items-ID]"],
            "allow_auto_complete": False,
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "not_null"
    assert data["draft"]["rules_to_add"][0]["rule_type"] == "not_null"


@pytest.mark.anyio
async def test_rule_draft_no_auto_complete_reuses_selected_variable_only(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """关闭自动补齐时，ready 草稿只能复用用户选择的变量池变量。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "not_null",
            "confidence": 0.9,
            "reasoning_summary": "复用已选 ID 变量生成非空规则。",
            "target": {"tag": "[src_demo-items-ID]"},
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "ID 不能为空",
            "input_mode": "structured",
            "allow_auto_complete": False,
            "selected_variable_tags": ["[src_demo-items-ID]"],
            "workflow_hints": {
                "rule_type_hint": "not_null",
                "target_variable_tag": "[src_demo-items-ID]",
                "source_id": "quests",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls",
                "sheet": "Quest",
                "target_field": "STR_ABSwitch",
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["[src_demo-items-ID]"]
    assert data["draft"]["rules_to_add"][0]["target_variable_tag"] == "[src_demo-items-ID]"


@pytest.mark.anyio
async def test_rule_draft_no_auto_complete_requires_selected_variable(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """关闭自动补齐且未选择变量时，不生成可添加的半成品规则。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "not_null",
            "confidence": 0.9,
            "reasoning_summary": "模型识别到 ID 非空。",
            "target": {"tag": "[src_demo-items-ID]"},
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "ID 不能为空",
            "input_mode": "structured",
            "allow_auto_complete": False,
            "selected_variable_tags": [],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["rules_to_add"] == []
    assert data["missing"][0]["kind"] == "variable"


@pytest.mark.anyio
@pytest.mark.parametrize(
    "description",
    [
        (
            "校验规则筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
            "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
        ),
        (
            "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
            "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
        ),
    ],
)
async def test_rule_draft_lightweight_selected_composite_variable_generates_set_conditions(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
    description: str,
) -> None:
    """轻量智能规则只选择变量 + 描述时，也能生成可执行组合分支规则。"""
    user_id = await _get_test_user_id()
    await _seed_test_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "confidence": 0.1,
            "reasoning_summary": "等待后端基于描述兜底解析。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": description,
            "input_mode": "free_text",
            "allow_auto_complete": False,
            "selected_variable_tags": ["test"],
            "workflow_hints": {
                "target_variable_tag": "test",
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["description"] == description
    assert data["rule_type"] == "composite_condition_check"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["test"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "test"
    assert rule["rule_type"] == "composite_condition_check"

    composite_config = rule["composite_config"]
    desc_filter = composite_config["global_filters"][0]
    assert desc_filter["field"] == "DESC3"
    assert desc_filter["operator"] == "eq"
    assert desc_filter["expected_value_mode"] == "set"
    assert desc_filter["expected_value"] == "升级p1建筑到p2级p4次,完成p2等级的p1科研"

    assertion = composite_config["branches"][0]["assertions"][0]
    assert assertion["field"] == "STR_ABSwitch  "
    assert assertion["operator"] == "eq"
    assert assertion["expected_value_mode"] == "set"
    assert assertion["expected_value"] == "GreenServer:0,SLG2:0"


@pytest.mark.anyio
async def test_rule_draft_bpcherks_comma_dual_filter_reuses_selected_variable(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """bpcherks 的 FIELD=V1,V2 + Key + 多字段相等描述应兜底生成跨组变量校验。"""
    user_id = await _get_test_user_id()
    await _seed_bpcherks_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise ProviderConnectionError(status_code=504, category="timeout", message="调用大模型超时，请稍后重试。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": BPCHERKS_COMMA_DUAL_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": False,
            "selected_variable_tags": ["bpcherks"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "dual_composite_compare"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["bpcherks"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "bpcherks"
    assert rule["reference_variable_tag"] == "bpcherks"
    assert rule["left_key_field"] == "INT_Level"
    assert rule["right_key_field"] == "INT_Level"
    assert rule["left_filters"][0]["field"] == "INT_Index"
    assert rule["left_filters"][0]["expected_value"] == "1012"
    assert rule["right_filters"][0]["field"] == "INT_Index"
    assert rule["right_filters"][0]["expected_value"] == "1010"
    assert [item["left_field"] for item in rule["comparisons"]] == BPCHERKS_COMPARE_FIELDS
    assert [item["right_field"] for item in rule["comparisons"]] == BPCHERKS_COMPARE_FIELDS


@pytest.mark.anyio
async def test_rule_draft_bpcherks_short_template_reuses_selected_variable(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """短模板输入应复用已选组合变量，直接生成双组 Key 对比规则。"""
    user_id = await _get_test_user_id()
    await _seed_bpcherks_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise ProviderConnectionError(status_code=504, category="timeout", message="调用大模型超时，请稍后重试。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": BPCHERKS_SHORT_TEMPLATE_DUAL_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": False,
            "selected_variable_tags": ["bpcherks"],
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "dual_composite_compare"
    assert data["draft"]["reuse_variable_tags"] == ["bpcherks"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "bpcherks"
    assert rule["reference_variable_tag"] == "bpcherks"
    assert rule["left_key_field"] == "INT_Level"
    assert rule["right_key_field"] == "INT_Level"
    assert rule["left_filters"][0]["field"] == "INT_Index"
    assert rule["left_filters"][0]["expected_value"] == "1012"
    assert rule["right_filters"][0]["field"] == "INT_Index"
    assert rule["right_filters"][0]["expected_value"] == "1010"
    assert [item["left_field"] for item in rule["comparisons"]] == BPCHERKS_COMPARE_FIELDS
    assert [item["right_field"] for item in rule["comparisons"]] == BPCHERKS_COMPARE_FIELDS


def test_hint_extractor_keeps_set_filter_and_fixed_value_out_of_dual_compare() -> None:
    """只有同时具备左右筛选、Key 和字段相等语义时才识别为跨组变量校验。"""
    composite_hints = extract_workflow_hints_from_text(
        "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
        "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
    )
    assert composite_hints.rule_type_hint == "composite_condition_check"
    assert composite_hints.filter_field == "DESC3"
    assert composite_hints.filter_value == "升级p1建筑到p2级p4次,完成p2等级的p1科研"
    assert composite_hints.left_filter_field is None
    assert composite_hints.right_filter_field is None

    fixed_hints = extract_workflow_hints_from_text("校验 Status 字段=0,1")
    assert fixed_hints.rule_type_hint == "fixed_value_compare"
    assert fixed_hints.left_filter_field is None
    assert fixed_hints.right_filter_field is None


def test_hint_extractor_optimized_description_ignores_meta_fixed_value() -> None:
    """优化提示词中的“更适合 AI 解析”不应污染业务比较值。"""
    hints = extract_workflow_hints_from_text(QUESTS_SWITCH_OPTIMIZED_DESCRIPTION)
    assert hints.rule_type_hint == "composite_condition_check"
    assert hints.source_id == "quests"
    assert hints.sheet == "Quest"
    assert hints.filter_field == "DESC3"
    assert hints.filter_value == "升级p1建筑到p2级p4次,完成p2等级的p1科研"
    assert hints.assertion_field == "STR_ABSwitch"
    assert hints.assertion_value == "GreenServer:0,SLG2:0"
    assert hints.key_column is None
    assert "Key" not in hints.composite_columns
    assert hints.expected_value is None


def test_hint_extractor_fixed_template_sections_are_structured() -> None:
    """固定模板里的三项配置和筛选/校验段应直接进入结构化线索。"""
    hints = extract_workflow_hints_from_text(
        """数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/battlepass.xls
sheet分页：level_reward
变量选择：INT_Level,INT_Index,INT_PayRewardValue

筛选规则1：INT_Index=1012

筛选规则2：INT_Index=1010

校验规则：以 INT_Level 为 Key，判断 INT_PayRewardValue 字段是否相等"""
    )
    assert hints.source_id == "battlepass"
    assert hints.source_url and hints.source_url.endswith("battlepass.xls")
    assert hints.sheet == "level_reward"
    assert hints.left_filter_field == "INT_Index"
    assert hints.left_filter_value == "1012"
    assert hints.right_filter_field == "INT_Index"
    assert hints.right_filter_value == "1010"
    assert hints.key_column == "INT_Level"
    assert hints.rule_type_hint == "dual_composite_compare"
    assert "INT_PayRewardValue" in hints.composite_columns


def test_hint_extractor_fixed_template_field_compare_uses_key_precondition() -> None:
    """固定模板不能串段；筛选唯一字段作为 Key，校验 A=B 作为字段对字段断言。"""
    hints = extract_workflow_hints_from_text(
        """数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls
sheet分页：Strategic_slg2
变量选择：INT_ID,INT_Faction,INT_Group

筛选规则1：INT_ID唯一

筛选规则2：INT_Faction=0

校验规则：INT_Group=INT_ID"""
    )

    assert hints.rule_type_hint == "composite_condition_check"
    assert hints.source_id == "quests"
    assert hints.sheet == "Strategic_slg2"
    assert hints.key_column == "INT_ID"
    assert hints.filter_field == "INT_Faction"
    assert hints.filter_value == "0"
    assert hints.assertion_field == "INT_Group"
    assert hints.assertion_operator == "eq"
    assert hints.assertion_value is None
    assert hints.assertion_value_source == "field"
    assert hints.assertion_expected_field == "INT_ID"
    assert hints.composite_columns == ["INT_ID", "INT_Group", "INT_Faction"]


def test_hint_extractor_v3_template_field_compare_uses_typed_slots() -> None:
    """v3 模板按固定槽位抽取规则类型、筛选、Key 和字段对字段断言。"""
    hints = extract_workflow_hints_from_text(
        """数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls
sheet分页：Strategic_slg2
变量选择：INT_ID,INT_Faction,INT_Group

规则类型：组合分支

目标字段：INT_Group
筛选条件：INT_Faction=0
Key字段：INT_ID
引用对象：无
比较字段：无

校验规则：INT_Group=INT_ID
规则参数：无"""
    )

    assert hints.rule_type_hint == "composite_condition_check"
    assert hints.source_id == "quests"
    assert hints.sheet == "Strategic_slg2"
    assert hints.target_field == "INT_Group"
    assert hints.filter_field == "INT_Faction"
    assert hints.filter_value == "0"
    assert hints.key_column == "INT_ID"
    assert hints.assertion_field == "INT_Group"
    assert hints.assertion_value_source == "field"
    assert hints.assertion_expected_field == "INT_ID"
    assert hints.composite_columns == ["INT_ID", "INT_Group", "INT_Faction"]


def test_hint_extractor_natural_template_field_compare_uses_key_and_filter() -> None:
    """自然句模板应抽取目标字段、筛选条件、Key 和字段对字段断言。"""
    hints = extract_workflow_hints_from_text(
        """我想检查INT_Group。

只检查满足 INT_Faction=0 的数据。

如果需要按同一条配置对齐，用INT_ID作为 Key；不需要就写“无”。

规则是：INT_Group 必须等于字段 INT_ID。

补充说明：无"""
    )

    assert hints.rule_type_hint == "composite_condition_check"
    assert hints.target_field == "INT_Group"
    assert hints.filter_field == "INT_Faction"
    assert hints.filter_value == "0"
    assert hints.key_column == "INT_ID"
    assert hints.assertion_field == "INT_Group"
    assert hints.assertion_operator == "eq"
    assert hints.assertion_value_source == "field"
    assert hints.assertion_expected_field == "INT_ID"
    assert hints.expected_value is None


def test_hint_extractor_short_template_uses_key_selection_and_judgement() -> None:
    """短模板应把 Key值选择 / 判定 识别为 Key 和最终断言。"""
    hints = extract_workflow_hints_from_text(
        "筛选：INT_ID唯一，INT_Faction!=0，Key值选择：INT_ID，判定：INT_Group必须重复"
    )

    assert hints.rule_type_hint == "composite_condition_check"
    assert hints.key_column == "INT_ID"
    assert hints.filter_field == "INT_Faction"
    assert hints.filter_operator == "ne"
    assert hints.filter_value == "0"
    assert hints.filters[0].field == "INT_Faction"
    assert hints.filters[0].operator == "ne"
    assert hints.filters[0].value == "0"
    assert hints.assertion_field == "INT_Group"
    assert hints.assertion_operator == "duplicate_required"
    assert hints.composite_columns == ["INT_ID", "INT_Group", "INT_Faction"]


@pytest.mark.parametrize(
    ("rule_type_text", "body", "expected_rule_type"),
    [
        ("非空 / not_null", "目标字段：ID\n筛选条件：无\nKey字段：无\n引用对象：无\n比较字段：无\n\n校验规则：ID 不能为空\n规则参数：无", "not_null"),
        ("唯一 / unique", "目标字段：ID\n筛选条件：无\nKey字段：无\n引用对象：无\n比较字段：无\n\n校验规则：ID 不能重复\n规则参数：无", "unique"),
        ("固定值比较 / fixed_value_compare", "目标字段：Status\n筛选条件：无\nKey字段：无\n引用对象：无\n比较字段：无\n\n校验规则：Status 只能是 0,1\n规则参数：期望值=0,1；期望值模式=set", "fixed_value_compare"),
        ("正则 / regex_check", "目标字段：Code\n筛选条件：无\nKey字段：无\n引用对象：无\n比较字段：无\n\n校验规则：Code 匹配正则\n规则参数：正则=^[A-Z]+$", "regex_check"),
        ("顺序 / sequence_order_check", "目标字段：Level\n筛选条件：无\nKey字段：无\n引用对象：无\n比较字段：无\n\n校验规则：Level 按升序连续\n规则参数：方向=升序；步长=1；起始=自动", "sequence_order_check"),
        ("引用存在 / cross_table_mapping", "目标字段：ItemID\n筛选条件：无\nKey字段：无\n引用对象：Item_ItemID\n比较字段：无\n\n校验规则：ItemID 必须存在于引用对象\n规则参数：引用对象=Item_ItemID", "cross_table_mapping"),
        ("组合分支 / composite_condition_check", "目标字段：Code\n筛选条件：Status=1\nKey字段：ID\n引用对象：无\n比较字段：无\n\n校验规则：Code 匹配正则\n规则参数：正则=^[A-Z]+$", "composite_condition_check"),
        ("跨组变量 / dual_composite_compare", "目标字段：无\n筛选条件：左侧 Type=1；右侧 Type=2\nKey字段：ID\n引用对象：无\n比较字段：Value\n\n校验规则：左右两组按 Key 对齐后比较字段必须相等\n规则参数：无", "dual_composite_compare"),
        ("多组串行 / multi_composite_pipeline_check", "目标字段：Code\n筛选条件：无\nKey字段：ID\n引用对象：无\n比较字段：无\n\n校验规则：按多组串行节点执行筛选和断言\n规则参数：节点1 -> 节点2", "multi_composite_pipeline_check"),
        ("多组映射 / multi_composite_mapping_check", "目标字段：Code\n筛选条件：无\nKey字段：ID\n引用对象：无\n比较字段：无\n\n校验规则：按多组映射节点独立筛选和判断\n规则参数：节点1：变量=A；筛选=Type=1", "multi_composite_mapping_check"),
        ("IAP礼包校验 / package_items_compare", "目标字段：无\n筛选条件：无\nKey字段：无\n引用对象：礼包配置变量\n比较字段：STR_Items\n\n校验规则：飞书礼包规划明细与 STR_Items 一致\n规则参数：无", "package_items_compare"),
    ],
)
def test_hint_extractor_v3_template_rule_type_aliases(
    rule_type_text: str,
    body: str,
    expected_rule_type: str,
) -> None:
    """v3 模板中的中英文规则类型别名应归一到当前 11 类规则。"""
    hints = extract_workflow_hints_from_text(
        f"""数据源：demo.xls
sheet分页：items
变量选择：ID,Status,Code,Type,Value

规则类型：{rule_type_text}

{body}"""
    )

    assert hints.rule_type_hint == expected_rule_type


@pytest.mark.anyio
async def test_rule_draft_auto_complete_quests_with_model_parameters_wrapper(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """开启自动补齐时，模型多返回 parameters 也应复用覆盖字段的已有组合变量。"""
    user_id = await _get_test_user_id()
    await _seed_test_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "composite_condition_check",
            "confidence": 0.74,
            "reasoning_summary": "模型把规则参数放进了非协议字段。",
            "parameters": {
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls",
                "sheet": "Quest",
                "filter_field": "DESC3",
                "assertion_field": "STR_ABSwitch",
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert "Extra inputs are not permitted" not in json.dumps(data, ensure_ascii=False)

    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["test"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "test"
    desc_filter = rule["composite_config"]["global_filters"][0]
    assert desc_filter["field"] == "DESC3"
    assert desc_filter["operator"] == "eq"
    assert desc_filter["expected_value_mode"] == "set"
    assert desc_filter["expected_value"] == "升级p1建筑到p2级p4次,完成p2等级的p1科研"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "STR_ABSwitch  "
    assert assertion["operator"] == "eq"
    assert assertion["expected_value_mode"] == "set"
    assert assertion["expected_value"] == "GreenServer:0,SLG2:0"


@pytest.mark.anyio
async def test_rule_draft_auto_complete_ignores_placeholder_key_and_trims_existing_variable(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型把 Key 当列名时，应忽略占位 Key、复用已有组合变量并回写真实字段名。"""
    user_id = await _get_test_user_id()
    await _seed_test_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "composite_condition_check",
            "confidence": 0.82,
            "reasoning_summary": "模型误把 Key 当成真实列。",
            "target": {
                "source_id": "quests",
                "sheet": "Quest",
                "variable_kind": "composite",
                "columns": ["Key", "DESC3", "STR_ABSwitch"],
                "key_column": "Key",
                "expected_type": "json",
            },
            "composite_config": {
                "global_filters": [
                    {
                        "condition_id": "ai-filter-desc3",
                        "field": "DESC3",
                        "operator": "eq",
                        "value_source": "literal",
                        "expected_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                        "expected_value_mode": "set",
                    }
                ],
                "branches": [
                    {
                        "branch_id": "ai-branch-switch",
                        "filters": [],
                        "assertions": [
                            {
                                "condition_id": "ai-assert-switch",
                                "field": "STR_ABSwitch",
                                "operator": "eq",
                                "value_source": "literal",
                                "expected_value": "GreenServer:0,SLG2:0",
                                "expected_value_mode": "set",
                            }
                        ],
                    }
                ],
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_OPTIMIZED_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests",
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "key_column": "Key",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["Key", "DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["test"]
    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "test"
    assert rule["composite_config"]["global_filters"][0]["field"] == "DESC3"
    assert rule["composite_config"]["branches"][0]["assertions"][0]["field"] == "STR_ABSwitch  "


@pytest.mark.anyio
async def test_rule_draft_auto_complete_new_variable_uses_trimmed_metadata_key(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    tmp_path: Path,
    test_db,
) -> None:
    """没有可复用变量时，元数据里的 INT_ID 和尾空格真实列名可补齐新组合变量。"""
    user_id = await _get_test_user_id()
    workbook_path = _create_quests_workbook(tmp_path / "quests.xlsx")
    await _seed_quests_source_only_config(test_project_id, user_id, workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.4,
            "reasoning_summary": "交给结构化线索自动补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests",
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "key_column": "Key",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["Key", "DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    variable = data["draft"]["variables_to_add"][0]
    assert variable["source_id"] == "quests"
    assert variable["sheet"] == "Quest"
    assert variable["key_column"] == "INT_ID"
    assert variable["columns"] == ["INT_ID", "DESC3", "STR_ABSwitch  "]
    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == variable["tag"]
    assert rule["composite_config"]["branches"][0]["assertions"][0]["field"] == "STR_ABSwitch  "


@pytest.mark.anyio
async def test_rule_draft_auto_complete_new_source_reads_temporary_metadata(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_db,
) -> None:
    """本次输入的新数据源未入库时，也应临时读取表头补齐数据源、变量和规则。"""
    workbook_path = _create_quests_workbook(tmp_path / "quests.xlsx")
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.42,
            "reasoning_summary": "由结构化线索补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests_new",
                "source_type": "local_excel",
                "source_url": str(workbook_path),
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    source = data["draft"]["sources_to_add"][0]
    assert source["id"] == "quests_new"
    assert source["pathOrUrl"] == str(workbook_path)

    variable = data["draft"]["variables_to_add"][0]
    assert variable["source_id"] == "quests_new"
    assert variable["sheet"] == "Quest"
    assert variable["key_column"] == "INT_ID"
    assert variable["columns"] == ["INT_ID", "DESC3", "STR_ABSwitch  "]
    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == variable["tag"]
    assert rule["composite_config"]["branches"][0]["assertions"][0]["field"] == "STR_ABSwitch  "


@pytest.mark.anyio
async def test_rule_draft_auto_complete_new_source_requires_metadata_key(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_db,
) -> None:
    """新数据源元数据没有常见 Key 且用户未指定 Key 时，不应兜底用业务字段生成规则。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quest"
    sheet.append(["DESC3", "STR_ABSwitch  "])
    workbook_path = tmp_path / "quests-no-key.xlsx"
    workbook.save(workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.4,
            "reasoning_summary": "缺少 Key。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests_no_key",
                "source_type": "local_excel",
                "source_url": str(workbook_path),
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["rules_to_add"] == []
    assert "Key" in "；".join(item["message"] for item in data["missing"])


@pytest.mark.anyio
async def test_rule_draft_auto_complete_new_source_reports_missing_field(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_db,
) -> None:
    """新数据源元数据可读但规则字段不存在时，应提示缺列而不是保存半成品。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quest"
    sheet.append(["INT_ID", "DESC3"])
    workbook_path = tmp_path / "quests-missing-field.xlsx"
    workbook.save(workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.4,
            "reasoning_summary": "由结构化线索补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests_missing_field",
                "source_type": "local_excel",
                "source_url": str(workbook_path),
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["rules_to_add"] == []
    assert "STR_ABSwitch" in "；".join(item["message"] for item in data["missing"])


@pytest.mark.anyio
async def test_rule_draft_auto_complete_needs_real_key_without_variable_or_metadata(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
) -> None:
    """没有已有变量和可读元数据时，占位 Key 不应兜底生成组合变量。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.4,
            "reasoning_summary": "缺少 Key。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    def fail_read_source_metadata(*_: Any, **__: Any) -> dict[str, Any]:
        raise FileNotFoundError("测试场景模拟元数据不可读。")

    monkeypatch.setattr("backend.app.ai.agent_service.read_source_metadata", fail_read_source_metadata)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": QUESTS_SWITCH_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "quests",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls",
                "sheet": "Quest",
                "rule_type_hint": "composite_condition_check",
                "key_column": "Key",
                "filter_field": "DESC3",
                "filter_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                "assertion_field": "STR_ABSwitch",
                "assertion_value": "GreenServer:0,SLG2:0",
                "composite_columns": ["Key", "DESC3", "STR_ABSwitch"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert "Key" in "；".join(item["message"] for item in data["missing"])


@pytest.mark.anyio
@pytest.mark.parametrize(
    ("description", "expected_missing"),
    [
        (
            "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
            "STR_ABSwitch字段=GreenServer:0 or SLG2:0。",
            "数据源",
        ),
        (
            "校验https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls配置表，"
            "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
            "STR_ABSwitch字段=GreenServer:0 or SLG2:0。",
            "Sheet",
        ),
    ],
)
async def test_rule_draft_auto_complete_requires_source_sheet_and_fields(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
    description: str,
    expected_missing: str,
) -> None:
    """自动补齐开启但缺少数据源路径或 Sheet 时，应提示补充而不是猜测新增配置。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "composite_condition_check",
            "confidence": 0.42,
            "reasoning_summary": "缺少自动补齐所需配置。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": description,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["rules_to_add"] == []
    combined_missing = "；".join(item["message"] for item in data["missing"])
    assert expected_missing in combined_missing
    assert "关闭自动补齐" in combined_missing or "关闭自动补齐" in data["reasoning_summary"]


@pytest.mark.anyio
async def test_rule_draft_normalizes_model_target_variable_tag_alias(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型把 target.tag 写成 target.variable_tag 时，后端应归一化后继续编译。"""
    user_id = await _get_test_user_id()
    await _seed_test_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "composite_condition_check",
            "confidence": 0.8,
            "reasoning_summary": "模型识别到组合分支校验。",
            "target": {"variable_tag": "test"},
            "composite_config": {
                "global_filters": [
                    {
                        "condition_id": "ai-condition-filter",
                        "field": "DESC3",
                        "operator": "eq",
                        "value_source": "literal",
                        "expected_value": "升级p1建筑到p2级p4次,完成p2等级的p1科研",
                        "expected_value_mode": "set",
                    }
                ],
                "branches": [
                    {
                        "branch_id": "ai-branch-switch",
                        "filters": [],
                        "assertions": [
                            {
                                "condition_id": "ai-condition-assert",
                                "field": "STR_ABSwitch  ",
                                "operator": "eq",
                                "value_source": "literal",
                                "expected_value": "GreenServer:0,SLG2:0",
                                "expected_value_mode": "set",
                            }
                        ],
                    }
                ],
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": (
                "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
                "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
            ),
            "input_mode": "free_text",
            "allow_auto_complete": False,
            "selected_variable_tags": ["test"],
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["draft"]["rules_to_add"][0]["target_variable_tag"] == "test"
    assert "Extra inputs are not permitted" not in json.dumps(data, ensure_ascii=False)


@pytest.mark.anyio
async def test_rule_draft_rejected_returns_extension_suggestions(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
) -> None:
    """规则库无法表达时，应返回面向扩展规则能力的建议。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "rejected",
            "confidence": 0.2,
            "reasoning_summary": "当前规则库不支持公式计算后比较。",
            "rejection_reason": "需要先把两列求和，再和第三列比较。",
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "把 A 列和 B 列求和后必须等于 C 列"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "rejected"
    assert data["extension_suggestions"]
    assert any("公式" in item or "表达式" in item for item in data["extension_suggestions"])


@pytest.mark.anyio
async def test_rule_draft_ready_compiles_composite_regex_workflow(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """组合分支正则草稿应保留过滤条件与正则断言，供前端预校验后应用。"""
    await _save_provider(auth_client)
    regex_pattern = r"^(?:(?:all|\d+(?:-\d+)?):[01](;(?:all|\d+(?:-\d+)?):[01])*)?$"

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "composite_condition_check",
            "confidence": 0.91,
            "reasoning_summary": "switch 表 STR_ServersParam 可通过组合分支筛选后执行正则校验。",
            "rule_name": "switch-STR_ServersParam-格式校验",
            "target": {
                "source_id": "server_config",
                "source_type": "svn",
                "path_or_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                "sheet": "switch",
                "variable_kind": "composite",
                "columns": ["INT_Id", "STR_Func", "STR_ServersParam", "DES"],
                "key_column": "INT_Id",
                "append_index_to_key": True,
                "expected_type": "json",
            },
            "composite_config": {
                "global_filters": [
                    {
                        "condition_id": "condition-switch-des-not-deprecated",
                        "field": "DES",
                        "operator": "not_contains",
                        "value_source": "literal",
                        "expected_value": "废弃",
                    }
                ],
                "branches": [
                    {
                        "branch_id": "branch-switch-servers-param-format",
                        "filters": [],
                        "assertions": [
                            {
                                "condition_id": "condition-switch-servers-param-regex",
                                "field": "STR_ServersParam",
                                "operator": "regex",
                                "value_source": "literal",
                                "expected_value": regex_pattern,
                            }
                        ],
                    }
                ],
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "server_config.xls 的 switch 页校验 STR_ServersParam 格式，并过滤 DES 包含废弃的行"
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["draft"]["sources_to_add"][0]["id"] == "server_config"
    variable = data["draft"]["variables_to_add"][0]
    assert variable["variable_kind"] == "composite"
    assert variable["columns"] == ["INT_Id", "STR_Func", "STR_ServersParam", "DES"]
    rule = data["draft"]["rules_to_add"][0]
    assert rule["rule_type"] == "composite_condition_check"
    assert rule["display_field"] is None
    assert rule["composite_config"]["global_filters"][0]["operator"] == "not_contains"
    assert rule["composite_config"]["global_filters"][0]["expected_value"] == "废弃"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["operator"] == "regex"
    assert assertion["expected_value"] == regex_pattern


@pytest.mark.anyio
async def test_rule_draft_workflow_hints_auto_complete_composite_rule(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型返回缺口时，完整结构化线索应自动补齐数据源、组合变量和规则。"""
    await _save_provider(auth_client)
    regex_pattern = r"^(?:(?:all|\d+(?:-\d+)?):[01](;(?:all|\d+(?:-\d+)?):[01])*)?$"

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "confidence": 0.52,
            "reasoning_summary": "缺少 server_config 数据源及 switch 分页的组合变量。",
            "missing": [
                {"kind": "source", "message": "缺少 server_config 数据源。"},
                {"kind": "variable", "message": "缺少 switch 组合变量。"},
                {"kind": "rule", "message": "缺少可用规则类型。"},
            ],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "读取 server_config.xls 的 switch 分页，校验 STR_ServersParam 字段配置格式，过滤 DES 包含废弃的行。",
            "workflow_hints": {
                "source_id": "server_config",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                "sheet": "switch",
                "target_field": "STR_ServersParam",
                "display_field": "STR_Func",
                "filter_field": "DES",
                "filter_value": "废弃",
                "regex_pattern": regex_pattern,
                "key_column": "INT_Id",
                "composite_columns": ["INT_Id", "STR_Func", "STR_ServersParam", "DES"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["missing"] == []

    source = data["draft"]["sources_to_add"][0]
    assert source["id"] == "server_config"
    assert source["type"] == "svn"
    assert source["pathOrUrl"].endswith("/server_config.xls")

    variable = data["draft"]["variables_to_add"][0]
    assert variable["source_id"] == "server_config"
    assert variable["sheet"] == "switch"
    assert variable["variable_kind"] == "composite"
    assert variable["key_column"] == "INT_Id"
    assert variable["append_index_to_key"] is True
    assert variable["columns"] == ["INT_Id", "STR_Func", "STR_ServersParam", "DES"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["rule_type"] == "composite_condition_check"
    assert rule["display_field"] == "STR_Func"
    assert rule["target_variable_tag"] == "[server_config-switch-INT_Id-mapping]"
    assert rule["composite_config"]["global_filters"][0]["operator"] == "not_contains"
    assert rule["composite_config"]["global_filters"][0]["field"] == "DES"
    assert rule["composite_config"]["global_filters"][0]["expected_value"] == "废弃"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "STR_ServersParam"
    assert assertion["operator"] == "regex"
    assert assertion["expected_value"] == regex_pattern


@pytest.mark.anyio
async def test_rule_draft_fixed_template_key_filter_field_compare_generates_field_assertion(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """FIELD唯一 + FIELD=VALUE + A=B 应生成组合分支字段对字段比较，不再提示规则不足。"""
    await _save_provider(auth_client)
    description = """数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls
sheet分页：Strategic_slg2
变量选择：INT_ID,INT_Faction,INT_Group

筛选规则1：INT_ID唯一

筛选规则2：INT_Faction=0

校验规则：INT_Group=INT_ID"""

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整固定模板应由确定性候选算法直接编译，不需要调用模型。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": description,
            "input_mode": "free_text",
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["missing"] == []

    variable = data["draft"]["variables_to_add"][0]
    assert variable["source_id"] == "quests"
    assert variable["sheet"] == "Strategic_slg2"
    assert variable["variable_kind"] == "composite"
    assert variable["key_column"] == "INT_ID"
    assert variable["columns"] == ["INT_ID", "INT_Group", "INT_Faction"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["rule_type"] == "composite_condition_check"
    global_filter = rule["composite_config"]["global_filters"][0]
    assert global_filter["field"] == "INT_Faction"
    assert global_filter["operator"] == "eq"
    assert global_filter["expected_value"] == "0"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "INT_Group"
    assert assertion["operator"] == "eq"
    assert assertion["value_source"] == "field"
    assert assertion["expected_field"] == "INT_ID"
    assert assertion["expected_value"] is None


@pytest.mark.anyio
async def test_rule_draft_user_regression_quests_set_filter_and_assertion(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_project_id: int,
    test_db,
) -> None:
    """用户回归 1：Quest 页多值筛选 + 多值断言应直接生成组合分支规则。"""
    user_id = await _get_test_user_id()
    workbook_path = _create_quests_workbook(tmp_path / "quests.xlsx")
    await _seed_quests_source_only_config(test_project_id, user_id, workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整 Quest 回归样例应由确定性链路直接编译。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": (
                "校验https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls配置表的Quest分页，"
                "筛选DESC3=升级p1建筑到p2级p4次,完成p2等级的p1科研，两种类型。"
                "STR_ABSwitch字段=GreenServer:0 or SLG2:0。"
            ),
            "input_mode": "free_text",
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    rule = data["draft"]["rules_to_add"][0]
    global_filter = rule["composite_config"]["global_filters"][0]
    assert global_filter["field"] == "DESC3"
    assert global_filter["operator"] == "eq"
    assert global_filter["expected_value_mode"] == "set"
    assert global_filter["expected_value"] == "升级p1建筑到p2级p4次,完成p2等级的p1科研"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "STR_ABSwitch  "
    assert assertion["operator"] == "eq"
    assert assertion["expected_value_mode"] == "set"
    assert assertion["expected_value"] == "GreenServer:0,SLG2:0"


@pytest.mark.anyio
async def test_rule_draft_user_regression_battlepass_dual_compare_six_fields(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_project_id: int,
    test_db,
) -> None:
    """用户回归 2：battlepass 两组筛选 + Key + 6 字段相等应生成跨组变量校验。"""
    user_id = await _get_test_user_id()
    workbook_path = _create_battlepass_user_regression_workbook(tmp_path / "battlepass.xlsx")
    await _seed_battlepass_source_only_config(test_project_id, user_id, workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整 battlepass 回归样例应由确定性链路直接编译。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": (
                "校验https://samosvn/data/project/samo/GameDatas/datas_qa88/battlepass.xls表的level_reward分页，"
                "筛选INT_Index=1012,1010,以INT_Level为key值，判断"
                "INT_FreeRewardSubType,INT_FreeRewardValue,INT_FreeRewardSubType1,INT_FreeRewardValue1,"
                "INT_FreeRewardSubType2,INT_FreeRewardValue2。6个字段值相等"
            ),
            "input_mode": "free_text",
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "dual_composite_compare"
    rule = data["draft"]["rules_to_add"][0]
    assert rule["left_key_field"] == "INT_Level"
    assert rule["right_key_field"] == "INT_Level"
    assert rule["left_filters"][0]["field"] == "INT_Index"
    assert rule["left_filters"][0]["expected_value"] == "1012"
    assert rule["right_filters"][0]["field"] == "INT_Index"
    assert rule["right_filters"][0]["expected_value"] == "1010"
    assert [item["left_field"] for item in rule["comparisons"]] == [
        "INT_FreeRewardSubType",
        "INT_FreeRewardValue",
        "INT_FreeRewardSubType1",
        "INT_FreeRewardValue1",
        "INT_FreeRewardSubType2",
        "INT_FreeRewardValue2",
    ]


@pytest.mark.anyio
async def test_rule_draft_user_regression_strategic_slg2_multi_filter_key_compare(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_project_id: int,
    test_db,
) -> None:
    """用户回归 3：INT_ID 唯一是 Key 前置条件，不生成最终唯一规则。"""
    user_id = await _get_test_user_id()
    workbook_path = _create_quests_workbook(tmp_path / "quests.xlsx")
    await _seed_quests_source_only_config(test_project_id, user_id, workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整 Strategic_slg2 回归样例应由确定性链路直接编译。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": (
                "校验https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls配置表的Strategic_slg2分页，"
                "INT_ID唯一，INT_Faction=0，INT_Group=INT_ID"
            ),
            "input_mode": "free_text",
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["draft"]["rules_to_add"][0]["rule_type"] == "composite_condition_check"
    variable = data["draft"]["variables_to_add"][0]
    assert variable["key_column"] == "INT_ID"
    rule = data["draft"]["rules_to_add"][0]
    assert rule["composite_config"]["global_filters"][0]["field"] == "INT_Faction"
    assert rule["composite_config"]["global_filters"][0]["expected_value"] == "0"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "INT_Group"
    assert assertion["value_source"] == "field"
    assert assertion["expected_field"] == "INT_ID"


@pytest.mark.anyio
async def test_rule_draft_short_template_duplicate_required_ready(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_project_id: int,
    test_db,
) -> None:
    """短模板输入应直接生成组合分支必须重复草稿。"""
    user_id = await _get_test_user_id()
    workbook_path = _create_quests_workbook(tmp_path / "quests.xlsx")
    await _seed_quests_source_only_config(test_project_id, user_id, workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise AssertionError("完整短模板应由确定性链路直接编译。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "\n".join(
                [
                    "数据源：https://samosvn/data/project/samo/GameDatas/datas_qa88/quests.xls",
                    "sheet分页：Strategic_slg2",
                    "变量选择：INT_ID,INT_Faction,INT_Group",
                    "",
                    "筛选：INT_ID唯一，INT_Faction!=0，Key值选择：INT_ID，判定：INT_Group必须重复",
                ]
            ),
            "input_mode": "template",
            "allow_auto_complete": True,
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    rule = data["draft"]["rules_to_add"][0]
    assert rule["composite_config"]["global_filters"][0]["field"] == "INT_Faction"
    assert rule["composite_config"]["global_filters"][0]["operator"] == "ne"
    assert rule["composite_config"]["global_filters"][0]["expected_value"] == "0"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "INT_Group"
    assert assertion["operator"] == "duplicate_required"
    assert assertion["expected_value"] is None


@pytest.mark.anyio
async def test_rule_draft_description_hints_auto_complete_wrapped_model_target(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """只输入自然语言时，后端应抽取变量线索并兼容 target.variable 包装字段。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "composite_condition_check",
            "confidence": 0.82,
            "reasoning_summary": "模型把目标变量多包了一层 variable。",
            "target": {
                "variable": {
                    "tag": "[server_config-switch-INT_Id-mapping]",
                }
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": SERVER_CONFIG_NATURAL_DESCRIPTION},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"

    source = data["draft"]["sources_to_add"][0]
    assert source["id"] == "server_config"
    assert source["type"] == "svn"
    assert source["pathOrUrl"].endswith("/server_config.xls")

    variable = data["draft"]["variables_to_add"][0]
    assert variable["sheet"] == "switch"
    assert variable["variable_kind"] == "composite"
    assert variable["key_column"] == "INT_Id"
    assert variable["append_index_to_key"] is True
    assert variable["columns"] == ["INT_Id", "STR_Func", "STR_ServersParam", "DES"]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["display_field"] == "STR_Func"
    assert rule["target_variable_tag"] == "[server_config-switch-INT_Id-mapping]"
    assert rule["composite_config"]["global_filters"][0]["field"] == "DES"
    assert rule["composite_config"]["global_filters"][0]["operator"] == "not_contains"
    assert rule["composite_config"]["global_filters"][0]["expected_value"] == "废弃"
    assertion = rule["composite_config"]["branches"][0]["assertions"][0]
    assert assertion["field"] == "STR_ServersParam"
    assert assertion["operator"] == "regex"
    assert assertion["expected_value"] == SERVER_CONFIG_PATTERN


@pytest.mark.anyio
async def test_rule_draft_description_hints_fallback_when_model_output_is_invalid(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型输出非法字段时，完整自然语言线索仍可兜底生成 ready 草稿。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "formula_calculate",
            "confidence": 0.8,
            "reasoning_summary": "模型误用了不存在的规则类型。",
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": SERVER_CONFIG_NATURAL_DESCRIPTION},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert data["draft"]["variables_to_add"][0]["columns"] == [
        "INT_Id",
        "STR_Func",
        "STR_ServersParam",
        "DES",
    ]
    assert "自然语言线索" in data["reasoning_summary"]


@pytest.mark.anyio
async def test_rule_draft_workflow_hints_do_not_override_rejected(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型明确判定现有能力不支持时，结构化线索不能强行转成可添加。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "rejected",
            "confidence": 0.2,
            "reasoning_summary": "当前系统不支持公式聚合计算。",
            "rejection_reason": "当前 11 类规则无法表达跨行聚合平均值判断。",
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "按服务器聚合分组后计算平均战力。",
            "workflow_hints": {
                "source_id": "server_config",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                "sheet": "switch",
                "target_field": "STR_ServersParam",
                "regex_pattern": ".*",
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "rejected"
    assert data["draft"]["rules_to_add"] == []
    assert "聚合" in data["rejection_reason"]
    assert "现有 11 类规则" in data["rejection_reason"]


@pytest.mark.anyio
async def test_rule_draft_workflow_hints_fallback_when_provider_times_out(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型超时时，完整结构化线索仍可生成可预校验草稿。"""
    await _save_provider(auth_client)
    regex_pattern = r"^(?:(?:all|\d+(?:-\d+)?):[01](;(?:all|\d+(?:-\d+)?):[01])*)?$"

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise ProviderConnectionError(status_code=504, category="timeout", message="调用大模型超时，请稍后重试。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "读取 server_config.xls 的 switch 分页，校验 STR_ServersParam 字段配置格式。",
            "workflow_hints": {
                "source_id": "server_config",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                "sheet": "switch",
                "target_field": "STR_ServersParam",
                "display_field": "STR_Func",
                "filter_field": "DES",
                "filter_value": "废弃",
                "regex_pattern": regex_pattern,
                "key_column": "INT_Id",
                "composite_columns": ["INT_Id", "STR_Func", "STR_ServersParam", "DES"],
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "composite_condition_check"
    assert "结构化线索" in data["reasoning_summary"]
    assert data["draft"]["rules_to_add"][0]["display_field"] == "STR_Func"


@pytest.mark.anyio
@pytest.mark.parametrize(
    ("rule_type", "workflow_hints"),
    [
        ("not_null", {"rule_type_hint": "not_null", "target_field": "ID"}),
        ("unique", {"rule_type_hint": "unique", "target_field": "ID"}),
        (
            "fixed_value_compare",
            {
                "rule_type_hint": "fixed_value_compare",
                "target_field": "Status",
                "operator": "eq",
                "expected_value": "1",
            },
        ),
        (
            "regex_check",
            {
                "rule_type_hint": "regex_check",
                "target_field": "Code",
                "regex_pattern": "^[A-Z]+$",
            },
        ),
        (
            "sequence_order_check",
            {
                "rule_type_hint": "sequence_order_check",
                "target_field": "Sort",
                "sequence_direction": "asc",
                "sequence_step": "1",
            },
        ),
        (
            "cross_table_mapping",
            {
                "rule_type_hint": "cross_table_mapping",
                "target_field": "ItemID",
                "reference_sheet": "dict",
                "reference_field": "ItemID",
            },
        ),
        (
            "composite_condition_check",
            {
                "rule_type_hint": "composite_condition_check",
                "target_field": "Code",
                "key_column": "ID",
                "composite_columns": ["ID", "Code", "Status"],
                "filter_field": "Status",
                "filter_operator": "eq",
                "filter_value": "1",
                "regex_pattern": "^[A-Z]+$",
            },
        ),
        (
            "dual_composite_compare",
            {
                "rule_type_hint": "dual_composite_compare",
                "key_column": "ID",
                "left_key_field": "ID",
                "right_key_field": "ID",
                "left_filter_field": "Type",
                "left_filter_operator": "eq",
                "left_filter_value": "1",
                "right_filter_field": "Type",
                "right_filter_operator": "eq",
                "right_filter_value": "2",
                "compare_fields": ["Value"],
                "composite_columns": ["ID", "Type", "Value"],
            },
        ),
        (
            "multi_composite_pipeline_check",
            {
                "rule_type_hint": "multi_composite_pipeline_check",
                "target_field": "Status",
                "key_column": "ID",
                "composite_columns": ["ID", "Status"],
                "filter_field": "Status",
                "filter_operator": "eq",
                "filter_value": "1",
                "assertion_field": "Status",
                "assertion_operator": "not_null",
            },
        ),
        (
            "multi_composite_mapping_check",
            {
                "rule_type_hint": "multi_composite_mapping_check",
                "target_field": "Status",
                "key_column": "ID",
                "composite_columns": ["ID", "Status"],
                "filter_field": "Status",
                "filter_operator": "eq",
                "filter_value": "1",
            },
        ),
    ],
)
async def test_rule_draft_workflow_hints_cover_all_existing_rule_types_when_provider_times_out(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    rule_type: str,
    workflow_hints: dict[str, Any],
    test_db,
) -> None:
    """模型不可用时，完整结构化线索仍能编译为当前标准 AI 规则。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise ProviderConnectionError(status_code=504, category="timeout", message="调用大模型超时，请稍后重试。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)
    payload_hints = {
        "source_id": "demo",
        "source_url": "https://samosvn/data/project/samo/GameDatas/demo.xls",
        "sheet": "items",
        **workflow_hints,
    }

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": f"items 表规则：{rule_type}",
            "workflow_hints": payload_hints,
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == rule_type
    assert data["draft"]["rules_to_add"][0]["rule_type"] == rule_type


@pytest.mark.anyio
async def test_rule_draft_battlepass_natural_description_compiles_dual_compare_when_provider_times_out(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_db,
) -> None:
    """battlepass 两组筛选按业务 Key 比较的自然语言应生成跨组变量校验。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        raise ProviderConnectionError(status_code=504, category="timeout", message="调用大模型超时，请稍后重试。")

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": BATTLEPASS_NATURAL_DESCRIPTION},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["rule_type"] == "dual_composite_compare"

    source = data["draft"]["sources_to_add"][0]
    assert source["id"] == "battlepass"
    assert source["type"] == "svn"
    assert source["pathOrUrl"].endswith("/battlepass.xls")

    variable = data["draft"]["variables_to_add"][0]
    assert variable["sheet"] == "level_reward"
    assert variable["variable_kind"] == "composite"
    assert variable["key_column"] == "INT_Level"
    assert variable["append_index_to_key"] is True
    assert variable["columns"] == [
        "INT_Level",
        "INT_Index",
        "INT_FreeRewardSubType",
        "INT_FreeRewardValue",
        "INT_PayRewardSubType1",
        "INT_PayRewardValue1",
    ]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["left_key_field"] == "INT_Level"
    assert rule["right_key_field"] == "INT_Level"
    assert rule["left_filters"][0]["field"] == "INT_Index"
    assert rule["left_filters"][0]["expected_value"] == "1011"
    assert rule["right_filters"][0]["field"] == "INT_Index"
    assert rule["right_filters"][0]["expected_value"] == "1010"
    assert [item["left_field"] for item in rule["comparisons"]] == [
        "INT_FreeRewardSubType",
        "INT_FreeRewardValue",
        "INT_PayRewardSubType1",
        "INT_PayRewardValue1",
    ]


@pytest.mark.anyio
async def test_rule_draft_auto_complete_new_source_fuzzy_corrects_battlepass_fields(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_db,
) -> None:
    """新增变量时，近似字段只有一个高置信表头候选才自动纠偏并生成规则。"""
    workbook_path = _create_battlepass_workbook(tmp_path / "battlepass.xlsx")
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "dual_composite_compare",
            "confidence": 0.42,
            "reasoning_summary": "由结构化线索补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": BPCHERKS_COMMA_DUAL_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "battlepass",
                "source_type": "local_excel",
                "source_url": str(workbook_path),
                "sheet": "level_reward",
                "rule_type_hint": "dual_composite_compare",
                "key_column": "INT_Level",
                "left_key_field": "INT_Level",
                "right_key_field": "INT_Level",
                "left_filter_field": "INT_Index",
                "left_filter_value": "1012",
                "right_filter_field": "INT_Index",
                "right_filter_value": "1010",
                "compare_fields": [
                    "INT_FreeRewardSubType",
                    "INT_FreeRewardValue",
                    "INT_FreeRewardSubType1",
                    "INT_FreeRewardValue1",
                    "INT_FreeRewardSubType2",
                    "INT_FreeRewardValue2",
                ],
            },
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert "修正为 INT_PayRewardSubType1" in data["reasoning_summary"]

    variable = data["draft"]["variables_to_add"][0]
    assert variable["source_id"] == "battlepass"
    assert variable["sheet"] == "level_reward"
    assert variable["key_column"] == "INT_Level"
    assert variable["columns"] == [
        "INT_Level",
        "INT_Index",
        "INT_FreeRewardSubType",
        "INT_FreeRewardValue",
        "INT_PayRewardSubType1",
        "INT_PayRewardValue1",
        "INT_PayRewardSubType2",
        "INT_PayRewardValue2",
    ]

    rule = data["draft"]["rules_to_add"][0]
    assert rule["rule_type"] == "dual_composite_compare"
    assert [item["left_field"] for item in rule["comparisons"]] == [
        "INT_FreeRewardSubType",
        "INT_FreeRewardValue",
        "INT_PayRewardSubType1",
        "INT_PayRewardValue1",
        "INT_PayRewardSubType2",
        "INT_PayRewardValue2",
    ]


@pytest.mark.anyio
async def test_rule_draft_auto_complete_reuses_existing_variable_after_fuzzy_correction(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """已有组合变量字段覆盖纠偏后的字段时，应复用变量而不是新增重复变量。"""
    user_id = await _get_test_user_id()
    await _seed_bpcherks_composite_variable_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "dual_composite_compare",
            "confidence": 0.4,
            "reasoning_summary": "由结构化线索补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": BPCHERKS_COMMA_DUAL_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "battlepass",
                "sheet": "level_reward",
                "rule_type_hint": "dual_composite_compare",
                "key_column": "INT_Level",
                "left_key_field": "INT_Level",
                "right_key_field": "INT_Level",
                "left_filter_field": "INT_Index",
                "left_filter_value": "1012",
                "right_filter_field": "INT_Index",
                "right_filter_value": "1010",
                "compare_fields": [
                    "INT_FreeRewardSubType",
                    "INT_FreeRewardValue",
                    "INT_FreeRewardSubType1",
                    "INT_FreeRewardValue1",
                    "INT_FreeRewardSubType2",
                    "INT_FreeRewardValue2",
                ],
            },
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "ready"
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["reuse_variable_tags"] == ["bpcherks"]
    rule = data["draft"]["rules_to_add"][0]
    assert rule["target_variable_tag"] == "bpcherks"
    assert rule["reference_variable_tag"] == "bpcherks"
    assert [item["left_field"] for item in rule["comparisons"]] == BPCHERKS_COMPARE_FIELDS


@pytest.mark.anyio
async def test_rule_draft_auto_complete_rejects_ambiguous_fuzzy_field(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    test_db,
) -> None:
    """近似字段存在多个高分候选时继续提示补充，不生成半成品。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "level_reward"
    sheet.append(["INT_Index", "INT_Level", "INT_PayRewardValue1", "INT_FreeRewardValue1"])
    workbook_path = tmp_path / "battlepass-ambiguous.xlsx"
    workbook.save(workbook_path)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "dual_composite_compare",
            "confidence": 0.4,
            "reasoning_summary": "由结构化线索补齐。",
            "missing": [],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": BPCHERKS_COMMA_DUAL_DESCRIPTION,
            "input_mode": "free_text",
            "allow_auto_complete": True,
            "selected_variable_tags": [],
            "workflow_hints": {
                "source_id": "battlepass",
                "source_type": "local_excel",
                "source_url": str(workbook_path),
                "sheet": "level_reward",
                "rule_type_hint": "dual_composite_compare",
                "key_column": "INT_Level",
                "left_filter_field": "INT_Index",
                "left_filter_value": "1012",
                "right_filter_field": "INT_Index",
                "right_filter_value": "1010",
                "compare_fields": ["INT_RewardValue1"],
            },
        },
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["rules_to_add"] == []


@pytest.mark.anyio
async def test_rule_draft_workflow_hints_still_need_rule_detail_when_incomplete(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """线索不足时不生成半成品配置，只提示修改输入后重试。"""
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "confidence": 0.4,
            "reasoning_summary": "缺少规则口径。",
            "missing": [{"kind": "rule", "message": "缺少可用规则类型。"}],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={
            "description": "检查 switch 表 STR_ServersParam。",
            "workflow_hints": {
                "source_id": "server_config",
                "source_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                "sheet": "switch",
                "target_field": "STR_ServersParam",
            },
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["draft"]["sources_to_add"] == []
    assert data["draft"]["variables_to_add"] == []
    assert data["draft"]["rules_to_add"] == []
    assert "修改上方规则描述" in data["reasoning_summary"]


@pytest.mark.anyio
async def test_rule_draft_needs_input_when_supported_rule_lacks_variable(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """支持的规则类型缺少 Sheet/列/变量时返回 needs_input 和可预填缺口。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "not_null",
            "confidence": 0.75,
            "reasoning_summary": "规则类型支持，但当前变量池没有 STR_ABSwitch。",
            "rule_name": "Quest-STR_ABSwitch-非空",
            "target": {
                "source_id": "src_demo",
                "sheet": "Quest",
                "variable_kind": "single",
                "column": "STR_ABSwitch",
                "expected_type": "str",
            },
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "Quest 表 STR_ABSwitch 不能为空"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["missing"][0]["suggested_action"] == "open_single_variable_dialog"
    assert data["missing"][0]["prefill"]["column"] == "STR_ABSwitch"


@pytest.mark.anyio
async def test_rule_draft_normalizes_missing_without_kind(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型 missing 缺少 kind 时应归一化为可展示缺口，而不是协议错误。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "needs_input",
            "rule_type": "not_null",
            "confidence": 0.61,
            "reasoning_summary": "规则可支持，但缺少目标字段信息。",
            "missing": [
                {
                    "message": "缺少 switch Sheet 的 STR_ServersParam 字段。",
                    "suggested_action": "edit_description",
                    "extra": "模型多返回的字段应被忽略",
                }
            ],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "switch 表 STR_ServersParam 需要格式校验"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "needs_input"
    assert data["missing"][0] == {
        "kind": "variable",
        "message": "缺少 switch Sheet 的 STR_ServersParam 字段。",
        "suggested_action": "edit_description",
        "prefill": {},
    }
    assert "字段不合法" not in json.dumps(data, ensure_ascii=False)


@pytest.mark.anyio
async def test_rule_draft_normalizes_missing_invalid_action(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型 missing 的 action 不合法时应兜底，不影响 rejected 原因展示。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "rejected",
            "confidence": 0.22,
            "reasoning_summary": "当前规则缺少可访问的数据源。",
            "rejection_reason": "请先补充 server_config 数据源。",
            "missing": [
                {
                    "kind": "source",
                    "message": "缺少 server_config SVN 数据源。",
                    "suggested_action": "open_dialog",
                    "prefill": {
                        "source_id": "server_config",
                        "path_or_url": "https://samosvn/data/project/samo/GameDatas/datas_qa88/server_config.xls",
                    },
                }
            ],
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "校验 server_config switch STR_ServersParam"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "rejected"
    assert data["rejection_reason"] == "请先补充 server_config 数据源。"
    assert data["missing"][0]["kind"] == "source"
    assert data["missing"][0]["suggested_action"] == "open_source_dialog"
    assert "字段不合法" not in json.dumps(data, ensure_ascii=False)


@pytest.mark.anyio
async def test_rule_draft_rejected_on_invalid_model_json(
    auth_client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    test_project_id: int,
    test_db,
) -> None:
    """模型返回非法字段或非法 JSON 结构时降级为 rejected，不写入可应用规则。"""
    user_id = await _get_test_user_id()
    await _seed_workbench_config(test_project_id, user_id)
    await _save_provider(auth_client)

    async def fake_call_provider_json(**_: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        return {
            "verdict": "ready",
            "rule_type": "formula_calculate",
            "reasoning_summary": "尝试生成一个当前不支持的公式规则。",
        }, {}

    monkeypatch.setattr("backend.app.ai.agent_service.call_provider_json", fake_call_provider_json)

    response = await auth_client.post(
        "/api/v1/ai/agents/rule-draft",
        json={"description": "把两列相加后比较"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["verdict"] == "rejected"
    assert data["draft"]["rules_to_add"] == []
    assert "字段不合法" in data["rejection_reason"]


@pytest.mark.anyio
async def test_draft_history_is_project_user_scoped_and_limited_to_20(
    auth_client: AsyncClient,
    test_project_id: int,
    test_db,
) -> None:
    """草稿历史按当前项目和用户隔离，并只返回最近 20 条。"""
    user_id = await _get_test_user_id()
    async with async_session_factory() as session:
        other_project = Project(name="other-ai-drafts", description="")
        session.add(other_project)
        await session.flush()
        for index in range(25):
            response_json = {
                "verdict": "rejected",
                "confidence": 0.1,
                "reasoning_summary": f"当前用户草稿 {index}",
                "draft": {
                    "sources_to_add": [],
                    "variables_to_add": [],
                    "rules_to_add": [],
                    "reuse_variable_tags": [],
                },
                "missing": [],
                "applied": False,
            }
            session.add(
                AiRuleDraftRecord(
                    project_id=test_project_id,
                    user_id=user_id,
                    description=f"desc-{index}",
                    verdict="rejected",
                    response_json=json.dumps(response_json, ensure_ascii=False),
                )
            )
        session.add(
            AiRuleDraftRecord(
                project_id=other_project.id,
                user_id=user_id,
                description="other-project",
                verdict="ready",
                response_json=json.dumps(
                    {
                        "verdict": "ready",
                        "confidence": 1,
                        "reasoning_summary": "不应返回",
                        "draft": {
                            "sources_to_add": [],
                            "variables_to_add": [],
                            "rules_to_add": [],
                            "reuse_variable_tags": [],
                        },
                        "missing": [],
                        "applied": False,
                    },
                    ensure_ascii=False,
                ),
            )
        )
        await session.commit()

    response = await auth_client.get("/api/v1/ai/drafts")
    assert response.status_code == 200
    items = response.json()["data"]["items"]
    assert len(items) == 20
    assert all(item["reasoning_summary"] != "不应返回" for item in items)
    assert items[0]["description"] == "desc-24"
    assert items[-1]["description"] == "desc-5"

    first_id = items[0]["draft_id"]
    apply_response = await auth_client.post(f"/api/v1/ai/drafts/{first_id}/apply")
    assert apply_response.status_code == 200

    delete_response = await auth_client.delete(f"/api/v1/ai/drafts/{first_id}")
    assert delete_response.status_code == 200

    clear_response = await auth_client.delete("/api/v1/ai/drafts")
    assert clear_response.status_code == 200
    assert clear_response.json()["data"]["deleted"] == 24
