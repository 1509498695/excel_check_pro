"""用例生成 V1 Excel 导出测试。"""

from __future__ import annotations

import datetime
import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.app.database import async_session_factory
from backend.app.models import (
    ExecutionRunRecord,
    SourceEvidenceResourceRecord,
    SourceEvidenceRunRecord,
    SourceEvidenceVisualObservationRecord,
)
from backend.app.test_cases import source_evidence_storage
from backend.app.test_cases.constants import STANDARD_CASE_FIELD_LABELS, STANDARD_CASE_FIELDS


@pytest.fixture(autouse=True)
def _source_evidence_runtime(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        source_evidence_storage,
        "settings",
        SimpleNamespace(source_evidence_dir=tmp_path / "source-evidence"),
    )


def _blueprint() -> dict[str, Any]:
    return {
        "modules": [{"name": "活动入口"}, {"name": "奖励领取"}],
        "flows": [{"name": "进入活动并领取奖励"}],
        "requirement_traces": [
            {
                "source_row_index": 2,
                "source_fragment": "按配置开放入口",
                "blueprint_node": "活动入口",
            }
        ],
        "coverage_dimensions": [{"name": "生命周期"}, {"name": "时间刷新"}],
        "risks": [{"name": "入口图语义未读取"}],
        "unmapped_requirements": [],
        "unsupported_or_unfounded_test_points": [],
        "open_questions": [],
        "warnings": [
            {
                "source": "blueprint",
                "level": "warning",
                "message": "入口图语义未读取，需人工确认。",
            }
        ],
    }


def _cases() -> list[dict[str, Any]]:
    return [
        {
            "case_id": "TC-001",
            "module": "活动入口",
            "feature": "入口开放",
            "scenario": "按配置开放入口",
            "title": "活动入口按配置展示",
            "preconditions": "活动配置已开启",
            "steps": "进入主界面并查看活动入口",
            "expected_results": "活动入口按配置展示",
            "priority": "P1",
            "case_type": "功能",
            "source_requirement": "按配置开放入口",
            "config_source": "ActivityConfig",
            "planning_answer": "",
            "initial_status": "未执行",
            "bug_link": "",
            "remarks": "入口图需人工确认",
        },
        {
            "case_id": "TC-002",
            "module": "奖励领取",
            "feature": "每日领取",
            "scenario": "每日领取一次",
            "title": "奖励每日仅可领取一次",
            "preconditions": "玩家满足领取条件",
            "steps": "领取奖励后再次点击领取",
            "expected_results": "首次成功，重复领取被拦截",
            "priority": "P2",
            "case_type": "边界",
            "source_requirement": "每日领取一次",
            "config_source": "",
            "planning_answer": "",
            "initial_status": "未执行",
            "bug_link": "",
            "remarks": "",
        },
    ]


def _warnings() -> list[dict[str, str]]:
    return [
        {
            "source": "snapshot",
            "level": "warning",
            "message": "V1 仅读取单元格文本，未读取图片、附件、批注或评论语义。",
        },
        {
            "source": "cases",
            "level": "warning",
            "message": "未使用参考案例增强。",
        },
    ]


def _stats() -> dict[str, Any]:
    return {
        "total": 2,
        "priority_counts": {"P1": 1, "P2": 1},
        "module_counts": {"活动入口": 1, "奖励领取": 1},
        "case_type_counts": {"功能": 1, "边界": 1},
        "warning_count": 2,
    }


def _export_payload(**overrides: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "blueprint": _blueprint(),
        "cases": _cases(),
        "warnings": _warnings(),
        "stats": _stats(),
        "export_columns": list(STANDARD_CASE_FIELDS),
        "source_summary": "上传 Excel：planning.xlsx / 策划案",
    }
    payload.update(overrides)
    return payload


def _load_response_workbook(response) -> Any:
    return load_workbook(BytesIO(response.content))


async def _execution_run_count() -> int:
    async with async_session_factory() as session:
        result = await session.execute(select(func.count(ExecutionRunRecord.id)))
        return int(result.scalar_one())


async def _seed_source_evidence_run(
    project_id: int,
    *,
    status: str = "ready",
    expires_at: datetime.datetime | None = None,
) -> int:
    expires_at = expires_at or (
        datetime.datetime.now(datetime.UTC) + datetime.timedelta(days=1)
    )
    async with async_session_factory() as session:
        run = SourceEvidenceRunRecord(
            project_id=project_id,
            source_type="feishu",
            source_url="https://demo.feishu.cn/docx/doccn-secret-token",
            source_token="doccn-secret-token",
            source_identifier="doccn-secret-token",
            source_title="活动需求文档",
            status=status,
            expires_at=expires_at,
            raw_manifest_json=json.dumps(
                {
                    "run_id": 0,
                    "project_id": project_id,
                    "source_type": "feishu",
                    "source_title": "活动需求文档",
                    "doc_type": "docx",
                    "counts": {
                        "source_unit_count": 1,
                        "resource_count": 1,
                        "downloaded_resource_count": 0,
                        "failed_resource_count": 1,
                        "warning_count": 1,
                    },
                    "warnings": [
                        {
                            "source": "source_evidence",
                            "level": "warning",
                            "message": "资源 docx_img_001 待观察。",
                        }
                    ],
                    "expires_at": expires_at.isoformat(),
                },
                ensure_ascii=False,
            ),
        )
        session.add(run)
        await session.flush()
        run.raw_manifest_json = json.dumps(
            {**json.loads(run.raw_manifest_json), "run_id": run.id},
            ensure_ascii=False,
        )
        session.add(
            SourceEvidenceResourceRecord(
                run_id=run.id,
                project_id=project_id,
                ref="docx_img_001",
                resource_type="image",
                position="docx:block:img",
                filename="ui.png",
                file_token="file-secret-token",
                status="unobserved",
                download_status="download_failed",
                local_path="D:/runtime/source-evidence/secret/ui.png",
                mime_type="image/png",
            )
        )
        await session.commit()
        return run.id


async def _seed_visual_evidence(
    project_id: int,
    run_id: int,
    *,
    status: str = "adopted",
) -> int:
    async with async_session_factory() as session:
        run = await session.get(SourceEvidenceRunRecord, run_id)
        assert run is not None
        resource = (
            await session.execute(
                select(SourceEvidenceResourceRecord).where(
                    SourceEvidenceResourceRecord.project_id == project_id,
                    SourceEvidenceResourceRecord.run_id == run_id,
                    SourceEvidenceResourceRecord.ref == "docx_img_001",
                )
            )
        ).scalar_one()
        observation = SourceEvidenceVisualObservationRecord(
            run_id=run_id,
            project_id=project_id,
            resource_id=resource.id,
            ref=resource.ref,
            position=resource.position,
            filename=resource.filename,
            status=status,
            observation_path="",
            created_by=run.created_by,
            adopted_by=run.created_by if status == "adopted" else None,
            adopted_at=datetime.datetime.now(datetime.UTC) if status == "adopted" else None,
        )
        session.add(observation)
        await session.flush()
        observation_path = f"visual_evidence/observations/{observation.id}.json"
        source_evidence_storage.write_source_evidence_json(
            project_id,
            run_id,
            observation_path,
            {
                "id": observation.id,
                "run_id": run_id,
                "resource_id": resource.id,
                "ref": resource.ref,
                "position": resource.position,
                "summary": "图中展示活动入口按钮，按钮文案为“参与活动”。",
                "visible_text": "参与活动",
                "confidence": 0.87,
                "limitations": ["只能确认截图可见内容，不能确认配置规则。"],
                "source": {"provider": "openai", "model": "gpt-4o-mini"},
                "created_by": run.created_by,
                "created_at": datetime.datetime.now(datetime.UTC).isoformat(),
            },
        )
        observation.observation_path = observation_path
        resource.status = status
        resource.local_path = "visual_evidence/images/secret.jpg"
        await session.commit()
        return observation.id


async def _seed_adopted_visual_evidence(project_id: int, run_id: int) -> int:
    return await _seed_visual_evidence(project_id, run_id, status="adopted")


async def _seed_textless_image_source_evidence_run(project_id: int) -> int:
    expires_at = datetime.datetime.now(datetime.UTC) + datetime.timedelta(days=1)
    async with async_session_factory() as session:
        run = SourceEvidenceRunRecord(
            project_id=project_id,
            source_type="local_file",
            source_url="",
            source_token="sha256:image",
            source_identifier="sha256:image",
            source_title="ui.png",
            status="ready",
            expires_at=expires_at,
            raw_manifest_json=json.dumps(
                {
                    "run_id": 0,
                    "project_id": project_id,
                    "source_type": "local_file",
                    "source_title": "ui.png",
                    "doc_type": "image",
                    "warnings": [],
                    "expires_at": expires_at.isoformat(),
                },
                ensure_ascii=False,
            ),
        )
        session.add(run)
        await session.flush()
        run.raw_manifest_json = json.dumps(
            {**json.loads(run.raw_manifest_json), "run_id": run.id},
            ensure_ascii=False,
        )
        source_evidence_storage.ensure_source_evidence_subdirs(
            project_id=project_id,
            run_id=run.id,
        )
        session.add(
            SourceEvidenceResourceRecord(
                run_id=run.id,
                project_id=project_id,
                ref="local_img_001",
                resource_type="image",
                position="local:image=1",
                filename="local_img_001.png",
                status="unobserved",
                download_status="downloaded",
                local_path="images/local_img_001.png",
                mime_type="image/png",
            )
        )
        source_evidence_storage.write_source_evidence_json(
            project_id,
            run.id,
            "raw/parsed_source.json",
            {
                "source_type": "local_file",
                "title": "ui.png",
                "doc_type": "image",
                "token": "sha256:image",
                "url": "",
                "markdown": '<image ref="local_img_001" position="local:image=1" />',
                "source_units": [],
                "resources": [],
                "raw_manifest": {},
                "warnings": [],
            },
        )
        await session.commit()
        return run.id


@pytest.mark.anyio
async def test_export_returns_xlsx_with_three_sheets_and_standard_fields(
    auth_client: AsyncClient,
) -> None:
    """无主参考时按标准字段顺序导出三个 Sheet。"""
    before_count = await _execution_run_count()

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(),
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    assert "test-cases-" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.xlsx"')

    workbook = _load_response_workbook(response)
    assert workbook.sheetnames == ["测试用例", "用例蓝图", "生成说明"]
    case_sheet = workbook["测试用例"]
    assert [cell.value for cell in case_sheet[1]] == [
        STANDARD_CASE_FIELD_LABELS[field] for field in STANDARD_CASE_FIELDS
    ]
    assert case_sheet["A2"].value == "TC-001"
    assert case_sheet["E2"].value == "活动入口按配置展示"
    assert workbook["用例蓝图"]["A1"].value == "区块"
    assert workbook["生成说明"]["A1"].value == "项目"
    assert any(
        "未读取图片" in str(row[2].value)
        for row in workbook["生成说明"].iter_rows(min_row=2)
        if row[0].value == "warning"
    )
    assert await _execution_run_count() == before_count


@pytest.mark.anyio
async def test_export_respects_primary_reference_profile_and_ignores_unknown_columns(
    auth_client: AsyncClient,
) -> None:
    """有主参考字段画像时只采用可映射标准字段，未知列不强行生成。"""
    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            primary_reference_profile={
                "name": "历史活动用例.xlsx",
                "columns": [
                    {"original_name": "历史标题", "standard_field": "title"},
                    {"original_name": "历史未知列", "standard_field": ""},
                    {"original_name": "历史优先级", "standard_field": "priority"},
                    {"original_name": "历史预期", "standard_field": "expected_results"},
                ],
                "raw_prompt": "sk-secret should never appear",
                "provider_response": {"api_key": "sk-secret"},
            }
        ),
    )

    assert response.status_code == 200, response.text
    workbook = _load_response_workbook(response)
    headers = [cell.value for cell in workbook["测试用例"][1]]

    assert headers[:3] == ["用例标题", "优先级", "预期结果"]
    assert "历史未知列" not in headers
    assert STANDARD_CASE_FIELD_LABELS["case_id"] in headers
    assert STANDARD_CASE_FIELD_LABELS["module"] in headers
    assert workbook["测试用例"]["A2"].value == "活动入口按配置展示"
    assert workbook["测试用例"]["B2"].value == "P1"
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "sk-secret" not in workbook_text
    assert "raw_prompt" not in workbook_text
    assert "provider_response" not in workbook_text


@pytest.mark.anyio
async def test_export_uses_selected_excel_reference_sheet_columns(
    auth_client: AsyncClient,
) -> None:
    """完整 Excel 主参考画像传入导出时，应按选中 Sheet 的可识别字段排序。"""
    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            primary_reference_profile={
                "source_type": "excel",
                "source_name": "历史活动用例.xlsx",
                "default_sheet_name": "测试用例",
                "selected_sheet_name": "边界用例",
                "reference_case_count": 7,
                "columns": [
                    {"original_name": "默认标题", "standard_field": "title"},
                ],
                "sheet_options": [
                    {
                        "name": "测试用例",
                        "reference_case_count": 2,
                        "header_row_index": 1,
                        "columns": [
                            {"original_name": "默认标题", "standard_field": "title"},
                        ],
                    },
                    {
                        "name": "边界用例",
                        "reference_case_count": 7,
                        "header_row_index": 1,
                        "columns": [
                            {"original_name": "历史优先级", "standard_field": "priority"},
                            {"original_name": "历史模块", "standard_field": "module"},
                            {"original_name": "历史未知列", "standard_field": None},
                            {"original_name": "历史步骤", "standard_field": "steps"},
                            {"original_name": "历史标题", "standard_field": "title"},
                        ],
                    },
                ],
            }
        ),
    )

    assert response.status_code == 200, response.text
    workbook = _load_response_workbook(response)
    headers = [cell.value for cell in workbook["测试用例"][1]]

    assert headers[:4] == ["优先级", "功能模块", "操作步骤", "用例标题"]
    assert "历史未知列" not in headers
    assert STANDARD_CASE_FIELD_LABELS["case_id"] in headers


@pytest.mark.anyio
async def test_export_writes_safe_source_evidence_summary_from_server(
    auth_client: AsyncClient,
    test_project_id: int,
) -> None:
    """传 run_id 时导出说明使用服务端安全摘要，并过滤敏感内容。"""
    run_id = await _seed_source_evidence_run(test_project_id)

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            source_evidence_summary=(
                "客户端摘要不可信 sk-secret doccn-secret-token file-secret-token "
                "raw prompt provider_response"
            ),
        ),
    )

    assert response.status_code == 200, response.text
    workbook = _load_response_workbook(response)
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "Source Evidence 摘要" in workbook_text
    assert "活动需求文档" in workbook_text
    assert "TTL" in workbook_text
    assert "资源总数=1" in workbook_text
    assert "未观察/未采纳" in workbook_text
    assert "客户端摘要不可信" not in workbook_text
    for forbidden in (
        "sk-secret",
        "doccn-secret-token",
        "file-secret-token",
        "docx_img_001",
        "D:/runtime/source-evidence",
        "raw prompt",
        "provider_response",
    ):
        assert forbidden not in workbook_text


@pytest.mark.anyio
async def test_export_includes_only_adopted_visual_evidence_summary(
    auth_client: AsyncClient,
    test_project_id: int,
) -> None:
    """导出说明只包含已采纳视觉证据摘要，不包含路径、token 或模型原始内容。"""
    run_id = await _seed_source_evidence_run(test_project_id)
    evidence_id = await _seed_adopted_visual_evidence(test_project_id, run_id)

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            adopted_visual_evidence_ids=[evidence_id],
        ),
    )

    assert response.status_code == 200, response.text
    workbook = _load_response_workbook(response)
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "已采纳视觉证据" in workbook_text
    assert "docx_img_001" in workbook_text
    assert "图中展示活动入口按钮" in workbook_text
    assert "只能确认截图可见内容" in workbook_text
    for forbidden in (
        "visual_evidence/observations",
        "visual_evidence/images",
        "file-secret-token",
        "doccn-secret-token",
        "prompt",
        "provider_response",
        "sk-",
    ):
        assert forbidden not in workbook_text


@pytest.mark.anyio
async def test_export_rejects_invalid_adopted_visual_evidence_ids(
    auth_client: AsyncClient,
    test_project_id: int,
) -> None:
    run_id = await _seed_source_evidence_run(test_project_id)
    other_run_id = await _seed_source_evidence_run(test_project_id)
    other_run_evidence_id = await _seed_adopted_visual_evidence(
        test_project_id,
        other_run_id,
    )
    observed_id = await _seed_visual_evidence(test_project_id, run_id, status="observed")

    missing_response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            adopted_visual_evidence_ids=[99999999],
        ),
    )
    other_run_response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            adopted_visual_evidence_ids=[other_run_evidence_id],
        ),
    )
    observed_response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            adopted_visual_evidence_ids=[observed_id],
        ),
    )

    assert missing_response.status_code == 404
    assert other_run_response.status_code == 404
    assert observed_response.status_code == 400
    assert "已采纳" in observed_response.text


@pytest.mark.anyio
async def test_export_rejects_textless_image_run_without_adopted_evidence(
    auth_client: AsyncClient,
    test_project_id: int,
) -> None:
    run_id = await _seed_textless_image_source_evidence_run(test_project_id)

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(source_evidence_run_id=run_id),
    )

    assert response.status_code == 409
    assert "采纳视觉证据" in response.text


@pytest.mark.anyio
async def test_export_rejects_payload_referencing_unadopted_visual_ref(
    auth_client: AsyncClient,
    test_project_id: int,
) -> None:
    run_id = await _seed_source_evidence_run(test_project_id)
    cases = _cases()
    cases[0]["remarks"] = "错误引用未采纳图片 docx_img_001。"

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            source_evidence_run_id=run_id,
            cases=cases,
        ),
    )

    assert response.status_code == 400
    assert "未采纳" in response.text


@pytest.mark.anyio
@pytest.mark.parametrize(
    ("status", "expires_at"),
    [
        ("cleaned", datetime.datetime.now(datetime.UTC) + datetime.timedelta(days=1)),
        ("ready", datetime.datetime.now(datetime.UTC) - datetime.timedelta(seconds=1)),
    ],
)
async def test_export_rejects_expired_or_cleaned_source_evidence_run(
    auth_client: AsyncClient,
    test_project_id: int,
    status: str,
    expires_at: datetime.datetime,
) -> None:
    """过期或 cleaned run 不允许导出证据复查说明。"""
    run_id = await _seed_source_evidence_run(
        test_project_id,
        status=status,
        expires_at=expires_at,
    )

    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(source_evidence_run_id=run_id),
    )

    assert response.status_code == 409
    assert "重新读取来源" in response.json()["detail"]


@pytest.mark.anyio
async def test_export_without_run_id_sanitizes_client_evidence_summary(
    auth_client: AsyncClient,
) -> None:
    """未传 run_id 时只写入脱敏后的客户端摘要。"""
    response = await auth_client.post(
        "/api/v1/test-cases/export",
        json=_export_payload(
            evidence_summary=(
                "Source Evidence：待观察资源；api_key=sk-client-secret；"
                "Authorization: Bearer token；prompt=raw"
            ),
        ),
    )

    assert response.status_code == 200, response.text
    workbook = _load_response_workbook(response)
    workbook_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "待观察资源" in workbook_text
    assert "sk-client-secret" not in workbook_text
    assert "Authorization" not in workbook_text
    assert "prompt=raw" not in workbook_text
