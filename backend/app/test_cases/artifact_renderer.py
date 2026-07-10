"""Deterministic renderer for the canonical local test-case workbook."""

from __future__ import annotations

from collections import Counter
from copy import copy
from io import BytesIO
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from backend.app.test_cases.case_contract import canonical_case_fields


DEFAULT_TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "用例模板.xlsx"
EXECUTION_COLUMN_COUNT = 3
DATA_START_ROW = 11


def build_canonical_test_case_workbook(
    *,
    cases: list[dict[str, Any]],
    title: str,
    source_summary: str,
    blueprint: dict[str, Any],
    atoms: list[dict[str, Any]],
    coverage_audit: dict[str, Any],
    quality_audit: dict[str, Any],
    metadata: dict[str, Any],
    template_path: Path | None = None,
) -> BytesIO:
    """Build the four-sheet workbook used by web runs and the standalone CLI."""
    template = Path(template_path or DEFAULT_TEMPLATE_PATH)
    workbook = load_workbook(template) if template.exists() else Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"
    for extra_sheet in list(workbook.worksheets)[1:]:
        workbook.remove(extra_sheet)

    canonical_cases = [
        canonical_case_fields(
            _json_object(item.get("fields", item)),
            case_id=str(item.get("case_id") or ""),
        )
        for item in cases
    ]
    _write_case_sheet(
        sheet,
        cases=canonical_cases,
        title=title,
        source_summary=source_summary,
    )
    _write_blueprint_sheet(workbook.create_sheet("用例蓝图"), blueprint)
    _write_generation_notes_sheet(
        workbook.create_sheet("生成说明"),
        cases=canonical_cases,
        metadata=metadata,
        source_summary=source_summary,
        coverage_audit=coverage_audit,
        quality_audit=quality_audit,
    )
    _write_audit_sheet(
        workbook.create_sheet("覆盖审计"),
        cases=cases,
        atoms=atoms,
        coverage_audit=coverage_audit,
        quality_audit=quality_audit,
    )
    workbook.active = 0
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _write_case_sheet(
    sheet: Any,
    *,
    cases: list[dict[str, str]],
    title: str,
    source_summary: str,
) -> None:
    if sheet.max_column > 12:
        sheet.delete_cols(13, sheet.max_column - 12)
    for merged in list(sheet.merged_cells.ranges):
        if merged.max_row >= DATA_START_ROW:
            sheet.unmerge_cells(str(merged))
    max_row = max(sheet.max_row, DATA_START_ROW + len(cases) - 1)
    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(1, 13):
            sheet.cell(row, col).value = None
    if hasattr(sheet, "data_validations"):
        sheet.data_validations.dataValidation = []

    _set_if_writable(sheet, "D1", title)
    _set_if_writable(sheet, "B7", f"来源：{source_summary or 'Generation Run 已登记来源'}")
    _write_case_headers(sheet)
    style_row = DATA_START_ROW
    for offset, case in enumerate(cases):
        row = DATA_START_ROW + offset
        _copy_row_style(sheet, style_row, row, 12)
        values = [
            offset + 1,
            case["primary_module"],
            case["secondary_module"],
            case["checkpoint"],
            case["preconditions"],
            case["steps"],
            case["expected_results"],
            case["priority"],
            case["remarks"],
            "",
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col)
            cell.value = value
            cell.alignment = Alignment(
                horizontal="center" if col in {1, 2, 3, 4, 8, 10, 11, 12} else "left",
                vertical="center",
                wrap_text=True,
            )
        _apply_priority_style(sheet.cell(row, 8), case["priority"])
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 0, 58)

    data_end = DATA_START_ROW + len(cases) - 1
    if cases:
        _merge_runs(sheet, 2, DATA_START_ROW, [case["primary_module"] for case in cases])
        _merge_runs(
            sheet,
            3,
            DATA_START_ROW,
            [case["secondary_module"] for case in cases],
            [case["primary_module"] for case in cases],
        )
        _add_validations(sheet, DATA_START_ROW, data_end)
        _write_stats_formulas(sheet, DATA_START_ROW, data_end)
    else:
        sheet["B2"] = 0
    sheet.freeze_panes = "A11"
    sheet.sheet_view.showGridLines = False


def _write_case_headers(sheet: Any) -> None:
    values = {
        "A9": "用例编号",
        "B9": "测试标题",
        "B10": "一级模块",
        "C10": "二级模块",
        "D10": "检查点",
        "E9": "前置条件",
        "F9": "操作步骤",
        "G9": "预期结果",
        "H9": "优先级",
        "I9": "备注",
        "J9": "执行列1",
        "K9": "执行列2",
        "L9": "执行列3",
        "J10": "T/F/D/N/A",
        "K10": "T/F/D/N/A",
        "L10": "T/F/D/N/A",
    }
    for cell, value in values.items():
        _set_if_writable(sheet, cell, value)


def _write_stats_formulas(sheet: Any, start_row: int, end_row: int) -> None:
    sheet["B2"] = f"=COUNTA(A{start_row}:A{end_row})"
    for col in range(10, 13):
        letter = get_column_letter(col)
        sheet[f"{letter}3"] = f'=COUNTIF({letter}{start_row}:{letter}{end_row},"T")'
        sheet[f"{letter}4"] = f'=COUNTIF({letter}{start_row}:{letter}{end_row},"F")'
        sheet[f"{letter}5"] = f'=COUNTIF({letter}{start_row}:{letter}{end_row},"D")'
        sheet[f"{letter}6"] = f'=COUNTIF({letter}{start_row}:{letter}{end_row},"N/A")'
        sheet[f"{letter}7"] = f"=IFERROR(({letter}3+{letter}4+{letter}5)/$B$2,0)"
        sheet[f"{letter}8"] = f"=IFERROR({letter}3/({letter}3+{letter}4),0)"


def _add_validations(sheet: Any, start_row: int, end_row: int) -> None:
    priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=False)
    status = DataValidation(type="list", formula1='"T,F,D,N/A"', allow_blank=True)
    sheet.add_data_validation(priority)
    sheet.add_data_validation(status)
    priority.add(f"H{start_row}:H{end_row}")
    status.add(f"J{start_row}:L{end_row}")


def _write_blueprint_sheet(sheet: Any, blueprint: dict[str, Any]) -> None:
    sheet.append(["蓝图节点", "内容"])
    if blueprint:
        for key, value in blueprint.items():
            sheet.append([str(key), _format_value(value)])
    else:
        sheet.append(["状态", "未生成蓝图"])
    _style_simple_sheet(sheet, widths=(28, 100))


def _write_generation_notes_sheet(
    sheet: Any,
    *,
    cases: list[dict[str, str]],
    metadata: dict[str, Any],
    source_summary: str,
    coverage_audit: dict[str, Any],
    quality_audit: dict[str, Any],
) -> None:
    priority_counts = Counter(case["priority"] for case in cases)
    module_counts = Counter(case["primary_module"] for case in cases)
    coverage_limitations = [
        str(item.get("message") or "")
        for item in coverage_audit.get("export_limitations", [])
        if isinstance(item, dict) and item.get("message")
    ]
    rows = [
        ("标题", metadata.get("title", "")),
        ("Generation Run ID", metadata.get("run_id", "")),
        ("运行状态", metadata.get("status", "")),
        ("来源摘要", source_summary),
        ("用例总数", len(cases)),
        ("优先级分布", dict(priority_counts)),
        ("一级模块分布", dict(module_counts)),
        ("严格模式", "是" if metadata.get("strict_mode") else "否"),
        ("覆盖审计状态", coverage_audit.get("status", "missing")),
        ("覆盖/导出限制", "；".join(coverage_limitations) or "无"),
        (
            "质量审计",
            (
                f"状态：{quality_audit.get('status', 'missing')}；"
                f"阻塞：{quality_audit.get('blocking_count', 0)}；"
                f"警告：{quality_audit.get('warning_count', 0)}"
            ),
        ),
        ("模板契约", "A-I 固定用例字段；J-L 三个中性执行列；P0/P1/P2/P3"),
        ("产物说明", "工作簿在 Generation Run 完成时自动生成；下载只读取已校验文件。"),
    ]
    sheet.append(["项目", "内容"])
    for row in rows:
        sheet.append([row[0], _format_value(row[1])])
    _style_simple_sheet(sheet, widths=(26, 100))


def _write_audit_sheet(
    sheet: Any,
    *,
    cases: list[dict[str, Any]],
    atoms: list[dict[str, Any]],
    coverage_audit: dict[str, Any],
    quality_audit: dict[str, Any],
) -> None:
    linked: dict[str, list[str]] = {}
    for case in cases:
        for atom_id in case.get("atom_refs", []):
            linked.setdefault(str(atom_id), []).append(str(case.get("case_id") or ""))
    uncovered = {
        str(item)
        for item in coverage_audit.get("uncovered_atom_ids", [])
    }
    limitation_by_atom: dict[str, list[str]] = {}
    global_limitations: list[str] = []
    for limitation in coverage_audit.get("export_limitations", []):
        if not isinstance(limitation, dict):
            continue
        message = str(limitation.get("message") or "")
        if not message:
            continue
        atom_ids = limitation.get("atom_ids")
        if isinstance(atom_ids, list) and atom_ids:
            for atom_id in atom_ids:
                limitation_by_atom.setdefault(str(atom_id), []).append(message)
        else:
            global_limitations.append(message)
    sheet.append(
        [
            "atom id",
            "atom type",
            "source sheet",
            "source rows",
            "requirement",
            "coverage status",
            "linked case ids",
            "limitation notes",
        ]
    )
    for atom in atoms:
        atom_id = str(atom.get("atom_id") or "")
        case_bearing = str(atom.get("atom_type") or "") not in {"open_question", "limitation"}
        if not case_bearing:
            coverage_status = "not_case_bearing"
        elif linked.get(atom_id):
            coverage_status = "covered"
        elif atom_id in uncovered:
            coverage_status = "uncovered"
        else:
            coverage_status = str(atom.get("coverage_status") or "unmapped")
        sheet.append(
            [
                atom_id,
                atom.get("atom_type", ""),
                atom.get("source_sheet_name", ""),
                _source_rows(atom),
                atom.get("requirement_text", ""),
                coverage_status,
                "、".join(linked.get(atom_id, [])),
                "；".join(limitation_by_atom.get(atom_id, [])),
            ]
        )
    coverage_start = sheet.max_row + 2
    sheet.cell(coverage_start, 1, "Coverage Audit Summary")
    sheet.cell(coverage_start + 1, 1, "状态")
    sheet.cell(coverage_start + 1, 2, coverage_audit.get("status", "missing"))
    sheet.cell(coverage_start + 2, 1, "未覆盖 Requirement Atom")
    sheet.cell(coverage_start + 2, 2, coverage_audit.get("uncovered_atoms", len(uncovered)))
    sheet.cell(coverage_start + 3, 1, "失败 chunk")
    sheet.cell(coverage_start + 3, 2, coverage_audit.get("failed_chunk_count", 0))
    sheet.cell(coverage_start + 4, 1, "全局限制")
    sheet.cell(coverage_start + 4, 2, "；".join(global_limitations) or "无")
    quality_start = sheet.max_row + 2
    sheet.cell(quality_start, 1, "Case Quality Audit")
    sheet.cell(quality_start + 1, 1, "状态")
    sheet.cell(quality_start + 1, 2, quality_audit.get("status", ""))
    sheet.cell(quality_start + 2, 1, "阻塞问题")
    sheet.cell(quality_start + 2, 2, quality_audit.get("blocking_count", 0))
    sheet.cell(quality_start + 3, 1, "警告")
    sheet.cell(quality_start + 3, 2, quality_audit.get("warning_count", 0))
    issue_row = quality_start + 5
    sheet.cell(issue_row, 1, "case id")
    sheet.cell(issue_row, 2, "severity")
    sheet.cell(issue_row, 3, "code")
    sheet.cell(issue_row, 4, "message")
    for issue in quality_audit.get("issues", []):
        if not isinstance(issue, dict):
            continue
        issue_row += 1
        sheet.append(
            [
                issue.get("case_id", ""),
                issue.get("severity", ""),
                issue.get("code", ""),
                issue.get("message", ""),
            ]
        )
    _style_simple_sheet(sheet, widths=(20, 20, 24, 18, 80, 20, 32, 60))
    for cell in sheet[coverage_start]:
        cell.font = Font(bold=True)
    for cell in sheet[quality_start]:
        cell.font = Font(bold=True)


def _style_simple_sheet(sheet: Any, *, widths: tuple[int, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _copy_row_style(sheet: Any, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)


def _merge_runs(
    sheet: Any,
    col: int,
    start_row: int,
    values: list[str],
    scope: list[str] | None = None,
) -> None:
    run_start = 0
    for index in range(1, len(values) + 1):
        boundary = index == len(values) or values[index] != values[run_start]
        if not boundary and scope is not None:
            boundary = scope[index] != scope[run_start]
        if boundary:
            if index - run_start > 1 and values[run_start]:
                sheet.merge_cells(
                    start_row=start_row + run_start,
                    start_column=col,
                    end_row=start_row + index - 1,
                    end_column=col,
                )
            run_start = index


def _apply_priority_style(cell: Any, priority: str) -> None:
    fills = {
        "P0": "F8CBAD",
        "P1": "FFF2CC",
        "P2": "D9EAD3",
        "P3": "D9EAF7",
    }
    cell.fill = PatternFill("solid", fgColor=fills.get(priority, fills["P2"]))
    cell.font = Font(bold=True, color="000000")


def _set_if_writable(sheet: Any, cell: str, value: Any) -> None:
    try:
        sheet[cell] = value
    except AttributeError:
        return


def _source_rows(atom: dict[str, Any]) -> str:
    start = atom.get("source_row_start")
    end = atom.get("source_row_end")
    if start in (None, ""):
        return ""
    if end in (None, "", start):
        return str(start)
    return f"{start}-{end}"


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _json_object(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}
