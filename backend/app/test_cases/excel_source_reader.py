"""Reader for local `.xlsx` Source Evidence uploads."""

from __future__ import annotations

from datetime import datetime
import shutil
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from backend.app.test_cases import source_evidence_storage
from backend.app.test_cases.schemas import (
    GenerationWarning,
    ParsedSource,
    ParsedSourceCell,
    ParsedSourceResource,
    ParsedSourceUnit,
)
from backend.config import settings

try:
    import xlrd
except ImportError:  # pragma: no cover - 依赖缺失时由 .xls 路径给出中文错误
    xlrd = None  # type: ignore[assignment]


def read_xlsx_source_evidence(
    *,
    project_id: int,
    run_id: int,
    source_path: Path,
    original_filename: str,
    source_sha256: str,
    origin: str = "local_file",
) -> ParsedSource:
    """Read visible workbook sheets and extract embedded images into the run."""
    try:
        workbook = load_workbook(source_path, data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError(f"读取本地上传文件失败：无法打开 .xlsx 工作簿：{exc}") from exc

    warnings: list[GenerationWarning] = []
    units: list[ParsedSourceUnit] = []
    resources: list[ParsedSourceResource] = []
    markdown_lines = [
        f"# Source: {original_filename}",
        "",
        "Type: xlsx",
        "",
    ]

    visible_sheet_number = 0
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            warnings.append(
                GenerationWarning(
                    source=origin,
                    level="warning",
                    message=f"隐藏 Sheet 已排除：{sheet.title}",
                )
            )
            continue

        visible_sheet_number += 1
        cells = _collect_sheet_cells(sheet)
        units.append(
            ParsedSourceUnit(
                unit_id=f"xlsx_s{visible_sheet_number:03d}",
                kind="sheet",
                title=sheet.title,
                cells=cells,
                metadata={
                    "sheet_index": visible_sheet_number,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                },
            )
        )
        markdown_lines.extend(_sheet_markdown(sheet.title, cells))
        sheet_resources = _extract_sheet_images(
            project_id=project_id,
            run_id=run_id,
            sheet=sheet,
            sheet_number=visible_sheet_number,
        )
        resources.extend(sheet_resources)
        for resource in sheet_resources:
            markdown_lines.append(
                f'<image ref="{resource.ref}" position="{resource.position}" />'
            )
        markdown_lines.append("")

    if not units:
        warnings.append(
            GenerationWarning(
                source=origin,
                level="warning",
                message="工作簿没有可见 Sheet，未提取到文本主体。",
            )
        )

    return ParsedSource(
        title=original_filename,
        source_type=origin,
        doc_type="xlsx",
        token=f"sha256:{source_sha256}",
        url="",
        markdown="\n".join(markdown_lines).strip() + "\n",
        source_units=units,
        resources=resources,
        raw_manifest={
            "doc_type": "xlsx",
            "original_filename": original_filename,
            "sha256": source_sha256,
            "visible_sheet_count": len(units),
            "resource_count": len(resources),
        },
        warnings=warnings,
    )


def read_xls_source_evidence(
    *,
    project_id: int,
    run_id: int,
    source_path: Path,
    original_filename: str,
    source_sha256: str,
    origin: str = "local_file",
) -> ParsedSource:
    """Read `.xls` text and best-effort extract images via controlled conversion."""
    parsed = _read_xls_text_source(
        project_id=project_id,
        run_id=run_id,
        source_path=source_path,
        original_filename=original_filename,
        source_sha256=source_sha256,
        origin=origin,
    )
    warnings = list(parsed.warnings)
    raw_manifest = dict(parsed.raw_manifest)
    resources = list(parsed.resources)
    markdown = parsed.markdown

    try:
        converted_relative_path = _convert_xls_to_xlsx(
            project_id=project_id,
            run_id=run_id,
            source_path=source_path,
            original_filename=original_filename,
        )
        converted_path = source_evidence_storage.resolve_source_evidence_path(
            project_id,
            run_id,
            converted_relative_path,
        )
        converted = read_xlsx_source_evidence(
            project_id=project_id,
            run_id=run_id,
            source_path=converted_path,
            original_filename=original_filename,
            source_sha256=source_sha256,
            origin=origin,
        )
        resources.extend(converted.resources)
        warnings.extend(converted.warnings)
        raw_manifest["xls_image_conversion"] = {
            "status": "success",
            "converted_relative_path": converted_relative_path,
            "resource_count": len(converted.resources),
        }
        if converted.resources:
            markdown = "\n".join(
                [
                    markdown.rstrip(),
                    "",
                    *[
                        f'<image ref="{resource.ref}" position="{resource.position}" />'
                        for resource in converted.resources
                    ],
                ]
            ).strip() + "\n"
    except Exception as exc:  # noqa: BLE001 - 转换失败按需求降级为 warning
        raw_manifest["xls_image_conversion"] = {
            "status": "failed",
            "message": _sanitize_converter_message(str(exc), source_path=source_path),
        }
        warnings.append(
            GenerationWarning(
                source=origin,
                level="warning",
                message=(
                    ".xls 图片转换失败："
                    f"{raw_manifest['xls_image_conversion']['message']}；"
                    "文本主体仍可用于生成，图片未参与理解。"
                ),
            )
        )

    raw_manifest["resource_count"] = len(resources)
    return parsed.model_copy(
        update={
            "resources": resources,
            "warnings": warnings,
            "raw_manifest": raw_manifest,
            "markdown": markdown,
        }
    )


def _read_xls_text_source(
    *,
    project_id: int,
    run_id: int,
    source_path: Path,
    original_filename: str,
    source_sha256: str,
    origin: str = "local_file",
) -> ParsedSource:
    if xlrd is None:
        raise ValueError("读取本地上传文件失败：当前环境缺少 xlrd，无法读取 .xls 文本。")
    try:
        workbook = xlrd.open_workbook(str(source_path), formatting_info=False)
    except Exception as exc:
        raise ValueError(f"读取本地上传文件失败：无法打开 .xls 工作簿：{exc}") from exc

    warnings: list[GenerationWarning] = []
    units: list[ParsedSourceUnit] = []
    markdown_lines = [
        f"# Source: {original_filename}",
        "",
        "Type: xls",
        "",
    ]
    visible_sheet_number = 0
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        if int(getattr(sheet, "visibility", 0) or 0) != 0:
            warnings.append(
                GenerationWarning(
                    source=origin,
                    level="warning",
                    message=f"隐藏 Sheet 已排除：{sheet.name}",
                )
            )
            continue
        visible_sheet_number += 1
        cells = _collect_xls_sheet_cells(sheet)
        units.append(
            ParsedSourceUnit(
                unit_id=f"xls_s{visible_sheet_number:03d}",
                kind="sheet",
                title=sheet.name,
                cells=cells,
                metadata={
                    "sheet_index": visible_sheet_number,
                    "max_row": sheet.nrows,
                    "max_column": sheet.ncols,
                },
            )
        )
        markdown_lines.extend(_sheet_markdown(sheet.name, cells))
        markdown_lines.append("")

    if not units:
        warnings.append(
            GenerationWarning(
                source=origin,
                level="warning",
                message="工作簿没有可见 Sheet，未提取到文本主体。",
            )
        )

    return ParsedSource(
        title=original_filename,
        source_type=origin,
        doc_type="xls",
        token=f"sha256:{source_sha256}",
        url="",
        markdown="\n".join(markdown_lines).strip() + "\n",
        source_units=units,
        resources=[],
        raw_manifest={
            "doc_type": "xls",
            "original_filename": original_filename,
            "sha256": source_sha256,
            "visible_sheet_count": len(units),
            "resource_count": 0,
        },
        warnings=warnings,
    )


def _collect_xls_sheet_cells(sheet: Any) -> list[ParsedSourceCell]:
    cells: list[ParsedSourceCell] = []
    for row_index in range(sheet.nrows):
        for col_index in range(sheet.ncols):
            raw_value = sheet.cell_value(row_index, col_index)
            text = _xls_cell_text(raw_value)
            if not text:
                continue
            row = row_index + 1
            col = col_index + 1
            cells.append(
                ParsedSourceCell(
                    coord=f"{get_column_letter(col)}{row}",
                    row=row,
                    col=col,
                    text=text,
                    raw=raw_value,
                )
            )
    return cells


def _xls_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _convert_xls_to_xlsx(
    *,
    project_id: int,
    run_id: int,
    source_path: Path,
    original_filename: str,
) -> str:
    executable = _resolve_soffice_executable()
    if executable is None:
        raise RuntimeError("未配置 LibreOffice/soffice 转换器")
    output_dir = source_evidence_storage.resolve_source_evidence_path(
        project_id,
        run_id,
        "raw/converted",
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    profile_dir = output_dir / "profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "source.xlsx"
    candidate_output = output_dir / f"{Path(original_filename).stem}.xlsx"
    args = [
        executable,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--norestore",
        f"-env:UserInstallation={profile_dir.as_uri()}",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(source_path),
    ]
    timeout = max(1, int(settings.source_evidence_xls_convert_timeout_seconds))
    try:
        completed = subprocess.run(
            args,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            check=False,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f"LibreOffice/soffice 转换超时（>{timeout}s）") from exc
    if completed.returncode != 0:
        message = (completed.stderr or completed.stdout or "转换命令返回非零退出码").strip()
        raise RuntimeError(_sanitize_converter_message(message, source_path=source_path))
    if candidate_output.exists() and candidate_output != output_path:
        candidate_output.replace(output_path)
    if not output_path.exists():
        raise RuntimeError("LibreOffice/soffice 转换未生成 .xlsx 文件")
    return "raw/converted/source.xlsx"


def _resolve_soffice_executable() -> str | None:
    configured = settings.source_evidence_soffice_executable.strip().strip('"').strip("'")
    if not configured:
        return None
    configured_path = Path(configured).expanduser()
    if configured_path.is_file():
        return str(configured_path.resolve(strict=False))
    discovered = shutil.which(configured)
    if discovered:
        return discovered
    return None


def _sanitize_converter_message(message: str, *, source_path: Path) -> str:
    safe_message = (message or "转换失败").strip()
    source_text = str(source_path)
    if source_text:
        safe_message = safe_message.replace(source_text, source_path.name)
    return safe_message.replace("\r", " ").replace("\n", " ")[:300]


def _collect_sheet_cells(sheet: Any) -> list[ParsedSourceCell]:
    cells: list[ParsedSourceCell] = []
    for row in sheet.iter_rows():
        for cell in row:
            text = _cell_text(cell)
            if not text:
                continue
            cells.append(
                ParsedSourceCell(
                    coord=cell.coordinate,
                    row=cell.row,
                    col=cell.column,
                    text=text,
                    raw=cell.value,
                )
            )
    return cells


def _cell_text(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _sheet_markdown(sheet_title: str, cells: list[ParsedSourceCell]) -> list[str]:
    lines = [f"## Sheet: {sheet_title}"]
    if not cells:
        lines.append("- 无文本单元格")
        return lines
    for cell in cells:
        lines.append(f"- {cell.coord}: {cell.text}")
    return lines


def _extract_sheet_images(
    *,
    project_id: int,
    run_id: int,
    sheet: Any,
    sheet_number: int,
) -> list[ParsedSourceResource]:
    resources: list[ParsedSourceResource] = []
    for image_index, image in enumerate(getattr(sheet, "_images", []) or [], start=1):
        ref = f"excel_img_s{sheet_number:03d}_{image_index:03d}"
        anchor = _image_anchor_cell(image)
        extension = _image_extension(image)
        mime_type = _image_mime_type(extension)
        filename = f"{ref}{extension}"
        relative_path = f"images/{filename}"
        image_bytes = _image_bytes(image)
        source_evidence_storage.write_source_evidence_bytes(
            project_id,
            run_id,
            relative_path,
            image_bytes,
        )
        resources.append(
            ParsedSourceResource(
                ref=ref,
                type="image",
                source_id=ref,
                position=f"excel:sheet={sheet.title}:image={image_index}:anchor={anchor}",
                filename=filename,
                file_token=ref,
                mime_type=mime_type,
                status="downloaded",
                metadata={
                    "local_path": relative_path,
                    "sheet": sheet.title,
                    "sheet_index": sheet_number,
                    "image_index": image_index,
                    "anchor": anchor,
                },
            )
        )
    return resources


def _image_bytes(image: Any) -> bytes:
    data = image._data() if hasattr(image, "_data") else b""
    if not data:
        raise ValueError("读取本地上传文件失败：无法提取 .xlsx 内嵌图片。")
    return data


def _image_anchor_cell(image: Any) -> str:
    marker = getattr(getattr(image, "anchor", None), "_from", None)
    if marker is None:
        return "unknown"
    row = int(getattr(marker, "row", 0)) + 1
    col = int(getattr(marker, "col", 0)) + 1
    return f"{get_column_letter(col)}{row}"


def _image_extension(image: Any) -> str:
    path = str(getattr(image, "path", "") or "").lower()
    suffix = Path(path).suffix
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return ".jpg" if suffix == ".jpeg" else suffix
    image_format = str(getattr(image, "format", "") or "").lower().strip(".")
    if image_format in {"png", "jpg", "jpeg", "webp"}:
        return ".jpg" if image_format == "jpeg" else f".{image_format}"
    return ".png"


def _image_mime_type(extension: str) -> str:
    if extension == ".jpg":
        return "image/jpeg"
    if extension == ".webp":
        return "image/webp"
    return "image/png"
