"""用例生成 V1 Excel 导出。"""

from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.models import (
    TestCaseCoverageAuditRecord,
    TestCaseGenerationCaseRecord,
    TestCaseGenerationChunkRecord,
    TestCaseReferenceFileRecord,
    TestCaseRequirementAtomRecord,
)
from backend.app.test_cases.constants import (
    STANDARD_CASE_FIELD_LABELS,
    STANDARD_CASE_FIELDS,
)
from backend.app.test_cases.generation_runs import (
    EXPIRED_DETAILS_MESSAGE,
    GENERATION_RUN_COMPLETED_STATUSES,
    GenerationRunError,
    NO_EXPORT_MESSAGE,
    RUN_NOT_EXPORTABLE_MESSAGE,
    STRICT_EXPORT_COVERAGE_GAP_MESSAGE,
    get_project_generation_run,
)
from backend.app.test_cases.schemas import (
    GeneratedTestCase,
    TestCaseBlueprint,
    TestCaseExportRequest,
)


TEST_CASE_EXPORT_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_SENSITIVE_PROFILE_KEYS = {
    "api_key",
    "apikey",
    "authorization",
    "encrypted_api_key",
    "raw_prompt",
    "prompt",
    "system_prompt",
    "user_prompt",
    "provider_response",
    "raw_provider_response",
    "raw_response",
    "source_token",
    "file_token",
    "token",
    "local_path",
    "observation_path",
    "password",
    "secret",
}

_WINDOWS_PATH_RE = re.compile(r"[A-Za-z]:[\\/][^\s\"']+")
_UNIX_PATH_RE = re.compile(r"(?<!\w)/(?:[^\s\"']+/){1,}[^\s\"']*")
_URL_RE = re.compile(r"https?://[^\s\"']+", re.IGNORECASE)


def build_test_case_export_workbook(payload: TestCaseExportRequest) -> BytesIO:
    """基于当前页面提交的结果构建用例生成导出工作簿。"""
    workbook = Workbook()
    case_sheet = workbook.active
    case_sheet.title = "测试用例"
    blueprint_sheet = workbook.create_sheet("用例蓝图")
    summary_sheet = workbook.create_sheet("生成说明")

    export_fields = resolve_export_fields(payload)
    _write_case_sheet(case_sheet, payload.cases, export_fields)
    _write_blueprint_sheet(blueprint_sheet, payload.blueprint)
    _write_summary_sheet(summary_sheet, payload, export_fields)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


async def build_generation_run_export_workbook(
    db: AsyncSession,
    *,
    project_id: int,
    run_id: int,
) -> BytesIO:
    """从 V3 Generation Run 的短期 DB 结果构建导出工作簿。"""
    run = await get_project_generation_run(db, project_id=project_id, run_id=run_id)
    if run.status == "expired":
        raise GenerationRunError(409, EXPIRED_DETAILS_MESSAGE)
    if run.status in {"cancelled", "failed"}:
        raise GenerationRunError(409, RUN_NOT_EXPORTABLE_MESSAGE)
    if run.status not in GENERATION_RUN_COMPLETED_STATUSES:
        raise GenerationRunError(409, NO_EXPORT_MESSAGE)

    audit = await _load_generation_run_audit(db, project_id=project_id, run_id=run.id)
    if run.strict_mode and audit is not None and audit.uncovered_atoms > 0:
        raise GenerationRunError(409, STRICT_EXPORT_COVERAGE_GAP_MESSAGE)

    cases = await _load_generation_run_cases(db, project_id=project_id, run_id=run.id)
    if not cases:
        raise GenerationRunError(409, NO_EXPORT_MESSAGE)
    atoms = await _load_generation_run_atoms(db, project_id=project_id, run_id=run.id)
    chunks = await _load_generation_run_chunks(db, project_id=project_id, run_id=run.id)

    workbook = Workbook()
    case_sheet = workbook.active
    case_sheet.title = "测试用例"
    blueprint_sheet = workbook.create_sheet("用例蓝图")
    summary_sheet = workbook.create_sheet("生成说明")
    coverage_sheet = workbook.create_sheet("覆盖审计")

    export_fields = await _resolve_generation_run_export_fields(
        db,
        project_id=project_id,
        primary_reference_id=run.primary_reference_id,
        primary_reference_sheet_name=run.primary_reference_sheet_name,
    )
    generated_cases = [_generated_case_from_record(record) for record in cases]
    _write_case_sheet(case_sheet, generated_cases, export_fields)
    _write_blueprint_sheet(
        blueprint_sheet,
        _blueprint_from_stage_payload(run.stage_payload_json),
    )
    _write_generation_run_summary_sheet(
        summary_sheet,
        run=run,
        audit=audit,
        export_fields=export_fields,
    )
    _write_coverage_audit_sheet(
        coverage_sheet,
        atoms=atoms,
        cases=cases,
        chunks=chunks,
        audit=audit,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def resolve_export_fields(payload: TestCaseExportRequest) -> list[str]:
    """解析导出字段顺序：主参考可映射字段优先，标准字段兜底。"""
    return resolve_export_fields_from_reference_profile(
        payload.primary_reference_profile,
        fallback_fields=payload.export_columns,
    )


def resolve_export_fields_from_reference_profile(
    profile: dict[str, Any] | None,
    *,
    fallback_fields: list[str] | None = None,
) -> list[str]:
    """按主参考画像可识别字段生成导出字段顺序，缺失标准字段后置补齐。"""
    recognized_fields = _extract_profile_standard_fields(profile)
    if not recognized_fields and fallback_fields:
        recognized_fields = [field for field in fallback_fields if field in STANDARD_CASE_FIELDS]
    if not recognized_fields:
        recognized_fields = list(STANDARD_CASE_FIELDS)

    ordered_fields = _unique_preserve_order(
        [field for field in recognized_fields if field in STANDARD_CASE_FIELDS]
    )
    for field in STANDARD_CASE_FIELDS:
        if field not in ordered_fields:
            ordered_fields.append(field)
    return ordered_fields


def _write_case_sheet(
    sheet: Any,
    cases: list[GeneratedTestCase],
    export_fields: list[str],
) -> None:
    sheet.append([STANDARD_CASE_FIELD_LABELS[field] for field in export_fields])
    for case in cases:
        data = case.model_dump(mode="json")
        sheet.append([_format_cell_value(data.get(field, "")) for field in export_fields])

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _write_blueprint_sheet(sheet: Any, blueprint: TestCaseBlueprint) -> None:
    sheet.append(["区块", "内容"])
    rows = [
        ("模块树", blueprint.modules),
        ("核心流程", blueprint.flows),
        ("需求追踪", blueprint.requirement_traces),
        ("覆盖维度", blueprint.coverage_dimensions),
        ("风险点", blueprint.risks),
        ("未映射需求", blueprint.unmapped_requirements),
        ("无依据测试点", blueprint.unsupported_or_unfounded_test_points),
        ("待确认问题", blueprint.open_questions),
        ("蓝图 warnings", [warning.model_dump(mode="json") for warning in blueprint.warnings]),
    ]
    for title, value in rows:
        sheet.append([title, _format_cell_value(value)])

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _write_summary_sheet(
    sheet: Any,
    payload: TestCaseExportRequest,
    export_fields: list[str],
) -> None:
    sheet.append(["项目", "类型", "内容"])
    rows = [
        ("生成时间", "meta", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("策划案来源", "source", payload.source_summary),
        ("Source Evidence 摘要", "source_evidence", _resolve_source_evidence_summary(payload)),
        ("导出字段", "fields", "、".join(STANDARD_CASE_FIELD_LABELS[field] for field in export_fields)),
        ("用例总数", "stats", payload.stats.total),
        ("优先级分布", "stats", payload.stats.priority_counts),
        ("模块分布", "stats", payload.stats.module_counts),
        ("用例类型分布", "stats", payload.stats.case_type_counts),
        ("warning 数", "stats", payload.stats.warning_count),
        ("主参考画像", "reference", _summarize_primary_reference_profile(payload.primary_reference_profile)),
        (
            "V1 限制",
            "limit",
            "导出完全基于当前页面提交的 blueprint/cases/warnings/stats；"
            "不读取生成历史，不写入完整 API Key、敏感请求内容或上游原始响应。",
        ),
    ]
    for row in rows:
        if row[0] == "Source Evidence 摘要" and not row[2]:
            continue
        sheet.append([_format_cell_value(item) for item in row])
    for warning in payload.warnings:
        sheet.append(["warning", warning.level, warning.message])

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


async def _load_generation_run_audit(
    db: AsyncSession,
    *,
    project_id: int,
    run_id: int,
) -> TestCaseCoverageAuditRecord | None:
    return (
        await db.execute(
            select(TestCaseCoverageAuditRecord).where(
                TestCaseCoverageAuditRecord.project_id == project_id,
                TestCaseCoverageAuditRecord.run_id == run_id,
            )
        )
    ).scalar_one_or_none()


async def _load_generation_run_cases(
    db: AsyncSession,
    *,
    project_id: int,
    run_id: int,
) -> list[TestCaseGenerationCaseRecord]:
    return list(
        (
            await db.execute(
                select(TestCaseGenerationCaseRecord)
                .where(
                    TestCaseGenerationCaseRecord.project_id == project_id,
                    TestCaseGenerationCaseRecord.run_id == run_id,
                    TestCaseGenerationCaseRecord.status == "official",
                )
                .order_by(TestCaseGenerationCaseRecord.id)
            )
        ).scalars()
    )


async def _load_generation_run_atoms(
    db: AsyncSession,
    *,
    project_id: int,
    run_id: int,
) -> list[TestCaseRequirementAtomRecord]:
    return list(
        (
            await db.execute(
                select(TestCaseRequirementAtomRecord)
                .where(
                    TestCaseRequirementAtomRecord.project_id == project_id,
                    TestCaseRequirementAtomRecord.run_id == run_id,
                    TestCaseRequirementAtomRecord.coverage_status != "unfounded_candidate",
                )
                .order_by(TestCaseRequirementAtomRecord.id)
            )
        ).scalars()
    )


async def _load_generation_run_chunks(
    db: AsyncSession,
    *,
    project_id: int,
    run_id: int,
) -> list[TestCaseGenerationChunkRecord]:
    return list(
        (
            await db.execute(
                select(TestCaseGenerationChunkRecord)
                .where(
                    TestCaseGenerationChunkRecord.project_id == project_id,
                    TestCaseGenerationChunkRecord.run_id == run_id,
                )
                .order_by(TestCaseGenerationChunkRecord.chunk_index)
            )
        ).scalars()
    )


async def _resolve_generation_run_export_fields(
    db: AsyncSession,
    *,
    project_id: int,
    primary_reference_id: int | None,
    primary_reference_sheet_name: str | None,
) -> list[str]:
    profile: dict[str, Any] | None = None
    if primary_reference_id is not None:
        reference = await db.get(TestCaseReferenceFileRecord, primary_reference_id)
        if reference is not None and reference.project_id == project_id:
            profile = _json_object(reference.profile_json)
            if primary_reference_sheet_name:
                profile = {
                    **profile,
                    "selected_sheet_name": primary_reference_sheet_name,
                }
    return resolve_export_fields_from_reference_profile(profile)


def _generated_case_from_record(record: TestCaseGenerationCaseRecord) -> GeneratedTestCase:
    fields = _json_object(record.fields_json)
    normalized = {
        field: fields.get(field, "")
        for field in STANDARD_CASE_FIELDS
    }
    normalized["case_id"] = record.case_id
    if not normalized.get("priority"):
        normalized["priority"] = "P2"
    if not normalized.get("initial_status"):
        normalized["initial_status"] = "未执行"
    return GeneratedTestCase.model_validate(normalized)


def _blueprint_from_stage_payload(stage_payload_json: str) -> TestCaseBlueprint:
    stage_payload = _json_object(stage_payload_json)
    blueprint = stage_payload.get("blueprint")
    if not isinstance(blueprint, dict):
        return TestCaseBlueprint()
    return TestCaseBlueprint.model_validate(_sanitize_json_value(blueprint))


def _write_generation_run_summary_sheet(
    sheet: Any,
    *,
    run: Any,
    audit: TestCaseCoverageAuditRecord | None,
    export_fields: list[str],
) -> None:
    sheet.append(["项目", "类型", "内容"])
    rows: list[tuple[Any, Any, Any]] = [
        ("导出时间", "meta", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ("Generation Run ID", "run", run.id),
        ("运行状态", "run", run.status),
        ("Planning Sheet", "source", run.planning_sheet_name),
        ("Source Evidence Run ID", "source", run.source_evidence_run_id),
        ("严格模式", "run", "是" if run.strict_mode else "否"),
        ("导出字段", "fields", "、".join(STANDARD_CASE_FIELD_LABELS[field] for field in export_fields)),
        ("用例总数", "stats", run.case_count),
        ("Requirement Atom 总数", "stats", run.atom_count),
        ("chunk 总数", "stats", run.total_chunks),
        ("失败 chunk 数", "stats", run.failed_chunks),
        ("warning 数", "stats", run.warning_count),
    ]
    if audit is not None:
        rows.extend(
            [
                ("Coverage Audit 状态", "audit", audit.status),
                ("覆盖 atoms", "audit", audit.covered_atoms),
                ("未覆盖 atoms", "audit", audit.uncovered_atoms),
                ("无依据候选用例数", "audit", audit.unfounded_case_count),
                ("补生成摘要", "audit", _json_object(audit.supplement_summary_json)),
            ]
        )
    rows.append(
        (
            "V3 导出限制",
            "limit",
            "导出完全基于 Generation Run 数据库短期结果；不接受前端回传 blueprint/cases/stats 作为事实。",
        )
    )
    for row in rows:
        sheet.append([_format_cell_value(item) for item in row])

    for warning in _generation_run_export_warnings(run=run, audit=audit):
        sheet.append(
            [
                "warning",
                _format_cell_value(warning.get("level", "warning")),
                _format_cell_value(warning.get("message", "")),
            ]
        )

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _write_coverage_audit_sheet(
    sheet: Any,
    *,
    atoms: list[TestCaseRequirementAtomRecord],
    cases: list[TestCaseGenerationCaseRecord],
    chunks: list[TestCaseGenerationChunkRecord],
    audit: TestCaseCoverageAuditRecord | None,
) -> None:
    sheet.append(
        [
            "atom id",
            "source sheet",
            "source rows",
            "source columns",
            "atom type",
            "atom text",
            "coverage status",
            "linked case ids",
            "failed chunk",
            "limitation notes",
        ]
    )
    linked_case_ids = _linked_case_ids_by_atom(cases)
    uncovered_atom_ids = set(_json_list(audit.uncovered_atom_ids_json if audit else "[]"))
    limitation_notes = _limitation_notes_by_atom(audit)
    failed_chunk_notes = _failed_chunk_notes_by_atom(chunks, atoms)
    for atom in atoms:
        atom_id = atom.atom_id
        linked = linked_case_ids.get(atom_id, [])
        if linked:
            coverage_status = "covered"
        elif atom_id in uncovered_atom_ids:
            coverage_status = "uncovered"
        else:
            coverage_status = atom.coverage_status or "unmapped"
        notes = [
            *limitation_notes.get(atom_id, []),
            *[str(item) for item in _json_list(atom.warnings_json)],
        ]
        sheet.append(
            [
                _format_cell_value(atom_id),
                _format_cell_value(atom.source_sheet_name),
                _format_cell_value(_source_rows_text(atom)),
                _format_cell_value("、".join(str(item) for item in _json_list(atom.source_columns_json))),
                _format_cell_value(atom.atom_type),
                _format_cell_value(atom.requirement_text),
                _format_cell_value(coverage_status),
                _format_cell_value("、".join(linked)),
                _format_cell_value("、".join(failed_chunk_notes.get(atom_id, []))),
                _format_cell_value("；".join(notes)),
            ]
        )

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _generation_run_export_warnings(
    *,
    run: Any,
    audit: TestCaseCoverageAuditRecord | None,
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    warnings.extend(item for item in _json_list(run.warnings_json) if isinstance(item, dict))
    if audit is not None:
        warnings.extend(
            item for item in _json_list(audit.warnings_json) if isinstance(item, dict)
        )
        for limitation in _json_list(audit.export_limitations_json):
            if isinstance(limitation, dict):
                warnings.append(
                    {
                        "source": "coverage",
                        "level": limitation.get("level", "warning"),
                        "message": limitation.get("message", ""),
                    }
                )
    return _dedupe_warnings(warnings)


def _linked_case_ids_by_atom(
    cases: list[TestCaseGenerationCaseRecord],
) -> dict[str, list[str]]:
    linked: dict[str, list[str]] = {}
    for case in cases:
        for atom_id in _json_list(case.atom_refs_json):
            linked.setdefault(str(atom_id), []).append(case.case_id)
    return {key: _unique_preserve_order(value) for key, value in linked.items()}


def _limitation_notes_by_atom(
    audit: TestCaseCoverageAuditRecord | None,
) -> dict[str, list[str]]:
    notes: dict[str, list[str]] = {}
    if audit is None:
        return notes
    for limitation in _json_list(audit.export_limitations_json):
        if not isinstance(limitation, dict):
            continue
        message = _sanitize_sensitive_text(str(limitation.get("message") or ""))
        for atom_id in _json_list(limitation.get("atom_ids")):
            notes.setdefault(str(atom_id), []).append(message)
    return notes


def _failed_chunk_notes_by_atom(
    chunks: list[TestCaseGenerationChunkRecord],
    atoms: list[TestCaseRequirementAtomRecord],
) -> dict[str, list[str]]:
    failed_chunks = [chunk for chunk in chunks if chunk.status == "failed"]
    notes: dict[str, list[str]] = {}
    if not failed_chunks:
        return notes
    for atom in atoms:
        for chunk in failed_chunks:
            if _atom_overlaps_chunk(atom, chunk):
                notes.setdefault(atom.atom_id, []).append(f"chunk-{chunk.chunk_index}")
    return notes


def _atom_overlaps_chunk(
    atom: TestCaseRequirementAtomRecord,
    chunk: TestCaseGenerationChunkRecord,
) -> bool:
    if atom.chunk_id is not None and atom.chunk_id == chunk.id:
        return True
    if (
        atom.source_row_start is None
        or atom.source_row_end is None
        or chunk.source_row_start is None
        or chunk.source_row_end is None
    ):
        return False
    return atom.source_row_start <= chunk.source_row_end and atom.source_row_end >= chunk.source_row_start


def _source_rows_text(atom: TestCaseRequirementAtomRecord) -> str:
    if atom.source_row_start is None:
        return ""
    if atom.source_row_end is None or atom.source_row_end == atom.source_row_start:
        return str(atom.source_row_start)
    return f"{atom.source_row_start}-{atom.source_row_end}"


def _dedupe_warnings(values: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    result: list[dict[str, Any]] = []
    for value in values:
        level = _sanitize_sensitive_text(str(value.get("level") or "warning"))
        message = _sanitize_sensitive_text(str(value.get("message") or ""))
        if not message:
            continue
        key = (level, message)
        if key in seen:
            continue
        seen.add(key)
        result.append({"level": level, "message": message})
    return result


def _extract_profile_standard_fields(profile: dict[str, Any] | None) -> list[str]:
    if not isinstance(profile, dict):
        return []

    fields: list[str] = []
    for item in _iter_profile_columns(profile):
        if isinstance(item, str):
            standard_field = item
        elif isinstance(item, dict):
            standard_field = _first_text(
                item,
                "standard_field",
                "mapped_standard_field",
                "mapped_field",
                "standard_key",
                "field",
            )
        else:
            continue

        if standard_field in STANDARD_CASE_FIELDS:
            fields.append(standard_field)
    return _unique_preserve_order(fields)


def _iter_profile_columns(profile: dict[str, Any]) -> list[Any]:
    selected_sheet_name = _first_text(
        profile,
        "selected_sheet_name",
        "sheet_name",
        "primary_reference_sheet_name",
    )
    default_sheet_name = _first_text(profile, "default_sheet_name")
    for sheet_key in ("sheet_options", "sheets"):
        sheets = profile.get(sheet_key)
        if not isinstance(sheets, list):
            continue
        if selected_sheet_name:
            selected_columns = _find_sheet_columns(sheets, selected_sheet_name)
            if selected_columns:
                return selected_columns
        if default_sheet_name:
            default_columns = _find_sheet_columns(sheets, default_sheet_name)
            if default_columns:
                return default_columns

    for key in ("columns", "fields", "ordered_columns", "field_order"):
        value = profile.get(key)
        if isinstance(value, list):
            return value

    for key in ("selected_sheet_profile", "sheet_profile", "primary_sheet_profile"):
        nested = profile.get(key)
        if isinstance(nested, dict):
            nested_columns = _iter_profile_columns(nested)
            if nested_columns:
                return nested_columns

    return []


def _find_sheet_columns(sheets: list[Any], sheet_name: str) -> list[Any]:
    for item in sheets:
        if not isinstance(item, dict):
            continue
        if _first_text(item, "name", "sheet_name") != sheet_name:
            continue
        columns = item.get("columns")
        if isinstance(columns, list):
            return columns
    return []


def _summarize_primary_reference_profile(profile: dict[str, Any] | None) -> str:
    if not isinstance(profile, dict):
        return "未使用主参考"

    safe_summary: dict[str, Any] = {}
    for key in ("name", "file_name", "filename", "sheet_name", "default_sheet_name"):
        value = profile.get(key)
        if value is not None:
            safe_summary[key] = value
    mapped_fields = _extract_profile_standard_fields(profile)
    if mapped_fields:
        safe_summary["mapped_standard_fields"] = mapped_fields
    return _format_cell_value(safe_summary or "已使用主参考画像")


def _first_text(item: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _style_header(sheet: Any, row_index: int) -> None:
    fill = PatternFill("solid", fgColor="EAF2FF")
    font = Font(bold=True, color="1F2A44")
    for cell in sheet[row_index]:
        cell.fill = fill
        cell.font = font


def _auto_fit_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            60,
        )


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(_sanitize_json_value(value), ensure_ascii=False)
    return _sanitize_sensitive_text(str(value))


def _resolve_source_evidence_summary(payload: TestCaseExportRequest) -> str:
    summary = payload.source_evidence_summary or payload.evidence_summary
    return _sanitize_sensitive_text(summary)


def _sanitize_sensitive_text(value: str) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = _URL_RE.sub("[url]", text)
    text = _WINDOWS_PATH_RE.sub("[path]", text)
    text = _UNIX_PATH_RE.sub("[path]", text)
    text = re.sub(
        r"(?i)\bauthorization\s*:\s*bearer\s*[^；;\n]+",
        "[REDACTED]",
        text,
    )
    text = re.sub(
        r"(?i)\b(api_key|apikey|prompt|raw_prompt|provider_response|raw_provider_response|raw_response|token|secret|password)\s*[:=]\s*[^；;\n]+",
        "[REDACTED]",
        text,
    )
    text = re.sub(r"sk-[A-Za-z0-9._-]+", "sk-[REDACTED]", text)
    for term in (
        "provider_response",
        "raw_provider_response",
        "raw_response",
        "raw prompt",
        "prompt",
        "observation detail",
    ):
        text = re.sub(re.escape(term), "[REDACTED]", text, flags=re.IGNORECASE)
    return text


def _sanitize_json_value(value: Any) -> Any:
    if isinstance(value, dict):
        sanitized: dict[str, Any] = {}
        for key, child in value.items():
            normalized_key = str(key).strip().lower()
            if normalized_key in _SENSITIVE_PROFILE_KEYS:
                continue
            sanitized[str(key)] = _sanitize_json_value(child)
        return sanitized
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    return value


def _json_value(value: Any, default: Any) -> Any:
    if isinstance(value, type(default)):
        return value
    if not value:
        return default
    if isinstance(value, str):
        try:
            loaded = json.loads(value)
        except (TypeError, ValueError):
            return default
        return loaded if isinstance(loaded, type(default)) else default
    return default


def _json_list(value: Any) -> list[Any]:
    loaded = _json_value(value, [])
    return loaded if isinstance(loaded, list) else []


def _json_object(value: Any) -> dict[str, Any]:
    loaded = _json_value(value, {})
    return loaded if isinstance(loaded, dict) else {}


def _unique_preserve_order(items: list[str]) -> list[str]:
    return list(dict.fromkeys(items))
